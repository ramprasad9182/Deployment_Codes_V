import re

from odoo import models, fields, _
from odoo.exceptions import ValidationError
import base64
from io import BytesIO
from openpyxl import load_workbook


class DeliveryImportWizard(models.TransientModel):
    _name = 'delivery.import.wizard'
    _description = "Delivery Import Wizard"

    picking_id = fields.Many2one('stock.picking', string="Delivery", required=True)
    file = fields.Binary(string="Excel File", required=True)
    filename = fields.Char(string="Filename")

    def action_import_barcodes(self):
        """
        Excel Format:
        -------------------------
        Barcode         Qty
        -------------------------
        ABC123          4
        RCODE001        7

        SERIAL PRODUCT:
        -------------------------
        If barcode has serials:
            R1,R2,R3,R4,R5

        Qty = 4
        -> Takes:
            R1,R2,R3,R4

        LOT PRODUCT:
        -------------------------
        Lot1 = 5 qty
        Lot2 = 6 qty

        Qty = 7
        -> Takes:
            Lot1 = 5
            Lot2 = 2
        """

        self.ensure_one()

        if not self.file:
            raise ValidationError(_("Please upload Excel file."))

        try:
            data = base64.b64decode(self.file)
            wb = load_workbook(filename=BytesIO(data), data_only=True)
            sheet = wb.active
        except Exception as e:
            raise ValidationError(_("Invalid file: %s") % str(e))

        StockQuant = self.env['stock.quant']
        StockMove = self.env['stock.move']

        location_id = self.env.ref('stock.stock_location_stock').id

        skipped_lines = []

        # =========================================================
        # EXISTING LINE
        # =========================================================
        def _find_existing_line(product, lot_id):
            return StockMove.search([
                ('picking_id', '=', self.picking_id.id),
                ('product_id', '=', product.id),
                ('dummy_lot_ids', 'in', [lot_id]),
                ('state', '!=', 'cancel')
            ], limit=1)

        # =========================================================
        # SERIAL AVAILABILITY
        # =========================================================
        def _serial_already_used(lot_id):

            move = StockMove.search([
                ('dummy_lot_ids', 'in', [lot_id]),
                ('state', '!=', 'cancel')
            ], limit=1)

            return bool(move)

        # =========================================================
        # LOOP EXCEL
        # =========================================================
        for row in sheet.iter_rows(min_row=2, values_only=True):

            barcode = str(row[0]).strip() if row[0] else False
            qty = float(row[1]) if row[1] else 0

            if not barcode or qty <= 0:
                skipped_lines.append((barcode, "Missing barcode or qty"))
                continue

            # =====================================================
            # FIND STOCK
            # =====================================================
            quants = StockQuant.search([
                ('lot_id.ref', '=', barcode),
                ('quantity', '>', 0),
                ('location_id', '=', location_id),
            ], order='lot_id asc')

            if not quants:
                quants = StockQuant.search([
                    ('lot_id.name', '=', barcode),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                ], order='lot_id asc')

            if not quants:
                skipped_lines.append((barcode, "No stock found"))
                continue

            product = quants[0].product_id

            # =====================================================
            # SERIAL TRACKING
            # =====================================================
            if product.tracking == 'serial':

                available_serials = []

                for q in quants:

                    lot = q.lot_id

                    if not lot:
                        continue

                    # skip used serial
                    if _serial_already_used(lot.id):
                        continue

                    available_serials.append(q)

                if qty > len(available_serials):
                    skipped_lines.append((
                        barcode,
                        f"Only {len(available_serials)} serials available"
                    ))
                    continue

                created_count = 0

                # =============================================
                # FIFO SERIAL ALLOCATION
                # =============================================
                for q in available_serials:

                    lot = q.lot_id

                    StockMove.create({
                        'picking_id': self.picking_id.id,
                        'product_id': product.id,
                        'dummy_lot_ids': [(6, 0, [lot.id])],
                        'name': product.display_name,
                        'product_uom_qty': 1,
                        'product_uom': product.uom_id.id,
                        'location_id': self.picking_id.location_id.id,
                        'location_dest_id': self.picking_id.location_dest_id.id,
                    })

                    created_count += 1

                    if created_count >= qty:
                        break

            # =====================================================
            # LOT TRACKING
            # =====================================================
            else:

                available_quants = quants.filtered(
                    lambda q: q.quantity > 0
                )

                total_available = sum(
                    available_quants.mapped('quantity')
                )

                if qty > total_available:
                    skipped_lines.append((
                        barcode,
                        f"Requested {qty} but only {total_available} available"
                    ))
                    continue

                remaining_qty = qty

                # =============================================
                # FIFO LOT ALLOCATION
                # =============================================
                for q in available_quants:

                    if remaining_qty <= 0:
                        break

                    lot = q.lot_id

                    available_qty = q.quantity

                    if available_qty <= 0:
                        continue

                    allocate_qty = min(
                        remaining_qty,
                        available_qty
                    )

                    existing_line = _find_existing_line(
                        product,
                        lot.id
                    )

                    # =========================================
                    # UPDATE EXISTING
                    # =========================================
                    if existing_line:

                        existing_line.write({
                            'product_uom_qty':
                                existing_line.product_uom_qty + allocate_qty
                        })

                    # =========================================
                    # CREATE NEW
                    # =========================================
                    else:

                        StockMove.create({
                            'picking_id': self.picking_id.id,
                            'product_id': product.id,
                            'dummy_lot_ids': [(6, 0, [lot.id])],
                            'name': product.display_name,
                            'product_uom_qty': allocate_qty,
                            'product_uom': product.uom_id.id,
                            'location_id': self.picking_id.location_id.id,
                            'location_dest_id': self.picking_id.location_dest_id.id,
                        })

                    remaining_qty -= allocate_qty

                if remaining_qty > 0:
                    skipped_lines.append((
                        barcode,
                        f"Short by {remaining_qty}"
                    ))

        # =========================================================
        # FINAL ERROR
        # =========================================================
        if skipped_lines:
            msg = "\n".join([
                f"{b or '<empty>'} : {reason}"
                for b, reason in skipped_lines
            ])

            raise ValidationError(_(
                "Some lines skipped:\n%s"
            ) % msg)
