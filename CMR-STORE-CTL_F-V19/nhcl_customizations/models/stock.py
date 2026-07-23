import base64
import io
import logging
import ast
import openpyxl
from openpyxl import load_workbook
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, UserError
import math
import re
import requests
from collections import defaultdict
import csv
from datetime import datetime, timedelta
import xlrd
import pytz

_logger = logging.getLogger(__name__)
from markupsafe import Markup


class Picking(models.Model):
    """Inherited stock.picking class to add fields and functions"""
    _inherit = "stock.picking"

    stock_picking_delivery_ids = fields.One2many("stock.picking.barcode", "stock_picking_delivery_id")
    nhcl_batch_number = fields.Char(string="Batch Number")
    allow_slip_printed = fields.Boolean(string="Allow Slip Printed", default=False)
    customer_phone = fields.Char(string="Customer Phone", copy=False)
    date_time_nh = fields.Datetime(string="Date Time", copy=False)
    # allow_slip_printed = fields.Boolean(string="Allow Slip Printed", default=False)
    allow_slip_done = fields.Boolean(string="Allow Slip Done", default=False)
    return_counter = fields.Many2one('pos.config', string='Return Counter')
    is_outgoing = fields.Boolean(string="Is Outgoing", compute="_compute_is_outgoing")

    total_quantity = fields.Float(
        string="Quantity",
        compute="_compute_totals",
        store=True
    )

    net_amount = fields.Float(
        string="Net Amount",
        compute="_compute_totals",
        store=True
    )
    gst_types = fields.Char(
        string="Term Name",
        compute="_compute_gst_types",
        store=True
    )
    hired_product_ids = fields.One2many(
        'sale.order.hired.product',
        'picking_id',
        string="Hired Products"
    )
    transfer_type = fields.Selection(
        [('regular', 'Regular'), ('damage', 'Damage'), ('stock_mismatch', 'Stock Mismatch'),
         ('hired', 'Hired Product')],
        string="Transfer Type", tracking=True, copy=False)
    return_scan_barcode = fields.Char(string="Scan Serial")
    nhcl_integration_count = fields.Selection([('no','No'),('yes','Yes')], default='no', copy=False)

    @api.depends('picking_type_id')
    def _compute_is_outgoing(self):
        for picking in self:
            if picking.picking_type_id.code == 'outgoing':
                picking.is_outgoing = picking.picking_type_id.code == 'outgoing'
            elif picking.state in ['draft', 'done', 'cancel'] and picking.stock_picking_type not in ['pos_order']:
                picking.is_outgoing = True
            else:
                picking.is_outgoing = False


    @api.constrains('no_of_parcel', 'stock_picking_type')
    def _check_no_of_parcel(self):
        for rec in self:
            if rec.stock_picking_type in ('regular', 'damage') and (not rec.no_of_parcel or rec.no_of_parcel <= 0) and rec.state in ('done'):
                raise ValidationError(
                    _("No Of Parcels must be greater than 0.")
                )

    @api.onchange('return_scan_barcode')
    def _onchange_return_scan_barcode(self):
        if not self.return_scan_barcode:
            return

        serial = self.return_scan_barcode

        line = self.hired_product_ids.filtered(
            lambda l: l.lot_number.name == serial
        )

        if not line:
            raise ValidationError("Serial not expected in this return.")

        if line.returned_scan:
            raise ValidationError("Serial already scanned.")

        line.returned_scan = True

        self.return_scan_barcode = False

    @api.depends(
        'move_ids_without_package.purchase_line_id.taxes_id',
        'move_ids_without_package.sale_line_id.tax_id'
    )
    def _compute_gst_types(self):
        for picking in self:

            purchase_taxes = picking.move_ids_without_package.mapped(
                'purchase_line_id.taxes_id'
            )

            sale_taxes = picking.move_ids_without_package.mapped(
                'sale_line_id.tax_id'
            )

            taxes = purchase_taxes | sale_taxes

            gst_set = set()

            for tax in taxes:
                children = tax.children_tax_ids if tax.amount_type == 'group' else tax

                for child in children:
                    name = child.name.upper()

                    if "CGST" in name:
                        gst_set.add("CGST")
                    elif "SGST" in name:
                        gst_set.add("SGST")
                    elif "IGST" in name:
                        gst_set.add("IGST")

            picking.gst_types = (
                "GST: " + ", ".join(sorted(gst_set))
                if gst_set else ""
            )

    # @api.depends('move_line_ids.quantity', 'move_line_ids.rs_price')
    # def _compute_totals(self):
    #     for picking in self:
    #         qty = 0
    #         amount = 0
    #
    #         for line in picking.move_line_ids:
    #             if line.product_id.detailed_type == 'product':
    #                 qty += line.quantity
    #                 amount += line.rs_price
    #
    #         picking.total_quantity = qty
    #         picking.net_amount = amount

    @api.depends('move_ids_without_package.quantity', 'move_ids_without_package.nhcl_price_total',
                 'move_ids_without_package.nhcl_price_tax')
    def _compute_totals(self):
        for picking in self:
            qty = 0
            amount = 0

            for line in picking.move_ids_without_package:
                # if line.nhcl_exchange:
                if line.quantity and line.product_id.detailed_type == 'product':
                    qty += line.quantity
                # amount += (line.nhcl_price_total - line.nhcl_price_tax)
                amount += line.nhcl_price_total

            picking.total_quantity = qty
            picking.net_amount = amount


    @api.model
    def default_get(self, fields):
        res = super().default_get(fields)
        if self.env.context.get('default_stock_picking_type'):
            picking_type = self.env['stock.picking.type'].search([
                ('stock_picking_type', '=', self.env.context['default_stock_picking_type'])], limit=1)
            counter = self.env['pos.config'].search([('name','=','RF Counter')])
            if picking_type:
                res['picking_type_id'] = picking_type.id
                if picking_type == 'exchange' and not counter:
                    raise ValidationError("RF counter is not available please create the shop.")
                elif picking_type.stock_picking_type == 'exchange' and counter:
                    res['return_counter'] = counter.id
        return res

    def button_import_receipts(self):
        return {
            "name": _("Import Receipts"),
            "type": "ir.actions.act_window",
            "res_model": "import.stock.receipt.wizard",
            "target": "new",
            "views": [[False, "form"]],
        }

    @api.model
    def _cron_notify_old_backorders(self):
        """Notify about incoming backorders older than 1 month."""

        # Find backorders older than 30 days
        one_month_ago = datetime.now() - timedelta(days=30)
        old_backorders = self.search([
            ('stock_picking_type', '=', 'receipt'),
            ('backorder_id', '!=', False),
            ('create_date', '<', one_month_ago),
            ('state', 'in', ['draft', 'assigned']),
        ])

        if not old_backorders:
            return  # nothing to notify

        # Find the group to notify
        group = self.env.ref('nhcl_customizations.group_backorder_notification', raise_if_not_found=False)
        if not group:
            return  # no group exists, stop

        # Collect partner IDs (simple safe list)
        group_partner_ids = group.users.mapped('partner_id.id')

        # Find or create the channel (Important: create without partners first!)
        channel_name = "Backorder Alerts"
        channel = self.env['discuss.channel'].search([('name', '=', channel_name)], limit=1)

        if not channel:
            # Create the channel WITHOUT members (Linux safe)
            channel = self.env['discuss.channel'].create({
                'name': channel_name,
                'channel_type': 'group',
            })

        # Now safely assign members
        channel.write({'channel_partner_ids': [(6, 0, group_partner_ids)]})

        # Exclude current user from unread notifications
        partners_to_notify = channel.channel_partner_ids.filtered(
            lambda p: p.id != self.env.user.partner_id.id
        )

        # Post alert message for each backorder
        for picking in old_backorders:
            receipt_link = (
                f'<a href="/web#id={picking.id}&model=stock.picking&view_type=form">'
                f'{picking.name}</a>'
            )
            msg = channel.message_post(
                body=Markup(
                    f"<b>Receipt:</b> {receipt_link}<br/>"
                    f"<b>Created On:</b> {picking.create_date.strftime('%d-%m-%Y %H:%M')}<br/>"
                    f"<b>This back order has been pending for more than one month.</b>"
                ),
                message_type='comment',
                subtype_xmlid='mail.mt_comment'
            )

            # Mark unread for other users
            if partners_to_notify:
                self._create_unread_notifications(msg, partners_to_notify)

    def _create_unread_notifications(self, message, partners):
        """Mark message as unread for each partner."""
        Notification = self.env['mail.notification'].sudo()
        for partner in partners:
            Notification.create({
                'mail_message_id': message.id,
                'res_partner_id': partner.id,
                'notification_type': 'inbox',
                'is_read': False,
            })

    @api.onchange('no_of_parcel')
    def _onchange_no_of_parcel(self):
        # if self.stock_picking_delivery_ids:
        #     return
        if self.no_of_parcel < 0:
            raise ValidationError(_("No of Parcels cannot be negative."))

        self.stock_picking_delivery_ids = [(5, 0, 0)]  # Clear existing lines

        for i in range(1, self.no_of_parcel + 1):
            self.stock_picking_delivery_ids += self.stock_picking_delivery_ids.new({
                'serial_no': i,
            })


    def print_barcodes_stock_picking(self):
        report_name = 'nhcl_customizations.stock_picking_delivery_barcode'
        return {
            'type': 'ir.actions.report',
            'report_name': report_name,
            'report_type': 'qweb-pdf',
            'res_id': self.id,
            'res_model': 'stock.picking',
        }

    def action_my_button(self):
        self.ensure_one()
        if self.ref_credit_note:
            report = self.env.ref('nhcl_customizations.credit_note_pos')  # This is your report ID
            return report.report_action(self.ref_credit_note)
        else:
            raise UserError("No related Credit Note found to print the report.")

    def action_import_excel(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Import Excel',
            'res_model': 'stock.verification.import',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_stock_picking_id': self.id},
        }

    def action_allow_slip(self):
        self.ensure_one()

        # ✅ Mark slip as completed
        self.allow_slip_done = True

        return self.env.ref(
            "nhcl_customizations.allow_slip_credit_note"
        ).report_action(self)

    # def dev_transport_entry_create(self, rec):
    #     data = {'partner_id': rec.partner_id and rec.partner_id.id or False,
    #             'picking_id': rec and rec.id or False,
    #             'lr_number': rec.lr_number or ' ',
    #             'transport_details_id': rec.transpoter_id and rec.transpoter_id.id or False,
    #             'contact_name': rec.transpoter_id and rec.transpoter_id.contact_name or ' ',
    #             'no_of_parcel': rec.no_of_parcel or 0,
    #             'name': rec.tracking_number or ' ',
    #             }
    #     tra_ent = self.env['dev.transport.entry'].create(data)
    #     if tra_ent and rec.transpoter_route_id:
    #         for line in rec.transpoter_route_id.location_details_ids:
    #             location_detail = {
    #                 'source_location_id': line.source_location_id and line.source_location_id.id or False,
    #                 'destination_location_id': line.destination_location_id and line.destination_location_id.id or False,
    #                 'distance': line.distance,
    #                 'transport_charges': line.transport_charges,
    #                 'time_hour': line.time_hour or ' ',
    #                 'tracking_number': rec.tracking_number or ' ',
    #                 'picking_id': rec.id,
    #                 'transport_entry_id': tra_ent.id,
    #             }
    #             self.env['transport.location.details'].create(location_detail)
    #     return tra_ent

    @api.model_create_multi
    def create(self, vals_list):
        seq = self.env['ir.sequence']
        for vals in vals_list:
            if not vals.get('tracking_number'):
                vals['tracking_number'] = self.env['ir.sequence'].next_by_code(
                    'stock.picking.tracking'
                ) or '/'
            vals['lr_number'] = seq.next_by_code('stock.picking.lr_number')

        return super(Picking, self).create(vals_list)

    @api.depends('move_ids_without_package.nhcl_tax_ids', 'move_ids_without_package.nhcl_price_total',
                 'nhcl_amount_total', 'nhcl_amount_untaxed')
    def _compute_nhcl_tax_totals_json(self):
        for order in self:
            order.nhcl_tax_totals_json = self.env['account.tax']._prepare_tax_totals(
                [x._convert_to_tax_base_line_dict() for x in order.move_ids_without_package],
                order.currency_id or order.company_id.currency_id,
            )

    @api.depends('move_ids_without_package.nhcl_price_total')
    def _amount_all(self):
        for order in self:
            order_lines = order.move_ids_without_package
            amount_untaxed = amount_tax = 0.00
            if order_lines:
                tax_results = self.env['account.tax']._compute_taxes([
                    line._convert_to_tax_base_line_dict()
                    for line in order_lines
                ])
                totals = tax_results['totals']
                amount_untaxed = round(totals.get(order.currency_id, {}).get('nhcl_amount_untaxed', 0.0), 2)
                amount_tax = round(totals.get(order.currency_id, {}).get('nhcl_amount_tax', 0.0), 2)
            order.nhcl_amount_untaxed = amount_untaxed
            order.nhcl_amount_tax = amount_tax
            order.nhcl_amount_total = order.nhcl_amount_untaxed + order.nhcl_amount_tax

    nhcl_last_serial_number = fields.Char('Last Serial Number', compute='_get_last_serial_number')
    is_confirm = fields.Boolean('Is Confirm', copy=False)
    stock_type = fields.Selection(
        [('advertisement', 'Advertisement'), ('ho_operation', 'HO Operation'), ('data_import', 'Data Import'),
         ('inter_state', 'Inter State'),
         ('intra_state', 'Intra State'),
         ('others', 'Others')], string='Stock Type', tracking=True)
    dummy_stock_type = fields.Selection(
        [('advertisement', 'Advertisement'), ('ho_operation', 'HO Operation'),
         ('others', 'Others')], string='Dummy Stock Type', compute='_compute_dummy_stock_type')
    stock_barcode = fields.Char(string='Barcode Scan')
    label_click_count = fields.Integer(string="Label Click Count", default=0)
    transpoter_id = fields.Many2one('dev.transport.details', string='Transport by')
    transpoter_route_id = fields.Many2one('dev.routes.details', string='Transporter Route')
    no_of_parcel = fields.Integer(string='No Of Parcels')
    nhcl_credit_note_count = fields.Integer(string='CN Count')
    tracking_number = fields.Char(string='Tracking Number')
    nhcl_tracking_number = fields.Char(string='Source Number')
    lr_number = fields.Char(string='LR Number')
    vehicle_number = fields.Char(string='Vehicle Number')
    driver_name = fields.Char(string='Driver Name')
    transport_location_line = fields.One2many('transport.location.details', 'picking_id', string='Transport Routes')
    transport_entry_ids = fields.One2many('dev.transport.entry', 'picking_id', string='Transport Entry')
    nhcl_replication_status = fields.Boolean(string='Replication Status')
    stock_verification_ids = fields.One2many('stock.verification', 'stock_picking_id')
    verify_barcode = fields.Char(string='Verification Scan')
    exchange_barcode = fields.Char(string='Exchange Scan')
    stock_picking_type = fields.Selection(string='Type',
                                          tracking=True, related='picking_type_id.stock_picking_type')
    nhcl_pos_order = fields.Many2one('pos.order', string="Same POS Order", copy=False)
    nhcl_purchased_store = fields.Char(string="Purchased Store", copy=False)
    nhcl_invoice_date = fields.Date(string="Bill Date", copy=False)
    currency_id = fields.Many2one('res.currency', 'Currency', required=True, readonly=True,
                                  default=lambda self: self.env.company.currency_id.id, copy=False)
    nhcl_tax_totals_json = fields.Binary(compute='_compute_nhcl_tax_totals_json', copy=False)
    nhcl_amount_untaxed = fields.Monetary(string='Untaxed Amount', store=True, readonly=True, compute='_amount_all',
                                          tracking=True, copy=False)
    nhcl_amount_tax = fields.Monetary(string='Taxes', store=True, readonly=True, compute='_amount_all', copy=False)
    nhcl_amount_total = fields.Monetary(string='Total', store=True, readonly=True, compute='_amount_all', copy=False)
    stock_operation_type = fields.Selection([('scan', 'Scan'), ('import', 'Import')
                                                , ('document', 'Document')], string='Operation Type',
                                            tracking=True, default='scan')
    stock_document = fields.Many2one('stock.picking', string="Document", copy=False,
                                     domain=[('stock_picking_type', '=', 'receipt'), ('state', '=', 'done')])
    company_type = fields.Selection([('same', 'Same'), ('other', 'Other')], string="Company Type", copy=False)
    store_pos_order = fields.Char('Other Pos Order', copy=False)
    store_name = fields.Many2one('nhcl.ho.store.master', string='Store Name', copy=False)
    ref_credit_note = fields.Many2one('account.move', string="Ref. Credit Note",
                                      domain="[('move_type', '=', 'out_refund')]")
    nhcl_phone = fields.Char(string="Phone", related='partner_id.phone')
    scan_or_import = fields.Selection([
        ('scan', 'Scan'),
        ('import', 'Import')
    ], string="Scan or Import", default='scan')
    lot_qty = fields.Float(string="LOT Qty.")
    receipt_lot_qty = fields.Float(string="LOT Qty.")
    exchange_reason = fields.Text(string='Reason', tracking=True)
    nhcl_excess_qty = fields.Float(string="Excess No's", compute='nhcl_get_excess_qty')
    nhcl_shortage_qty = fields.Float(string="Shortage Qty")
    nhcl_dm_reason = fields.Selection([('torn', 'Torn'),('fade', 'Fade'),
        ('strains', 'Strains'),('other', 'Other')], string="Damage Reason", copy=False, tracking=True)
    nhcl_total_grc_qty =  fields.Float(string="Total Qty", compute='compute_nhcl_total_grc_qty', store=True)

    @api.depends('stock_verification_ids.stock_qty')
    def compute_nhcl_total_grc_qty(self):
        for rec in self:
            if rec.stock_verification_ids:
                rec.nhcl_total_grc_qty = sum(rec.stock_verification_ids.mapped('stock_qty'))
            else:
                rec.nhcl_total_grc_qty = 0.00

    def auto_validate_ob_grc(self):
        store_ob_records = self.env['stock.picking'].sudo().search(
            [('state', 'not in', ['done', 'cancel']),
             ('stock_type', '=', 'data_import'), ])
        if store_ob_records:
            for order in store_ob_records:
                order.sudo().button_validate()

    def _get_last_serial_number(self):
        auto_generate_seq_rec = self.env['nhcl.master.sequence'].search(
            [('nhcl_code', '=', 'Auto Serial Number'), ('nhcl_state', '=', 'activate')])
        if auto_generate_seq_rec and auto_generate_seq_rec.nhcl_next_number > 1:
            self.nhcl_last_serial_number = auto_generate_seq_rec.nhcl_prefix + str(
                auto_generate_seq_rec.nhcl_next_number - 1)
        else:
            self.nhcl_last_serial_number = '0'

    def nhcl_get_excess_qty(self):
        for rec in self:
            if rec.name:
                last_scanned = self.env['last.scanned.serial.number'].search_count([('receipt_number', '=', rec.name)])
                if last_scanned:
                    self.nhcl_excess_qty = last_scanned
                else:
                    self.nhcl_excess_qty = 0
            else:
                self.nhcl_excess_qty = 0

    def write(self, vals):
        res = super(Picking, self).write(vals)
        for pick in self:
            if pick.stock_picking_type == 'receipt':
                for picking in pick.filtered(lambda p: p.state != 'done'):
                    verification_map = {}
                    for ver in picking.stock_verification_ids:
                        if ver.stock_product_id and ver.type_product:
                            verification_map[ver.stock_product_id.id] = ver.type_product
                    for move in picking.move_ids_without_package:
                        if not move.type_product and move.product_id.id in verification_map:
                            move.type_product = verification_map[move.product_id.id]
        return res

    def get_receipt_picking_lines(self):
        self.ensure_one()
        if self.stock_document:
            # Check if this stock document is already used in another picking
            existing_usage = self.search([
                ('id', '!=', self.id),
                ('stock_document', '=', self.stock_document.id)
            ])
            if existing_usage:
                raise ValidationError("This document has already been used in another picking.")
            # Clear existing lines before generating new ones
            self.move_ids_without_package.unlink()
            self.move_line_ids_without_package.unlink()
            # Get stock moves from the selected document
            moves = self.stock_document.move_ids_without_package.filtered(lambda m: m.state == 'done')
            for move in moves:
                # Create new move for the current picking
                new_move = self.move_ids_without_package.create({
                    'name': move.name,
                    'product_id': move.product_id.id,
                    'product_uom_qty': move.product_uom_qty,
                    'product_uom': move.product_uom.id,
                    'picking_id': self.id,
                    'location_id': move.location_id.id,
                    'location_dest_id': move.location_dest_id.id,
                })
                # Get stock move lines
                for move_line in move.move_line_ids:
                    # Create new move line for the current move
                    self.env['stock.move.line'].create({
                        'move_id': new_move.id,
                        'product_id': move_line.product_id.id,
                        'lot_id': move_line.lot_id.id,
                        'quantity': move_line.quantity,
                        'product_uom_id': move_line.product_uom_id.id,
                        'location_id': move_line.location_id.id,
                        'location_dest_id': move_line.location_dest_id.id,
                        'picking_id': self.id,
                    })

    @api.constrains('nhcl_pos_order')
    def check_nhcl_pos_order(self):
        for rec in self:
            if rec.nhcl_pos_order and rec.stock_picking_type == 'exchange':
                if not rec.move_ids_without_package:
                    raise ValidationError("No products are available Exchanged...")

    # @api.constrains('store_pos_order')
    # def check_store_pos_order(self):
    #     for rec in self:
    #         if rec.store_pos_order:
    #             used_other = self.env['stock.picking'].search_count([('store_pos_order', '=', rec.store_pos_order)])
    #             if used_other > 1:
    #                 raise ValidationError(f"Already Used in another Exchange.")

    # @api.onchange('company_type')
    # def set_stock_type(self):
    #     for rec in self:
    #         rec.stock_type = 'ho_operation'

    @api.onchange('nhcl_pos_order')
    def nhcl_get_pos_order(self):
        for rec in self:

            if rec.state == 'done':
                raise ValidationError(_("Picking in Done State"))

            # ---------------------------------------
            # RESET VALUES
            # ---------------------------------------
            rec.partner_id = False
            rec.nhcl_invoice_date = False
            rec.nhcl_purchased_store = False
            rec.customer_phone = False

            # unlink existing moves safely
            rec.move_line_ids_without_package.unlink()
            rec.move_ids_without_package.unlink()

            if not rec.nhcl_pos_order:
                continue

            # ---------------------------------------
            # HO CONFIG
            # ---------------------------------------
            ho_id = self.env['nhcl.ho.store.master'].search([
                ('nhcl_store_type', '=', 'ho'),
                ('nhcl_active', '=', True)
            ], limit=1)

            if not ho_id:
                raise ValidationError(_("HO Configuration not found"))

            ho_ip = ho_id.nhcl_terminal_ip
            ho_port = ho_id.nhcl_port_no
            ho_api_key = ho_id.nhcl_api_key

            headers_source = {
                'api-key': ho_api_key,
                'Content-Type': 'application/json'
            }

            # ---------------------------------------
            # GET POS ORDER
            # ---------------------------------------
            pos_url = f"http://{ho_ip}:{ho_port}/api/pos.order/search"

            pos_domain = [
                ('pos_reference', '=', rec.nhcl_pos_order.pos_reference),
                ('is_pos_order_used', '=', False)
            ]

            pos_full_url = f"{pos_url}?domain={pos_domain}"

            pos_response = requests.get(
                pos_full_url,
                headers=headers_source
            ).json()

            if not pos_response.get("data"):
                raise ValidationError(_("POS Order not found"))

            pos_order = pos_response.get("data")[0]

            # ---------------------------------------
            # PARTNER
            # ---------------------------------------
            partner_id = False

            if pos_order.get('partner_id'):

                partner_api = f"http://{ho_ip}:{ho_port}/api/res.partner/search"

                partner_domain = [
                    ('id', '=', pos_order['partner_id'][0]['id'])
                ]

                partner_url = f"{partner_api}?domain={partner_domain}"

                partner_response = requests.get(
                    partner_url,
                    headers=headers_source
                ).json()

                if partner_response.get("data"):

                    partner_data = partner_response.get("data")[0]

                    phone = partner_data.get('phone')

                    if not phone:
                        raise ValidationError(_("Customer phone number missing"))

                    partner_id = self.env['res.partner'].search([
                        ('phone', '=', phone)
                    ], limit=1)

                    if not partner_id:

                        partner_category = self.env['res.partner.category'].search([
                            ('name', '=', 'Customer')
                        ], limit=1)

                        if not partner_category:
                            partner_category = self.env['res.partner.category'].sudo().create({
                                'name': 'Customer'
                            })

                        partner_id = self.env['res.partner'].sudo().create({
                            'name': partner_data.get('name'),
                            'phone': phone,
                            'vat': "1234567890123Z1",
                            'group_contact': partner_category.id
                        })

                    rec.partner_id = partner_id.id
                    rec.customer_phone = phone

            # ---------------------------------------
            # HEADER VALUES
            # ---------------------------------------
            rec.nhcl_invoice_date = pos_order.get("date_order")
            rec.nhcl_purchased_store = pos_order.get("company_id")[0]['name']
            rec.stock_type = 'ho_operation'

            # ---------------------------------------
            # POS LINES
            # ---------------------------------------
            move = []
            move_line = []
            for line_data in pos_order.get("lines", []):

                # ---------------------------------------
                # GET POS ORDER LINE
                # ---------------------------------------
                pos_line_api = f"http://{ho_ip}:{ho_port}/api/pos.order.line/search"

                line_domain = [
                    ('id', '=', line_data['id']),
                    ('is_pos_order_used_line', '=', False)
                ]

                line_url = f"{pos_line_api}?domain={line_domain}"

                line_response = requests.get(
                    line_url,
                    headers=headers_source
                ).json()

                if not line_response.get("data"):
                    continue

                line = line_response.get("data")[0]

                # ---------------------------------------
                # PRODUCT
                # ---------------------------------------
                product_id = False

                product_api = f"http://{ho_ip}:{ho_port}/api/product.product/search"

                product_domain = [
                    ('id', '=', line['product_id'][0]['id'])
                ]

                product_url = f"{product_api}?domain={product_domain}"

                product_response = requests.get(
                    product_url,
                    headers=headers_source
                ).json()

                if product_response.get("data"):

                    product_data = product_response.get("data")[0]

                    product_code = product_data.get("default_code")

                    if product_code:
                        product_id = self.env['product.product'].search([
                            ('default_code', '=', product_code)
                        ], limit=1)

                    if not product_id:
                        product_id = self.env['product.product'].search([
                            ('name', 'ilike', line['full_product_name'])
                        ], limit=1)

                    if not product_id:
                        product_id = self.env['product.product'].create({
                            'name': line['full_product_name'],
                            'detailed_type': 'service'
                        })

                if not product_id:
                    continue

                # ---------------------------------------
                # LOT DETAILS
                # ---------------------------------------
                pack_lot_name = False
                lot_ref = False
                type_product = False
                lot_data = False

                if line.get("lot_ids"):

                    lot_api = f"http://{ho_ip}:{ho_port}/api/stock.lot/search"

                    lot_domain = [
                        ('name', '=', line["lot_ids"][0]['name'])
                    ]

                    lot_url = f"{lot_api}?domain={lot_domain}"

                    lot_response = requests.get(
                        lot_url,
                        headers=headers_source
                    ).json()

                    if lot_response.get("data"):
                        pack_lot_name = lot_response.get("data")[0]

                        lot_ref = pack_lot_name.get("ref")
                        type_product = pack_lot_name.get("type_product")

                        lot_data = self.env['stock.lot'].search([
                            ('name', '=', pack_lot_name['name'])
                        ], limit=1)

                # ---------------------------------------
                # TAX
                # ---------------------------------------
                tax_id = False

                if line.get("tax_ids"):
                    tax = self.env['account.tax'].search([
                        ('name', '=', line["tax_ids"][0]["name"])
                    ], limit=1)

                    tax_id = tax.id

                # ---------------------------------------
                # CREATE STOCK MOVE
                # ---------------------------------------
                move_id = {

                    'location_dest_id': rec.location_dest_id.id,
                    'location_id': rec.location_id.id,

                    'name': product_id.display_name,
                    'picking_id': rec.id,

                    'product_id': product_id.id,
                    'product_uom': product_id.uom_id.id,

                    'product_uom_qty': line["qty"],
                    'quantity': line["qty"],
                    'nhcl_old_qty': line["qty"],

                    'nhcl_rsp': line.get('price_unit', 0.0),

                    'nhcl_tax_ids': [(6, 0, [tax_id])] if tax_id else False,

                    'nhcl_discount': line.get('discount', 0.0),
                    'nhcl_gdiscount': line.get('gdiscount', 0.0),
                    'nhcl_disc_lines': line.get('disc_lines'),

                    'type_product': type_product,

                    'serial_no': pack_lot_name["name"] if pack_lot_name else False,

                    'ref_pos_order_line_id': line['id'],

                    'nhcl_cost_price': line.get('nhcl_cost_price', 0.0),
                    'nhcl_rs_price': line.get('nhcl_rs_price', 0.0),
                    'nhcl_mr_price': line.get('nhcl_mr_price', 0.0),

                    'move_brand_barcode': lot_ref,
                }

                move.append((0, 0, move_id))

                # ---------------------------------------
                # CREATE MOVE LINE
                # ---------------------------------------
                move_line_vals = {
                    'product_id': product_id.id,
                    'product_uom_id': product_id.uom_id.id,

                    'location_id': rec.location_id.id,
                    'location_dest_id': rec.location_dest_id.id,

                    'quantity': line["qty"],

                    'internal_ref_lot': lot_ref,

                    'rs_price': line.get('nhcl_rs_price', 0.0),
                    'mr_price': line.get('nhcl_mr_price', 0.0),
                    'cost_price': line.get('nhcl_cost_price', 0.0),

                    'type_product': type_product,

                    'lot_name': pack_lot_name["name"] if pack_lot_name else False,

                    'categ_1': lot_data.category_1.id if lot_data and lot_data.category_1 else False,
                    'categ_2': lot_data.category_2.id if lot_data and lot_data.category_2 else False,
                    'categ_3': lot_data.category_3.id if lot_data and lot_data.category_3 else False,
                    'categ_4': lot_data.category_4.id if lot_data and lot_data.category_4 else False,
                    'categ_5': lot_data.category_5.id if lot_data and lot_data.category_5 else False,
                    'categ_6': lot_data.category_6.id if lot_data and lot_data.category_6 else False,
                    'categ_7': lot_data.category_7.id if lot_data and lot_data.category_7 else False,
                    'categ_8': lot_data.category_8.id if lot_data and lot_data.category_8 else False,

                    'descrip_1': lot_data.description_1.id if lot_data and lot_data.description_1 else False,
                    'descrip_2': lot_data.description_2.id if lot_data and lot_data.description_2 else False,
                    'descrip_3': lot_data.description_3.id if lot_data and lot_data.description_3 else False,
                    'descrip_4': lot_data.description_4.id if lot_data and lot_data.description_4 else False,
                    'descrip_5': lot_data.description_5.id if lot_data and lot_data.description_5 else False,
                    'descrip_6': lot_data.description_6.id if lot_data and lot_data.description_6 else False,
                    'descrip_7': lot_data.description_7.id if lot_data and lot_data.description_7 else False,
                    'descrip_8': lot_data.description_8.id if lot_data and lot_data.description_8 else False,
                }
                move_line.append((0, 0, move_line_vals))
            rec.move_ids_without_package = move
            rec.move_line_ids_without_package = move_line

    @api.onchange('company_type')
    def _onchange_company_type(self):
        self.store_name = False
        self.store_pos_order = False


    def nhcl_get_pos_order_from_diff_store(self):
        for rec in self:
            if rec.state == 'done':
                raise ValidationError(_("Picking in Done State"))
            rec.partner_id = False
            rec.nhcl_invoice_date = False
            rec.nhcl_purchased_store = False
            rec.move_ids_without_package = [(5, 0, 0)]
            rec.move_line_ids_without_package = [(5, 0, 0)]
            if rec.store_pos_order != False:
                ho_id = self.env['nhcl.ho.store.master'].search(
                    [('nhcl_store_type', '=', 'ho'), ('nhcl_active', '=', True)])

                ho_ip = ho_id.nhcl_terminal_ip
                ho_port = ho_id.nhcl_port_no
                ho_api_key = ho_id.nhcl_api_key
                headers_source = {'api-key': f"{ho_api_key}", 'Content-Type': 'application/json'}
                pos_data = f"http://{ho_ip}:{ho_port}/api/pos.order/search"
                pos_data_domain = [('pos_reference', '=', rec.store_pos_order), ('is_pos_order_used', '=', False)]
                pos_data_url = f"{pos_data}?domain={pos_data_domain}"
                pos = requests.get(pos_data_url, headers=headers_source).json()
                if pos.get("data"):
                    pos_partner = False
                    line_partner_data = f"http://{ho_ip}:{ho_port}/api/res.partner/search"
                    if pos.get("data")[0]['partner_id']:
                        line_partner_data_domain = [('id', '=', pos.get("data")[0]['partner_id'][0]['id'])]
                        line_partner_data_url = f"{line_partner_data}?domain={line_partner_data_domain}"
                        pos_partner = requests.get(line_partner_data_url, headers=headers_source).json()
                    partner_id = False
                    if pos_partner == False:
                        raise ValidationError(_('Invalid Customer for %s in %s store', rec.store_pos_order,
                                                rec.store_name.nhcl_store_name))
                    if pos_partner and pos_partner.get("data"):
                        if pos_partner.get("data")[0]['phone'] == False:
                            raise ValidationError(_('Customer Invalid Phone Number'))
                        phone = pos_partner.get("data")[0]['phone']

                        partner_id = self.env['res.partner'].search([('phone', '=', phone)])
                        if not partner_id:
                            partner_category = self.env['res.partner.category'].search([('name', '=', 'Customer')])
                            if not partner_category:
                                partner_category = self.env['res.partner.category'].sudo().create({'name': 'Customer'})
                            partner_id = self.env['res.partner'].sudo().create(
                                {'name': pos_partner.get("data")[0]['name'], 'phone': phone,
                                 'vat': "1234567890123Z1", 'group_contact': partner_category.id})
                    pos_order_id = pos.get("data")[0]["id"]
                    rec.partner_id = partner_id
                    rec.nhcl_invoice_date = pos.get("data")[0]["date_order"]
                    rec.nhcl_purchased_store = pos.get("data")[0]["company_id"][0]['name']
                    rec.stock_type = 'ho_operation'
                    for line_data in pos.get("data")[0]["lines"]:
                        pos_line_data = f"http://{ho_ip}:{ho_port}/api/pos.order.line/search"
                        pos_line_data_domain = [('id', '=', line_data['id']), ('order_id', '=', pos_order_id),
                                                ('is_pos_order_used_line', '=', False)]
                        pos_line_data_url = f"{pos_line_data}?domain={pos_line_data_domain}"
                        pos_line = requests.get(pos_line_data_url, headers=headers_source).json()
                        line = False
                        product_id = False
                        pack_lot_ids = False
                        if pos_line.get("data"):
                            line = pos_line.get("data")[0]
                            line_product_data = f"http://{ho_ip}:{ho_port}/api/product.product/search"
                            line_product_data_domain = [('id', '=', pos_line.get("data")[0]['product_id'][0]['id'])]
                            line_product_data_url = f"{line_product_data}?domain={line_product_data_domain}"
                            pos_line_product = requests.get(line_product_data_url, headers=headers_source).json()
                            if pos_line_product.get("data"):
                                product_data = pos_line_product.get("data")[0]
                                product_type = product_data.get("nhcl_detailed_type")  # product / consumable / service
                                product_name = product_data.get("name")
                                ser_prod = line["full_product_name"]
                                product_nhcl_id = product_data.get("nhcl_id")
                                product_code = product_data.get("default_code")
                                # 1. Storable Product → Match using nhcl_id
                                # -----------------------------------------------
                                if product_type == 'product' and product_code:
                                    product_id = self.env['product.product'].search([
                                        ('default_code', '=', product_code)
                                    ], limit=1)
                                # 2. Other product types → Match by name
                                # ------------------------------------------------
                                if not product_id:
                                    product_id = self.env['product.product'].search([
                                        '|',
                                        ('name', 'ilike', ser_prod),
                                        ('product_tmpl_id.name', 'ilike', ser_prod)
                                    ], limit=1)
                                # 3. Final check
                                # ------------------------------------------------
                                if not product_id:
                                    product_id = self.env['product.product'].create({'name': ser_prod,
                                                                                     'detailed_type': 'service'})
                        pack_lot_name = False
                        nhcl_categ_1 = False
                        nhcl_categ_2 = False
                        nhcl_categ_3 = False
                        nhcl_categ_4 = False
                        nhcl_categ_5 = False
                        nhcl_categ_6 = False
                        nhcl_categ_7 = False
                        nhcl_categ_8 = False
                        nhcl_descrip_1 = False
                        nhcl_descrip_2 = False
                        nhcl_descrip_3 = False
                        nhcl_descrip_4 = False
                        nhcl_descrip_5 = False
                        nhcl_descrip_6 = False
                        nhcl_descrip_7 = False
                        nhcl_descrip_8 = False
                        if line:
                            if "lot_ids" in line:
                                pack_lot_ids = line["lot_ids"]

                                lot_id_data = f"http://{ho_ip}:{ho_port}/api/stock.lot/search"
                                lot_id_data_domain = [('name', '=', pack_lot_ids[0]['name'])]
                                lot_id_data_url = f"{lot_id_data}?domain={lot_id_data_domain}"
                                product_lot = requests.get(lot_id_data_url, headers=headers_source).json()
                                pack_lot_name = product_lot.get("data")[0]
                                if pack_lot_name["category_1"]:
                                    nhcl_categ_1 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["category_1"][0]["name"]),
                                         ('attribute_id.name', '=', 'Color')])
                                if pack_lot_name["category_2"]:
                                    nhcl_categ_2 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["category_2"][0]["name"]),
                                         ('attribute_id.name', '=', 'Fit')])
                                if pack_lot_name["category_3"]:
                                    nhcl_categ_3 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["category_3"][0]["name"]),
                                         ('attribute_id.name', '=', 'Brand')])
                                if pack_lot_name["category_4"]:
                                    nhcl_categ_4 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["category_4"][0]["name"]),
                                         ('attribute_id.name', '=', 'Pattern')])
                                if pack_lot_name["category_5"]:
                                    nhcl_categ_5 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["category_5"][0]["name"]),
                                         ('attribute_id.name', '=', 'Border Type')])
                                if pack_lot_name["category_6"]:
                                    nhcl_categ_6 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["category_6"][0]["name"]),
                                         ('attribute_id.name', '=', 'Border Size')], limit=1)
                                if pack_lot_name["category_7"]:
                                    nhcl_categ_7 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["category_7"][0]["name"]),
                                         ('attribute_id.name', '=', 'Size')])
                                if pack_lot_name["category_8"]:
                                    nhcl_categ_8 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["category_8"][0]["name"]),
                                         ('attribute_id.name', '=', 'Design')])
                                if pack_lot_name["description_1"]:
                                    nhcl_descrip_1 = self.env['product.aging.line'].search(
                                        [('name', '=', pack_lot_name["description_1"][0]["name"])])
                                    print('nhcl_descrip_1', nhcl_descrip_1.mapped('name'), nhcl_descrip_1.mapped('id'))
                                if pack_lot_name["description_2"]:
                                    nhcl_descrip_2 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["description_2"][0]["name"]),
                                         ('attribute_id.name', '=', 'Range')])
                                if pack_lot_name["description_3"]:
                                    nhcl_descrip_3 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["description_3"][0]["name"]),
                                         ('attribute_id.name', '=', 'Collection')])
                                if pack_lot_name["description_4"]:
                                    nhcl_descrip_4 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["description_4"][0]["name"]),
                                         ('attribute_id.name', '=', 'Fabric')], limit=1)
                                if pack_lot_name["description_5"]:
                                    nhcl_descrip_5 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["description_5"][0]["name"]),
                                         ('attribute_id.name', '=', 'Exclusive')])
                                if pack_lot_name["description_6"]:
                                    nhcl_descrip_6 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["description_6"][0]["name"]),
                                         ('attribute_id.name', '=', 'Print')])
                                if pack_lot_name["description_7"]:
                                    nhcl_descrip_7 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["description_7"][0]["name"]),
                                         ('attribute_id.name', '=', 'Days Ageing')])
                                if pack_lot_name["description_8"]:
                                    nhcl_descrip_8 = self.env['product.attribute.value'].search(
                                        [('name', '=', pack_lot_name["description_8"][0]["name"])])

                            tax_id = False
                            if 'tax_ids' in line:
                                tax = self.env['account.tax'].search([('name', '=', line["tax_ids"][0]["name"])],
                                                                     limit=1)
                                tax_id = tax.id
                            move_id = self.env['stock.move'].create({
                                'location_dest_id': rec.location_dest_id.id,
                                'location_id': rec.location_id.id,
                                'name': product_id.display_name,
                                'picking_id': rec.id,
                                'product_id': product_id.id,
                                'product_uom_qty': line["qty"],
                                'quantity': line["qty"],
                                'nhcl_old_qty': line["qty"],
                                'nhcl_rsp': line['price_unit'] if line['price_unit'] != False else 0.0,
                                'nhcl_tax_ids': [(6, 0, [tax_id])] if tax_id else [],
                                'nhcl_discount': line['discount'],
                                'nhcl_gdiscount': line['gdiscount'],
                                # 'nhcl_is_fix_discount_line': line['is_fix_discount_line'],
                                'nhcl_disc_lines': line['disc_lines'],
                                'type_product': pack_lot_name["type_product"] if pack_lot_name != False else False,
                                "serial_no": pack_lot_name["name"] if pack_lot_name != False else False,
                                'ref_pos_order_line_id': line['id'],
                                'nhcl_cost_price': line['nhcl_cost_price'],
                                'nhcl_rs_price': line['nhcl_rs_price'],
                                'nhcl_mr_price': line['nhcl_mr_price'],
                                'move_brand_barcode': pack_lot_name["ref"],
                            })
                            for move_line in move_id.move_line_ids:
                                move_line.write(
                                    {"internal_ref_lot": pack_lot_name["ref"] if pack_lot_name != False else False,
                                     'rs_price': move_id.nhcl_rs_price,
                                     'mr_price': move_id.nhcl_mr_price,
                                     'cost_price': move_id.nhcl_cost_price,
                                     'type_product': move_id.type_product,
                                     "lot_name": pack_lot_name["name"] if pack_lot_name != False else False,
                                     'categ_1': nhcl_categ_1.id if nhcl_categ_1 else False,
                                     'categ_2': nhcl_categ_2.id if nhcl_categ_2 else False,
                                     'categ_3': nhcl_categ_3.id if nhcl_categ_3 else False,
                                     'categ_4': nhcl_categ_4.id if nhcl_categ_4 else False,
                                     'categ_5': nhcl_categ_5.id if nhcl_categ_5 else False,
                                     'categ_6': nhcl_categ_6.id if nhcl_categ_6 else False,
                                     'categ_7': nhcl_categ_7.id if nhcl_categ_7 else False,
                                     'categ_8': nhcl_categ_8.id if nhcl_categ_8 else False,
                                     'descrip_1': nhcl_descrip_1.id if nhcl_descrip_1 else False,
                                     'descrip_2': nhcl_descrip_2.id if nhcl_descrip_2 else False,
                                     'descrip_3': nhcl_descrip_3.id if nhcl_descrip_3 else False,
                                     'descrip_4': nhcl_descrip_4.id if nhcl_descrip_4 else False,
                                     'descrip_5': nhcl_descrip_5.id if nhcl_descrip_5 else False,
                                     'descrip_6': nhcl_descrip_6.id if nhcl_descrip_6 else False,
                                     'descrip_7': nhcl_descrip_7.id if nhcl_descrip_7 else False,
                                     'descrip_8': nhcl_descrip_8.id if nhcl_descrip_8 else False,
                                     })
                # else:
                #     raise ValidationError(
                #         _('Pos Order Number does not exist in %s or Already Used.', rec.store_name.nhcl_store_name))

    def stock_create_credit_note(self):
        """Creates and posts a credit note from stock picking"""
        for picking in self:
            if picking.nhcl_credit_note_count >= 1:
                raise ValidationError("Credit Note is already generated against " + picking.name)
            if not picking.partner_id:
                raise ValidationError("Customer is missing for this picking!")
            picking.nhcl_credit_note_count += 1
            journal = self.env['account.journal'].search([('type', '=', 'cash'), ('name', '=', 'Credit Note Issue')],
                                                         limit=1)
            if not journal:
                raise ValidationError(
                    "No journal found with type 'cash' and name 'Credit Note Issue'. Please configure it.")
            if picking.nhcl_pos_order:
                pos_bill_ref = picking.nhcl_pos_order.pos_reference
            else:
                pos_bill_ref = picking.store_pos_order
            credit_note_vals = {
                'move_type': 'out_refund',
                'partner_id': picking.partner_id.id,
                'invoice_origin': picking.name,
                'picking_ref': picking.name,
                'pos_bill_ref': pos_bill_ref,
                'currency_id': self.env.company.currency_id.id,
                'journal_id': journal.id,
                'invoice_date': fields.Date.context_today(self),
                'invoice_line_ids': [],
            }
            credit_note = self.env['account.move'].create(credit_note_vals)
            has_lines = False
            for move in picking.move_ids_without_package:
                if move.quantity > 0:
                    product = move.product_id
                    account = product.property_account_income_id
                    if not account:
                        raise ValidationError(f"Income account is not defined for product {product.display_name}.")
                    price_unit = move.nhcl_rsp
                    if move.nhcl_discount:
                        price_unit -= (price_unit * move.nhcl_discount / 100)
                    if move.nhcl_gdiscount:
                        price_unit -= (price_unit * move.nhcl_gdiscount / 100)

                    self.env['account.move.line'].create({
                        'move_id': credit_note.id,
                        'product_id': product.id,
                        'quantity': move.quantity,
                        'price_unit': price_unit,
                        'name': "Credit Note",
                        'account_id': account.id,
                        'tax_ids': [(6, 0, move.nhcl_tax_ids.ids)],
                    })
                    has_lines = True
            if not has_lines:
                raise ValidationError("No credit note lines created. Credit note cannot be posted.")
            credit_note.action_post()
            if credit_note.state == 'posted':
                picking.partner_id.wallet_amount += credit_note.amount_total
                picking.ref_credit_note = credit_note.id
                picking.partner_id.credit_note_ids = [(0, 0, {
                    'voucher_number': picking.ref_credit_note.name,
                    'pos_bill_number': picking.nhcl_pos_order.pos_reference,
                    'pos_bill_date': picking.nhcl_invoice_date,
                    'total_amount': credit_note.amount_total,
                })]

    def update_pos_order_status_other_store_ho(self):
        for rec in self:

            ho_id = self.env['nhcl.ho.store.master'].search(
                [('nhcl_store_type', '=', 'ho'), ('nhcl_active', '=', True)])

            ho_ip = ho_id.nhcl_terminal_ip
            ho_port = ho_id.nhcl_port_no
            ho_api_key = ho_id.nhcl_api_key

            headers = {
                'api-key': ho_api_key,
                'Content-Type': 'application/json'
            }

            try:
                # --------------------------------------------------
                # STEP 1: Get POS Order
                # --------------------------------------------------
                pos_search_url = f"http://{ho_ip}:{ho_port}/api/pos.order/search"
                if rec.company_type == 'other':
                    pos_domain = [('pos_reference', '=', rec.store_pos_order)]
                else:
                    pos_domain = [('pos_reference', '=', rec.nhcl_pos_order.pos_reference)]
                pos_url = f"{pos_search_url}?domain={pos_domain}"

                response = requests.get(pos_url, headers=headers)
                response.raise_for_status()

                pos_data = response.json().get("data", [])
                if not pos_data:
                    return

                pos_order_id = pos_data[0]['id']

                # --------------------------------------------------
                # STEP 2: Get validated serial numbers from picking
                # --------------------------------------------------
                validated_serials = rec.move_line_ids_without_package.filtered(
                    lambda x: x.lot_name
                ).mapped('lot_name')

                if not validated_serials:
                    return

                # --------------------------------------------------
                # STEP 3: Get ALL POS Order Lines
                # --------------------------------------------------
                line_search_url = f"http://{ho_ip}:{ho_port}/api/pos.order.line/search"
                line_domain = [('order_id', '=', pos_order_id)]
                line_url = f"{line_search_url}?domain={line_domain}"

                line_response = requests.get(line_url, headers=headers)
                line_response.raise_for_status()

                pos_lines = line_response.json().get("data", [])

                # --------------------------------------------------
                # STEP 4: Check pack_lot_ids for each line
                # --------------------------------------------------
                for line in pos_lines:

                    pack_lot_ids = line.get("lot_ids")

                    if not pack_lot_ids:
                        continue

                    lot_id = pack_lot_ids[0]['name']

                    # Compare serial
                    if lot_id in validated_serials:
                        update_line_url = (
                            f"http://{ho_ip}:{ho_port}/api/pos.order.line/{line['id']}"
                        )

                        requests.put(
                            update_line_url,
                            headers=headers,
                            json={'is_pos_order_used_line': True}
                        )

                # --------------------------------------------------
                # STEP 5: Check if ALL lines are used
                # --------------------------------------------------
                # Re-fetch lines
                line_response = requests.get(line_url, headers=headers)
                line_response.raise_for_status()

                updated_lines = line_response.json().get("data", [])

                if not updated_lines:
                    return

                all_used = all(
                    l.get('is_pos_order_used_line') == True
                    for l in updated_lines
                )

                # --------------------------------------------------
                # STEP 6: Update POS Order if all lines used
                # --------------------------------------------------
                if all_used:
                    update_order_url = (
                        f"http://{ho_ip}:{ho_port}/api/pos.order/{pos_order_id}"
                    )

                    requests.put(
                        update_order_url,
                        headers=headers,
                        json={'is_pos_order_used': True}
                    )

                    _logger.info(
                        f"POS Order '{rec.store_pos_order}' marked as fully used."
                    )

            except requests.exceptions.RequestException as e:
                _logger.error(
                    f"Failed to update POS order '{rec.store_pos_order}'. Error: {e}"
                )

    def update_pos_order_status_other_store(self):
        for rec in self:

            store_ip = rec.store_name.nhcl_terminal_ip
            store_port = rec.store_name.nhcl_port_no
            store_api_key = rec.store_name.nhcl_api_key

            headers = {
                'api-key': store_api_key,
                'Content-Type': 'application/json'
            }

            try:
                # --------------------------------------------------
                # STEP 1: Get POS Order
                # --------------------------------------------------
                pos_search_url = f"http://{store_ip}:{store_port}/api/pos.order/search"
                pos_domain = [('pos_reference', '=', rec.store_pos_order)]
                pos_url = f"{pos_search_url}?domain={pos_domain}"

                response = requests.get(pos_url, headers=headers)
                response.raise_for_status()

                pos_data = response.json().get("data", [])
                if not pos_data:
                    return

                pos_order_id = pos_data[0]['id']

                # --------------------------------------------------
                # STEP 2: Get validated serial numbers from picking
                # --------------------------------------------------
                validated_serials = rec.move_line_ids_without_package.filtered(
                    lambda x: x.lot_name
                ).mapped('lot_name')

                if not validated_serials:
                    return

                # --------------------------------------------------
                # STEP 3: Get ALL POS Order Lines
                # --------------------------------------------------
                line_search_url = f"http://{store_ip}:{store_port}/api/pos.order.line/search"
                line_domain = [('order_id', '=', pos_order_id)]
                line_url = f"{line_search_url}?domain={line_domain}"

                line_response = requests.get(line_url, headers=headers)
                line_response.raise_for_status()

                pos_lines = line_response.json().get("data", [])

                # --------------------------------------------------
                # STEP 4: Check pack_lot_ids for each line
                # --------------------------------------------------
                for line in pos_lines:

                    pack_lot_ids = line.get("pack_lot_ids")

                    if not pack_lot_ids:
                        continue

                    lot_id = pack_lot_ids[0]['id']

                    # Fetch lot record
                    lot_search_url = f"http://{store_ip}:{store_port}/api/pos.pack.operation.lot/search"
                    lot_domain = [('id', '=', lot_id)]
                    lot_url = f"{lot_search_url}?domain={lot_domain}"

                    lot_response = requests.get(lot_url, headers=headers)
                    lot_response.raise_for_status()

                    lot_data = lot_response.json().get("data", [])

                    if not lot_data:
                        continue

                    lot_name = lot_data[0].get("lot_name")

                    # Compare serial
                    if lot_name in validated_serials:
                        update_line_url = (
                            f"http://{store_ip}:{store_port}/api/pos.order.line/{line['id']}"
                        )

                        requests.put(
                            update_line_url,
                            headers=headers,
                            json={'is_pos_order_used_line': True}
                        )

                # --------------------------------------------------
                # STEP 5: Check if ALL lines are used
                # --------------------------------------------------
                # Re-fetch lines
                line_response = requests.get(line_url, headers=headers)
                line_response.raise_for_status()

                updated_lines = line_response.json().get("data", [])

                if not updated_lines:
                    return

                all_used = all(
                    l.get('is_pos_order_used_line') == True
                    for l in updated_lines
                )

                # --------------------------------------------------
                # STEP 6: Update POS Order if all lines used
                # --------------------------------------------------
                if all_used:
                    update_order_url = (
                        f"http://{store_ip}:{store_port}/api/pos.order/{pos_order_id}"
                    )

                    requests.put(
                        update_order_url,
                        headers=headers,
                        json={'is_pos_order_used': True}
                    )

                    _logger.info(
                        f"POS Order '{rec.store_pos_order}' marked as fully used."
                    )

            except requests.exceptions.RequestException as e:
                _logger.error(
                    f"Failed to update POS order '{rec.store_pos_order}'. Error: {e}"
                )

    def update_the_grc_status_ho(self):
        for rec in self:
            ho_id = self.env['nhcl.ho.store.master'].search(
                [('nhcl_store_type', '=', 'ho'), ('nhcl_active', '=', True)])

            ho_ip = ho_id.nhcl_terminal_ip
            ho_port = ho_id.nhcl_port_no
            ho_api_key = ho_id.nhcl_api_key

            headers = {
                'api-key': ho_api_key,
                'Content-Type': 'application/json'
            }
            try:
                ho_packet_search_url = f"http://{ho_ip}:{ho_port}/api/stock.picking/search"
                ho_packet_domain = [('name', '=', rec.origin)]
                ho_packet_url = f"{ho_packet_search_url}?domain={ho_packet_domain}"

                ho_packet_response = requests.get(ho_packet_url, headers=headers)
                ho_packet_response.raise_for_status()

                ho_packet_lines = ho_packet_response.json().get("data", [])
                if ho_packet_lines:
                    picking_data = {
                        'is_opened' : rec.is_opened,
                        'is_received' : rec.is_received,
                        'nhcl_excess_qty' : rec.nhcl_excess_qty,
                        'nhcl_shortage_qty' : rec.nhcl_shortage_qty,
                    }
                    packet_order_id = ho_packet_lines[0]['id']
                    update_order_url = (
                        f"http://{ho_ip}:{ho_port}/api/stock.picking/{packet_order_id}"
                    )

                    requests.put(
                        update_order_url,
                        headers=headers,
                        json=picking_data
                    )

            except requests.exceptions.RequestException as e:
                ho_id.create_cmr_transaction_server_replication_log('GRC Update', rec.id, rec.name, 500,
                                                                                    'add', "failure", e)

    @api.onchange('partner_id')
    def get_stock_type(self):
        for rec in self:
            rec.stock_type = False

            if rec.picking_type_id.code == 'outgoing':
                if rec.partner_id and rec.env.company.state_id:
                    if rec.partner_id.state_id.id == rec.env.company.state_id.id:
                        rec.stock_type = 'intra_state'
                    else:
                        rec.stock_type = 'inter_state'

            elif rec.picking_type_id.code == 'internal':
                if rec.partner_id and rec.partner_id.name:
                    partner_name = rec.partner_id.name.upper()
                    if 'TEXTILE' in partner_name:
                        rec.stock_type = 'ho_operation'

    @api.onchange('picking_type_id')
    def nhcl_picking_type(self):
        if self.picking_type_id and self.picking_type_code == 'internal':
            self.stock_type = 'intra_state'

    def apply_rounding_line(self):
        # total_rsp = sum(self.move_ids_without_package.filtered(lambda x:x.serial_no != False).mapped('nhcl_price_total'))
        total_rsp = round(self.nhcl_tax_totals_json.get('amount_total'), 2)
        total = total_rsp
        decimal_part = round(total - math.floor(total), 2)
        diff = 0.0
        rounding_product = None
        # 0.51 → 0.89 → Round Up → .90
        if 0.50 <= decimal_part <= 0.99:
            rounded_total = math.ceil(total)
            diff = round(rounded_total - total, 2)
            rounding_product = self.env.ref('nhcl_customizations.nhcl_product_product_round_up')
        # 0.01 → 0.49 → Round Down → .00
        elif 0.01 <= decimal_part <= 0.49:
            rounded_total = math.floor(total)
            diff = round(rounded_total - total, 2)
            rounding_product = self.env.ref('nhcl_customizations.nhcl_product_product_round_down')
        else:
            return  # .00 or .50 — no rounding change
        if not diff:
            return
        # Create rounding move
        self.env['stock.move'].create({
            'picking_id': self.id,
            'product_id': rounding_product.id,
            'name': rounding_product.name,
            'product_uom': rounding_product.uom_id.id,
            'product_uom_qty': 1,
            'nhcl_rsp': diff,
            'location_id': self.location_id.id,
            'location_dest_id': self.location_dest_id.id,
        })

    # Store button
    def button_validate(self):
        res = None
        for rec in self:
            exchange_moves = rec.move_ids_without_package.filtered(
                lambda m: m.nhcl_exchange
            )

            if exchange_moves and not rec.allow_slip_done:
                raise UserError("Please click 'Allow Slip' before validating.")
            if rec.stock_picking_type == 'exchange' and rec.picking_type_code == 'incoming':
                zero_qty_moves = exchange_moves.filtered(lambda m: m.nhcl_move_qty <= 0 and m.tracking == 'lot')
                if zero_qty_moves:
                    raise UserError("Quantity cannot be zero. Please enter a valid quantity.")
                if not self.env.context.get('bypass_exchange_wizard') and any(
                        rec.move_ids_without_package.filtered(lambda x: x.nhcl_exchange == True)) == True:
                    return {
                        'name': _('Exchange Confirmation'),
                        'type': 'ir.actions.act_window',
                        'target': 'new',
                        'res_model': 'pos.exchange.wizard',
                        'view_mode': 'form',
                        'view_id': self.env.ref('nhcl_customizations.view_pos_exchange_wizard_wizard').id,
                        'context': {'default_nhcl_picking_id': rec.id,
                                    'bypass_exchange_wizard': True,
                                    },
                    }
                discount_lines = rec.move_ids_without_package.filtered(
                    lambda line: not line.serial_no and line.nhcl_rsp < 0 and '% on your order' in (
                            line.product_id.name or '').lower()
                )
                all_saleable_lines =  rec.move_ids_without_package.filtered(
                    lambda line: not line.serial_no and line.nhcl_rsp < 0 and 'All-Saleable-PoS-DISCOUNT' in (
                            line.product_id.name or '')
                )
                if discount_lines:
                    valid_lines = rec.move_ids_without_package.filtered(
                        lambda line: line.serial_no and line.nhcl_rsp >= 0)
                    order_total_qty = sum(valid_lines.mapped('quantity'))
                    for disc_line in discount_lines:
                        if valid_lines:
                            share = abs(disc_line.nhcl_rsp) / order_total_qty
                            for valid_line in valid_lines:
                                valid_line.nhcl_rsp -= share

                if all_saleable_lines:
                    valid_all_saleable_lines = rec.move_ids_without_package.filtered(
                        lambda line: line.serial_no and line.nhcl_rsp >= 0)
                    order_total_qty = sum(valid_all_saleable_lines.mapped('quantity'))
                    for valid_all_saleable_line in all_saleable_lines:
                        if valid_all_saleable_lines:
                            share = abs(valid_all_saleable_line.nhcl_rsp) / order_total_qty
                            for valid_all_line in valid_all_saleable_lines:
                                valid_all_line.nhcl_rsp -= share
                cheapest_lines = rec.move_ids_without_package.filtered(
                    lambda line: not line.serial_no and line.nhcl_rsp < 0 and '% on the cheapest product' in (
                        (line.product_id.name or '').lower())
                )

                if cheapest_lines:
                    # Find all valid sale lines with serial and positive rsp
                    valid_lines = rec.move_ids_without_package.filtered(
                        lambda line: line.serial_no and line.nhcl_rsp >= 0)
                    if valid_lines:
                        cheapest_valid_line = valid_lines.sorted(lambda l: l.nhcl_rsp)[0]
                        for discount_line in cheapest_lines:
                            # Apply entire discount to the cheapest product only
                            cheapest_valid_line.nhcl_rsp += discount_line.nhcl_rsp
                gift_card_line = rec.move_ids_without_package.filtered(
                    lambda line: not line.serial_no and line.nhcl_rsp < 0 and 'Gift Card' in (
                            line.product_id.name or '')
                )
                if gift_card_line:
                    valid_lines_list = rec.move_ids_without_package.filtered(
                        lambda line: line.serial_no and line.nhcl_rsp >= 0)
                    total_qty = sum(valid_lines_list.mapped('quantity'))
                    if total_qty > 0:
                        for gift_line in gift_card_line:
                            gift_share_per_unit = abs(gift_line.nhcl_rsp) / total_qty
                            for valid_line in valid_lines_list:
                                valid_line.nhcl_rsp -= gift_share_per_unit
                discount_amount_line = rec.move_ids_without_package.filtered(
                    lambda line: not line.serial_no and line.nhcl_rsp < 0 and 'Discount' in (
                            line.product_id.name or '')
                )
                if discount_amount_line:
                    valid_lines_list = rec.move_ids_without_package.filtered(
                        lambda line: line.serial_no and line.nhcl_rsp >= 0)
                    total_qty = sum(valid_lines_list.mapped('quantity'))
                    if total_qty > 0:
                        for disc_amt_line in discount_amount_line:
                            disc_amt_line_share_per_unit = abs(disc_amt_line.nhcl_rsp) / total_qty
                            for valid_disc_amt_line in valid_lines_list:
                                valid_disc_amt_line.nhcl_rsp -= disc_amt_line_share_per_unit
                # Free Product case: negative RSP lines that represent a free product offer
                free_product_lines = rec.move_ids_without_package.filtered(
                    lambda line: not line.serial_no
                                 and line.nhcl_rsp < 0
                                 and ('free product' in (line.product_id.name or '').lower())
                )

                if free_product_lines:
                    valid_serial_lines = rec.move_ids_without_package.filtered(
                        lambda line: line.serial_no and line.nhcl_rsp >= 0
                    )

                    # If you only want to adjust respective product lines
                    for free_line in free_product_lines:
                        free_name = (free_line.product_id.name or '').lower()

                        # Remove "free product -" prefix and strip spaces
                        if free_name.startswith('free product -'):
                            target_name = free_name.replace('free product -', '').strip()
                        else:
                            continue  # safety fallback
                        # Match valid serial line by name instead of product_id
                        matching_line = valid_serial_lines.filtered(
                            lambda v: (v.product_id.name or '').lower().strip() == target_name
                        )
                        if matching_line:
                            # Move entire negative RSP to the matched product line
                            matching_line[0].nhcl_rsp += free_line.nhcl_rsp
                        else:
                            # If no proper match → apply to cheapest serial item
                            if valid_serial_lines:
                                cheapest_valid_line = valid_serial_lines.sorted(lambda l: l.nhcl_rsp)[0]
                                cheapest_valid_line.nhcl_rsp += free_line.nhcl_rsp
                rec.stock_pos_exchange()
                rec.stock_disc_pos_exchange()
                # Filter out move lines where nhcl_exchange is False and unlink them
                rec.move_ids_without_package.filtered(lambda line: not line.nhcl_exchange).unlink()
                rec.apply_rounding_line()
                # rec.nhcl_pos_order.is_pos_order_used = True
            # Update lot_ids from dummy_lot_ids before validation
            for move_line in rec.move_ids_without_package:
                if move_line.dummy_lot_ids:
                    move_line.lot_ids = [(6, 0, move_line.dummy_lot_ids.ids)]
                    # if rec.company_type == "same":
                    #     move_line.move_line_ids.update({'cost_price': move_line.dummy_lot_ids.cost_price})
                    if rec.company_type in ["other",'same']:
                        move_line.move_line_ids.lot_id.update({
                            'cost_price': move_line.nhcl_cost_price,
                            'category_1': move_line.categ_1.id if move_line.categ_1 else False,
                            'category_2': move_line.categ_2.id if move_line.categ_2 else False,
                            'category_3': move_line.categ_3.id if move_line.categ_3 else False,
                            'category_4': move_line.categ_4.id if move_line.categ_4 else False,
                            'category_5': move_line.categ_5.id if move_line.categ_5 else False,
                            'category_6': move_line.categ_6.id if move_line.categ_6 else False,
                            'category_7': move_line.categ_7.id if move_line.categ_7 else False,
                            'category_8': move_line.categ_8.id if move_line.categ_8 else False,
                            'description_1': move_line.descrip_1.id if move_line.descrip_1 else False,
                            'description_2': move_line.descrip_2.id if move_line.descrip_2 else False,
                            'description_3': move_line.descrip_3.id if move_line.descrip_3 else False,
                            'description_4': move_line.descrip_4.id if move_line.descrip_4 else False,
                            'description_5': move_line.descrip_5.id if move_line.descrip_5 else False,
                            'description_6': move_line.descrip_6.id if move_line.descrip_6 else False,
                            'description_7': move_line.descrip_7.id if move_line.descrip_7 else False,
                            'description_8': move_line.descrip_8.id if move_line.descrip_8 else False,
                        })
            # if rec.batch_id and rec.stock_picking_type in ('return', 'damage') and rec.picking_type_code == 'outgoing':
            #     rec.write({
            #         # 'transpoter_id': rec.batch_id.transpoter_id.id,
            #         # 'transpoter_route_id': rec.batch_id.transpoter_route_id.id,
            #         'no_of_parcel': rec.batch_id.no_of_parcel,
            #         # 'lr_number': rec.batch_id.lr_number,
            #         # 'driver_name': rec.batch_id.driver_name,
            #         # 'vehicle_number': rec.batch_id.vehicle_number,
            #     })
            # rec.transpoter_id == rec.batch_id.transpoter_id.id
            if rec.picking_type_id.code == 'incoming' and rec.stock_type == 'data_import':
                # ---------- Barcode validation ----------
                missing_barcodes = rec.move_ids_without_package.filtered(
                    lambda m: not m.product_id.barcode
                )
                if missing_barcodes:
                    raise ValidationError(
                        _("Please generate barcode for products: %s") %
                        ", ".join(missing_barcodes.mapped('product_id.display_name'))
                    )

                # ---------- Brand type validation ----------
                missing_brand_type = rec.move_ids_without_package.filtered(
                    lambda m: not m.type_product
                )
                if missing_brand_type:
                    raise ValidationError(_("Please Select The Brand Type."))

                # ---------- Serial check ----------
                serial_moves = rec.move_ids_without_package.filtered(
                    lambda m: m.product_id.tracking != 'none'
                )

                serial_move_lines = serial_moves.mapped('move_line_ids')

                no_serial_lines = serial_move_lines.filtered(
                    lambda ml: not ml.lot_id and not ml.lot_name
                )

                if no_serial_lines and not self.env.context.get('skip_serial_popup'):
                    return {
                        'name': _('Auto Serial No Confirmation'),
                        'type': 'ir.actions.act_window',
                        'target': 'new',
                        'res_model': 'nhcl.serial.no.popup',
                        'view_mode': 'form',
                        'view_id': self.env.ref(
                            'nhcl_customizations.crm_serial_no_confirm_popup_view'
                        ).id,
                        'context': {
                            'default_nhcl_picking_id': rec.id,
                            'skip_serial_popup': True
                        },
                    }

                # ---------- qty_done safety ----------
                for move in rec.move_ids_without_package:
                    total_done = sum(move.move_line_ids.mapped('quantity'))

                    if total_done == 0:
                        raise UserError(
                            _("Done quantity is not set for product %s") %
                            move.product_id.display_name
                        )

                # ---------- Optional : auto fill qty_done from demand ----------
                for move in rec.move_ids_without_package:
                    for ml in move.move_line_ids:
                        if ml.quantity == 0:
                            ml.quantity = ml.product_uom_qty
                            if ml.lot_id:
                                ml.lot_id.write({
                                    'cost_price': ml.cost_price,
                                    'mr_price': ml.mr_price,
                                    'rs_price': ml.rs_price,
                                })
            if rec.stock_picking_type == 'receipt' and rec.picking_type_code == 'incoming' and rec.stock_type != 'data_import':

                total_grc_qty = sum(rec.stock_verification_ids.mapped('stock_qty'))
                actual_qty = sum(rec.stock_verification_ids.mapped('stock_actual_qty'))

                diff_qty = total_grc_qty - actual_qty

                rec.nhcl_shortage_qty = diff_qty if diff_qty > 0 else 0

                # --------------------------------------------------
                # Update Move Quantities
                # --------------------------------------------------
                product_qty_map = {}

                for line in rec.stock_verification_ids:
                    product_id = line.stock_product_id.id

                    if product_id not in product_qty_map:
                        product_qty_map[product_id] = {
                            'stock_qty': 0.0,
                            'actual_qty': 0.0,
                        }

                    product_qty_map[product_id]['stock_qty'] += line.stock_qty
                    product_qty_map[product_id]['actual_qty'] += line.stock_actual_qty

                for move in rec.move_ids:
                    vals = product_qty_map.get(move.product_id.id)

                    if vals:
                        move.write({
                            'product_uom_qty': vals['stock_qty'],
                            'quantity': vals['actual_qty'],
                        })
                # --------------------------------------------------
                # Delete Existing Move Lines
                # --------------------------------------------------
                rec.move_line_ids_without_package.unlink()

                # --------------------------------------------------
                # Collect Attribute Names
                # --------------------------------------------------
                attribute_names = set()
                aging_names = set()

                for line in rec.stock_verification_ids:

                    attribute_names.update(filter(None, [
                        line.nhcl_categ_1,
                        line.nhcl_categ_2,
                        line.nhcl_categ_3,
                        line.nhcl_categ_4,
                        line.nhcl_categ_5,
                        line.nhcl_categ_6,
                        line.nhcl_categ_7,
                        line.nhcl_categ_8,
                        line.nhcl_descrip_2,
                        line.nhcl_descrip_3,
                        line.nhcl_descrip_4,
                        line.nhcl_descrip_5,
                        line.nhcl_descrip_6,
                    ]))

                    if line.nhcl_descrip_1:
                        aging_names.add(line.nhcl_descrip_1)

                attribute_map = {
                    attr.name: attr.id
                    for attr in self.env['product.attribute.value'].search([
                        ('name', 'in', list(attribute_names))
                    ])
                }

                aging_map = {
                    aging.name: aging.id
                    for aging in self.env['product.aging.line'].search([
                        ('name', 'in', list(aging_names))
                    ])
                }

                # --------------------------------------------------
                # Prepare Move Line Values
                # --------------------------------------------------
                move_line_vals = []

                for line in rec.stock_verification_ids:
                    move_line_vals.append({
                        'picking_id': rec.id,
                        'company_id': rec.company_id.id,
                        'product_id': line.stock_product_id.id,

                        'quantity': line.stock_actual_qty,
                        'lot_name': line.stock_serial,

                        'type_product': line.type_product,
                        'internal_ref_lot': line.stock_product_barcode,
                        'mr_price': line.mr_price,
                        'rs_price': line.rs_price or 0.0,
                        'cost_price': line.cost_price or 0.0,
                        'segment': line.segment,
                        'nhcl_lot_hsn_code': line.nhcl_lot_hsn_code,

                        'sale_tax_ids': [(6, 0, line.sale_tax_ids.ids)],
                        'purchase_tax_ids': [(6, 0, line.purchase_tax_ids.ids)],

                        'categ_1': attribute_map.get(line.nhcl_categ_1),
                        'categ_2': attribute_map.get(line.nhcl_categ_2),
                        'categ_3': attribute_map.get(line.nhcl_categ_3),
                        'categ_4': attribute_map.get(line.nhcl_categ_4),
                        'categ_5': attribute_map.get(line.nhcl_categ_5),
                        'categ_6': attribute_map.get(line.nhcl_categ_6),
                        'categ_7': attribute_map.get(line.nhcl_categ_7),
                        'categ_8': attribute_map.get(line.nhcl_categ_8),

                        'descrip_1': aging_map.get(line.nhcl_descrip_1),
                        'descrip_2': attribute_map.get(line.nhcl_descrip_2),
                        'descrip_3': attribute_map.get(line.nhcl_descrip_3),
                        'descrip_4': attribute_map.get(line.nhcl_descrip_4),
                        'descrip_5': attribute_map.get(line.nhcl_descrip_5),
                        'descrip_6': attribute_map.get(line.nhcl_descrip_6),
                    })

                # --------------------------------------------------
                # Bulk Create Move Lines
                # --------------------------------------------------
                if move_line_vals:
                    self.env['stock.move.line'].create(move_line_vals)
                # ----------------------------------------
                # Update Stock Moves
                # ----------------------------------------
                product_qty_map = {}
                for line in rec.stock_verification_ids:
                    product_id = line.stock_product_id.id
                    if product_id not in product_qty_map:
                        product_qty_map[product_id] = {
                            'stock_qty': 0.0,
                            'actual_qty': 0.0,
                        }
                    product_qty_map[product_id]['stock_qty'] += line.stock_qty
                    product_qty_map[product_id]['actual_qty'] += line.stock_actual_qty
                for move in rec.move_ids:
                    qty_vals = product_qty_map.get(move.product_id.id)
                    if not qty_vals:
                        continue
                    move.write({
                        'product_uom_qty': qty_vals['stock_qty'],
                        'quantity': qty_vals['actual_qty'],
                    })
            res = super(Picking, self).button_validate()
            if rec.stock_picking_type == 'hpi':
                not_scanned = rec.hired_product_ids.filtered(lambda l: not l.returned_scan)

                if not_scanned:
                    raise UserError("Please scan all serial numbers before validating return.")

            # Unlock hired serials after return delivery is done
            if rec.state == 'done' and rec.stock_picking_type == 'hpi':
                for line in rec.hired_product_ids:
                    if line.lot_number:
                        line.lot_number.hired_product = False

            # if rec.state == 'done' and rec.transpoter_id and rec.stock_type in ['inter_state', 'intra_state']:
            #     rec.dev_transport_entry_create(rec)
            if rec.state == 'done' and rec.company_type == 'same':
                rec.update_same_store_flag()
                rec.update_pos_order_status_other_store()
                rec.update_pos_order_status_other_store_ho()
            if rec.company_type == 'other':
                rec.update_pos_order_status_other_store()
                rec.update_pos_order_status_other_store_ho()
        return res

    def update_same_store_flag(self):
        for pick in self:
            pos_order = pick.nhcl_pos_order
            if not pos_order or not pos_order.lines:
                continue
            # Get only exchange moves with serial numbers
            exchange_moves = pick.move_ids_without_package.filtered(lambda m: m.nhcl_exchange and m.serial_no)
            if not exchange_moves:
                continue
            # Create a set of serial numbers from moves
            move_serials = set(exchange_moves.mapped('serial_no'))
            # Filter POS lines whose lot_name matches move serial
            matched_lines = pos_order.lines.filtered(
                lambda l: l.pack_lot_ids and l.pack_lot_ids.lot_name in move_serials)
            matched_lines.write({
                'is_pos_order_used_line': True
            })

    def stock_pos_exchange(self):
        for picking in self:
            for move in picking.move_ids_without_package:
                if move.nhcl_disc_lines:
                    try:
                        prod_list = ast.literal_eval(move.nhcl_disc_lines or "[]")
                    except Exception as e:
                        print(f"Error parsing nhcl_disc_lines: {e}")
                        continue
                    if not prod_list:
                        continue
                    # Normalize and clean prod names from discount line
                    prod_list_clean = [str(p).strip().lower() for p in prod_list]
                    # Find all matching lines
                    matching_moves = picking.move_ids_without_package.filtered(
                        lambda m: m != move and m.product_id.nhcl_display_name.strip().split(']', 1)[
                            -1].strip().lower() in prod_list_clean)
                    if matching_moves:
                        share = move.nhcl_rsp / len(matching_moves)
                        for matched_move in matching_moves:
                            print(f" - Reducing {share} from {matched_move.product_id.nhcl_display_name}")
                            matched_move.nhcl_rsp += share

    def stock_disc_pos_exchange(self):
        """
        Split POS discount line amount into mapped product lines
        and update RSP / unit_price accordingly.
        """
        import ast

        for picking in self:
            for move in picking.move_ids_without_package:

                # Skip if no discount mapping
                if not move.nhcl_disc_lines:
                    continue

                # -----------------------------------------
                # Parse mapping list safely
                # -----------------------------------------
                try:
                    disc_product_list = ast.literal_eval(move.nhcl_disc_lines or "[]")
                except Exception:
                    disc_product_list = []

                if not disc_product_list:
                    continue

                # Normalize IDs to integers
                disc_product_list = [
                    int(i)
                    for i in disc_product_list
                    if str(i).isdigit()
                ]

                # Discount value to be distributed
                discount_value = move.nhcl_rsp
                if not discount_value:
                    continue

                # -----------------------------------------
                # Collect matching serial & lot moves
                # -----------------------------------------
                matching_serial_moves = picking.move_ids_without_package.filtered(
                    lambda m: (
                            m != move
                            and m.product_id.tracking == 'serial'
                            and m.product_id.nhcl_id in disc_product_list
                    )
                )

                matching_lot_moves = picking.move_ids_without_package.filtered(
                    lambda m: (
                            m.product_id.tracking == 'lot'
                            and m.product_id.nhcl_id in disc_product_list
                    )
                )

                # -----------------------------------------
                # SERIAL PRODUCTS — split equally
                # -----------------------------------------
                if matching_serial_moves:
                    matching_count = len(matching_serial_moves)
                    if matching_count == 0:
                        continue

                    share = discount_value / matching_count  # equal split

                    for matched_move in matching_serial_moves:
                        matched_move.nhcl_rsp = matched_move.nhcl_rsp + share

                    continue  # don't apply lot logic also

                # -----------------------------------------
                # LOT PRODUCTS — split by quantity
                # -----------------------------------------
                if matching_lot_moves:
                    total_qty = sum(matching_lot_moves.mapped("product_uom_qty"))
                    if total_qty == 0:
                        continue

                    # Discount per unit
                    per_unit_share = discount_value / total_qty

                    for matched_move in matching_lot_moves:
                        qty = matched_move.product_uom_qty
                        added_discount = qty * per_unit_share
                        matched_move.nhcl_rsp = matched_move.nhcl_rsp + added_discount

                    continue

    @api.depends('stock_type')
    def _compute_dummy_stock_type(self):
        for i in self:
            if i.stock_type == 'ho_operation':
                i.dummy_stock_type = 'ho_operation'
            elif i.stock_type == 'advertisement':
                i.dummy_stock_type = 'advertisement'
            elif i.stock_type == 'others':
                i.dummy_stock_type = 'others'
            elif i.stock_type == 'inter_state':
                i.dummy_stock_type = 'ho_operation'
            elif i.stock_type == 'intra_state':
                i.dummy_stock_type = 'ho_operation'
            elif i.stock_type == 'data_import':
                i.dummy_stock_type = 'ho_operation'
            else:
                i.dummy_stock_type = ''

    @api.onchange('exchange_barcode')
    def _onchange_exchange_barcode(self):
        if not self.exchange_barcode:
            return
        barcode = self.exchange_barcode.strip()
        gs1_pattern = r'01(\d{14})21([A-Za-z0-9]+)'
        ean13_pattern = r'^\d{13}$'
        custom_serial_pattern = r'(^(?=.*\d)[A-Za-z0-9_/\- ]+$)'
        gs1_match = re.match(gs1_pattern, barcode)
        ean13_match = re.match(ean13_pattern, barcode)
        custom_match = re.match(custom_serial_pattern, barcode)
        # ---------------- GS1 (UNCHANGED - YOUR WORKING LOGIC) ----------------
        if gs1_match:
            product_barcode = gs1_match.group(1)
            code = gs1_match.group(2)
            matched_line = self.move_ids_without_package.filtered(
                lambda x: x.serial_no == code and not x.nhcl_exchange)
            if not matched_line:
                raise ValidationError('No matching product or serial/lot found.')
            line = matched_line[0]
            tracking = line.product_id.tracking
            if tracking == 'serial':
                if self.receipt_lot_qty != 0:
                    raise ValidationError("Qty not allowed for serial.")
                line.nhcl_exchange = True
            elif tracking == 'lot':
                if self.receipt_lot_qty <= 0:
                    raise ValidationError("Enter qty for lot.")
                if self.receipt_lot_qty > line.nhcl_old_qty:
                    raise ValidationError("Qty exceeds available.")
                line.nhcl_move_qty = self.receipt_lot_qty
                if self.receipt_lot_qty == line.nhcl_old_qty:
                    line.nhcl_exchange = True
        # ---------------- EAN-13 (FIXED) ----------------
        elif ean13_match:
            ean = barcode
            lines = self.move_ids_without_package.filtered(
                lambda x: x.move_brand_barcode == ean and not x.nhcl_exchange)
            if not lines:
                raise ValidationError('No matching product found.')
            tracking = lines[0].product_id.tracking
            # -------- SERIAL --------
            if tracking == 'serial':
                line = lines.filtered(lambda l: not l.nhcl_exchange)[:1]
                if not line:
                    raise ValidationError("All serials already matched.")
                if self.receipt_lot_qty != 0:
                    raise ValidationError("Qty not allowed for serial.")
                line.nhcl_exchange = True
            # -------- LOT (FIFO SAME AS GS1 STYLE) --------
            elif tracking == 'lot':
                if self.receipt_lot_qty <= 0:
                    raise ValidationError("Enter qty.")
                total_available = sum(l.nhcl_old_qty - l.nhcl_move_qty for l in lines)
                if self.receipt_lot_qty > total_available:
                    raise ValidationError(
                        f"Given {self.receipt_lot_qty} exceeds {total_available}"
                    )

                qty = self.receipt_lot_qty

                for line in lines:
                    if qty <= 0:
                        break
                    remaining = line.nhcl_old_qty - line.nhcl_move_qty
                    if remaining <= 0:
                        continue

                    consume = min(qty, remaining)

                    line.stock_actual_qty += consume
                    if line.nhcl_move_qty == line.nhcl_old_qty:
                        line.nhcl_exchange = True
                    qty -= consume
        # ---------------- CUSTOM (FIXED & CLEANED) ----------------
        elif custom_match:
            code = custom_match.group(1)

            unbranded = self.move_ids_without_package.filtered(
                lambda x: x.serial_no == code and x.type_product == 'un_brand' and x.nhcl_exchange != True)
            branded = self.move_ids_without_package.filtered(
                lambda x: x.move_brand_barcode == code and x.type_product == 'brand')
            if not unbranded and not branded:
                raise ValidationError('No matching product found.')
            # -------- UNBRANDED (GS1 STYLE SINGLE LINE) --------
            if unbranded:
                line = unbranded[0]
                tracking = line.product_id.tracking
                if tracking == 'serial':
                    if self.receipt_lot_qty != 0:
                        raise ValidationError("Qty not allowed.")
                    line.nhcl_exchange = True
                elif tracking == 'lot':
                    if self.receipt_lot_qty <= 0:
                        raise ValidationError("Enter qty.")
                    if self.receipt_lot_qty > line.nhcl_old_qty:
                        raise ValidationError("Qty exceeds available.")
                    line.nhcl_move_qty += self.receipt_lot_qty
                    if line.nhcl_move_qty == line.nhcl_old_qty:
                        line.nhcl_exchange = True
            # -------- BRANDED (FIFO LIKE EAN / GS1 LOT STYLE) --------
            else:
                lines = branded.filtered(lambda l: l.nhcl_exchange != True)
                tracking = lines[0].product_id.tracking
                if tracking == 'serial':
                    line = lines[:1]
                    if line.nhcl_exchange == True:
                        raise ValidationError("All serials matched.")
                    if self.receipt_lot_qty != 0:
                        raise ValidationError("Qty not allowed.")
                    line.nhcl_exchange = True
                elif tracking == 'lot':
                    if self.receipt_lot_qty <= 0:
                        raise ValidationError("Enter qty.")
                    total_available = sum(
                        l.nhcl_old_qty - l.nhcl_move_qty for l in lines
                    )
                    if self.receipt_lot_qty > total_available:
                        raise ValidationError("Qty exceeds total available.")
                    qty = self.receipt_lot_qty
                    for line in lines:
                        if qty <= 0:
                            break
                        remaining = line.nhcl_old_qty - line.nhcl_move_qty
                        if remaining <= 0:
                            continue
                        consume = min(qty, remaining)

                        line.nhcl_move_qty += consume
                        if line.nhcl_move_qty == line.nhcl_old_qty:
                            line.nhcl_exchange = True
                        qty -= consume
        else:
            raise ValidationError('Invalid barcode format.')
        self.exchange_barcode = False
        self.receipt_lot_qty = 0.0


    @api.onchange('verify_barcode')
    def _onchange_verify_barcode(self):
        if self.verify_barcode:
            if not any(line.status == 'matched' for line in self.stock_picking_delivery_ids):
                raise ValidationError(_("You cannot proceed. Please scan at least one bale barcode first."))
            barcode = self.verify_barcode
            # Patterns for barcode formats
            gs1_pattern = r'01(\d{14})21([A-Za-z0-9]+)'
            ean13_pattern = r'(\d{13})'
            custom_serial_pattern = r'(^([A-Za-z]+)[A-Za-z0-9_/\- ]*\d.*$)'
            if re.match(gs1_pattern, barcode):
                # GS1 Barcode
                product_barcode = re.match(gs1_pattern, barcode).group(1)
                code = re.match(gs1_pattern, barcode).group(2)
                matched_line = self.stock_verification_ids.filtered(lambda x: x.stock_serial == code)
                if not matched_line:
                    if not matched_line:
                        existing = self.env['last.scanned.serial.number'].search([
                            ('stock_serial', '=', code)
                        ], limit=1)
                        if not existing:
                            self.env['last.scanned.serial.number'].create({
                                'stock_serial': code,
                                'stock_product_barcode': barcode,
                                'receipt_number': self.name,
                                'document_number': self.origin,
                                'store_name': self.company_id.name,
                                'type_product': matched_line.type_product,
                                'stock_qty': matched_line.stock_qty,
                            })
                            self.env.cr.commit()
                            self.nhcl_excess_qty += 1
                    raise ValidationError('No matching product or serial/lot number found.')
                tracking_type = matched_line.stock_product_id.tracking
                product_id = matched_line.stock_product_id
                if tracking_type == 'serial':
                    if self.receipt_lot_qty != 0:
                        raise ValidationError("You not should enter qty.")
                    if matched_line.stock_status == 'matched':
                        raise ValidationError(f"Already Matched {matched_line.stock_serial}.")
                    matched_line.stock_status = 'matched'
                    matched_line.stock_actual_qty = 1.0
                elif tracking_type == 'lot':
                    if self.receipt_lot_qty == 0 or self.receipt_lot_qty < 0:
                        raise ValidationError("You should enter qty.")
                    if matched_line.stock_qty < self.receipt_lot_qty:
                        raise ValidationError(
                            f"You have given {self.receipt_lot_qty} more than {matched_line.stock_qty}.")
                    if matched_line.stock_status == 'matched':
                        raise ValidationError(f"Already Matched {matched_line.stock_serial}.")
                    if matched_line.stock_qty == self.receipt_lot_qty:
                        matched_line.stock_status = 'matched'
                    matched_line.stock_actual_qty = self.receipt_lot_qty

            elif re.match(ean13_pattern, barcode):
                # EAN-13 Barcode
                ean13_barcode = re.match(ean13_pattern, barcode).group(1)

                # Fetch ALL un-matched lines for this barcode
                matched_lines = self.stock_verification_ids.filtered(
                    lambda x: x.stock_product_barcode == ean13_barcode and x.stock_status == 'un_matched'
                )

                if not matched_lines:
                    existing = self.env['last.scanned.serial.number'].search([
                        ('stock_product_barcode', '=', ean13_barcode)
                    ], limit=1)
                    if not existing:
                        self.env['last.scanned.serial.number'].create({
                            'stock_product_barcode': ean13_barcode,
                            'receipt_number': self.name,
                            'document_number': self.origin,
                            'store_name': self.company_id.name,
                            'type_product': matched_lines.type_product,
                            'stock_qty': matched_lines.stock_qty,
                        })
                        self.env.cr.commit()
                        self.nhcl_excess_qty += 1
                    raise ValidationError('No matching product or serial/lot number found.')

                tracking_type = matched_lines[0].stock_product_id.tracking

                if tracking_type == 'serial':
                    if self.receipt_lot_qty != 0:
                        raise ValidationError("You should not enter qty.")
                    if matched_lines[0].stock_status == 'matched':
                        raise ValidationError(f"Already matched {matched_lines[0].stock_serial}.")
                    matched_lines[0].stock_status = 'matched'
                    matched_lines[0].stock_actual_qty = 1.0

                elif tracking_type == 'lot':

                    if self.receipt_lot_qty <= 0:
                        raise ValidationError("You should enter qty.")

                    # --- TOTAL REMAINING QTY ACROSS ALL LINES ---
                    total_available = sum(
                        line.stock_qty - line.stock_actual_qty
                        for line in matched_lines
                    )

                    # --- VALIDATE USER DOES NOT EXCEED TOTAL ---
                    if self.receipt_lot_qty > total_available:
                        raise ValidationError(
                            f"You have given {self.receipt_lot_qty} more than {total_available}."
                        )

                    qty_to_allocate = self.receipt_lot_qty

                    # --- FIFO ALLOCATION ACROSS MULTIPLE LINES ---
                    for line in matched_lines:
                        if qty_to_allocate <= 0:
                            break

                        line_remaining = line.stock_qty - line.stock_actual_qty

                        if qty_to_allocate <= line_remaining:
                            line.stock_actual_qty += qty_to_allocate
                            if line.stock_actual_qty == line.stock_qty:
                                line.stock_status = 'matched'
                            qty_to_allocate = 0
                        else:
                            # consume entire line
                            line.stock_actual_qty = line.stock_qty
                            line.stock_status = 'matched'
                            qty_to_allocate -= line_remaining
            elif re.match(custom_serial_pattern, barcode):
                code = re.match(custom_serial_pattern, barcode).group(1)

                # ---------------- UNBRANDED (SINGLE LINE) ----------------
                matched_line = self.stock_verification_ids.filtered(
                    lambda x: x.stock_serial == code and x.type_product == 'un_brand'
                )

                # ---------------- BRANDED (MULTIPLE LINES) ----------------
                branded_lines = self.stock_verification_ids.filtered(
                    lambda x: x.stock_product_barcode == code and x.type_product == 'brand'
                )

                if not matched_line and not branded_lines:
                    existing = self.env['last.scanned.serial.number'].search([
                        ('stock_serial', '=', code)
                    ], limit=1)
                    if not existing:
                        self.env['last.scanned.serial.number'].create({
                            'stock_serial': code,
                            'receipt_number': self.name,
                            'document_number': self.origin,
                            'store_name': self.company_id.name,
                            'type_product': matched_line.type_product,
                            'stock_qty': matched_line.stock_qty,
                        })
                        self.env.cr.commit()
                        self.nhcl_excess_qty += 1
                    raise ValidationError('No matching product or serial/lot number found.')

                # ---------------- PROCESS UNBRANDED ----------------
                if matched_line:
                    tracking_type = matched_line.stock_product_id.tracking

                    if tracking_type == 'serial':
                        if self.receipt_lot_qty != 0:
                            raise ValidationError("You should not enter qty.")
                        if matched_line.stock_status == 'matched':
                            raise ValidationError(f"Already Matched {matched_line.stock_serial}.")
                        matched_line.stock_status = 'matched'
                        matched_line.stock_actual_qty = 1.0

                    elif tracking_type == 'lot':
                        if self.receipt_lot_qty <= 0:
                            raise ValidationError("You should enter qty.")
                        if matched_line.stock_qty < self.receipt_lot_qty:
                            raise ValidationError(
                                f"You have given {self.receipt_lot_qty} more than {matched_line.stock_qty}."
                            )
                        matched_line.stock_actual_qty += self.receipt_lot_qty
                        if matched_line.stock_actual_qty == matched_line.stock_qty:
                            matched_line.stock_status = 'matched'

                # ---------------- PROCESS BRANDED ----------------
                else:
                    tracking_type = branded_lines[0].stock_product_id.tracking

                    if tracking_type == 'serial':
                        # pick first un-matched branded serial
                        serial_line = branded_lines.filtered(
                            lambda x: x.stock_status == 'un_matched'
                        )[:1]

                        if not serial_line:
                            raise ValidationError(f"All serials already matched for {code}.")

                        if self.receipt_lot_qty != 0:
                            raise ValidationError("You should not enter qty.")

                        serial_line.stock_status = 'matched'
                        serial_line.stock_actual_qty = 1.0

                    elif tracking_type == 'lot':
                        if self.receipt_lot_qty <= 0:
                            raise ValidationError("You should enter qty.")

                        # --- TOTAL AVAILABLE QTY ACROSS ALL BRANDED LOT LINES ---
                        total_available = sum(
                            line.stock_qty - line.stock_actual_qty
                            for line in branded_lines
                        )

                        if self.receipt_lot_qty > total_available:
                            raise ValidationError(
                                f"You have given {self.receipt_lot_qty} more than {total_available}."
                            )

                        qty_to_allocate = self.receipt_lot_qty

                        # --- FIFO ALLOCATION ---
                        for line in branded_lines:
                            if qty_to_allocate <= 0:
                                break

                            remaining = line.stock_qty - line.stock_actual_qty
                            if remaining <= 0:
                                continue

                            if qty_to_allocate <= remaining:
                                line.stock_actual_qty += qty_to_allocate
                                if line.stock_actual_qty == line.stock_qty:
                                    line.stock_status = 'matched'
                                qty_to_allocate = 0
                            else:
                                line.stock_actual_qty = line.stock_qty
                                line.stock_status = 'matched'
                                qty_to_allocate -= remaining
            else:
                raise ValidationError('Invalid barcode format.')
            self.verify_barcode = False
            self.receipt_lot_qty = 0.0

    def action_open_label_type(self):
        # Increment the click count before or after calling super
        self.label_click_count += 1
        # Call the original method using super()
        return super(Picking, self).action_open_label_type()

    @api.constrains('stock_verification_ids')
    def nhcl_assign_qty_receipt(self):
        for pick in self:
            if not pick.stock_verification_ids or not pick.move_line_ids_without_package:
                continue
            # Build lookup: {lot_name: actual_qty}
            verify_qty_map = {
                v.stock_serial: v.stock_actual_qty
                for v in pick.stock_verification_ids
                if v.stock_product_id.tracking == 'lot'
            }
            if not verify_qty_map:
                continue
            for move_line in pick.move_line_ids_without_package:
                qty = verify_qty_map.get(move_line.lot_name)
                if qty is not None:
                    move_line.quantity = qty

    @api.onchange('stock_barcode')
    def _onchange_stock_barcode(self):
        if not self.stock_type:
            if self.stock_barcode:
                raise ValidationError('Please select a Stock Type before scanning a barcode.')
            return

        if self.sale_id:
            raise ValidationError("You cannot scan the serial numbers.")

        barcode = str(self.stock_barcode or '').strip()

        # Patterns
        gs1_pattern = r'01(\d{14})21([A-Za-z0-9]+)'
        ean13_pattern = r'(\d{13})'
        custom_serial_pattern = r'(^(?=.*\d)[A-Za-z0-9_/\- ]+$)'

        # ---------------- HELPER: Search product ----------------
        def search_product(field, value):
            product = self.env['product.product'].search([(field, '=', value)], limit=1)
            if not product:
                template = self.env['product.template'].search([(field, '=', value)], limit=1)
                if template:
                    product = template.product_variant_id
            return product

        # --------------- HELPER: Merge or Create lot line ---------------
        def _merge_or_create_lot_line(product, lot, qty):
            """Merge quantity into existing move line or create a new one.
               Validate against available qty in stock.quant (same location).
            """

            # find quant only for this location (Option A)
            quant = self.env['stock.quant'].search([
                ('lot_id', '=', lot.id),
                ('product_id', '=', product.id),
                ('location_id', '=', self.location_id.id),
            ], limit=1)

            if not quant:
                raise ValidationError(f"No available stock for lot {lot.name} in this location.")
            available_qty = quant.quantity
            # Qty already in this picking
            existing_line = False
            current_qty = 0
            for ml in self.move_ids_without_package:
                if ml.product_id.id == product.id and any(l.name == lot.name for l in ml.dummy_lot_ids):
                    existing_line = ml
                    current_qty += ml.product_uom_qty

            # Qty already used in other open pickings of same type
            other_qty = sum(
                ml.product_uom_qty
                for ml in self.env['stock.move'].search([
                    ('dummy_lot_ids', 'in', lot.ids),
                    ('product_id', '=', product.id),
                    ('move_picking_type', '=', self.stock_picking_type),
                    ('picking_id.state', 'not in', ['cancel', 'done']),
                ])
            )

            # Validate against total available qty
            total_after = current_qty + other_qty + qty
            if total_after > available_qty:
                raise ValidationError(
                    f"Lot {lot.name} exceeds available quantity. "
                    f"Requested {total_after}, Available {available_qty}."
                )

            # Merge
            if existing_line:
                existing_line.product_uom_qty += qty
                return

            # Create new line
            new_vals = {
                'product_id': product.id,
                'product_uom_qty': qty,
                'location_id': self.location_id.id,
                'location_dest_id': self.location_dest_id.id,
                'name': product.display_name,
                'dummy_lot_ids': [(4, lot.id)],
            }
            new_line = self.env['stock.move'].new(new_vals)
            self.move_ids_without_package |= new_line

        # ============ LOCATION HANDLING =================
        location = False
        if self.stock_picking_type in ['main_damage', 'return']:
            location = self.env.ref('stock.stock_location_stock').id
        elif self.stock_picking_type in ['damage_main', 'damage']:
            location = self.env['stock.location'].search([('name', 'ilike', '-DM')], limit=1).id
        elif self.stock_picking_type in ['return_main']:
            location = self.env['stock.location'].search([('name', 'ilike', '-RE')], limit=1).id

        # ====================== GS1 ======================
        if re.match(gs1_pattern, barcode):
            match = re.match(gs1_pattern, barcode)
            product_barcode = match.group(1)
            lot_number = match.group(2)

            product = search_product('barcode', product_barcode)
            if not product:
                raise ValidationError(f'No product found with barcode {product_barcode}')

            # Quant search
            lots = self.env['stock.quant'].search([
                ('lot_id.ref', '=', lot_number),
                ('quantity', '>', 0),
                ('location_id', '=', location)
            ])
            lot = lots.lot_id
            if not lot:
                raise ValidationError(f"No lot/serial found: {lot_number}")

            # Serial tracked
            if product.tracking == 'serial':
                _merge_or_create_lot_line(product, lot[0], 1)

            # Lot tracked
            else:
                if not self.lot_qty or self.lot_qty <= 0:
                    raise ValidationError('Please provide a valid quantity for lot tracked products.')
                _merge_or_create_lot_line(product, lot[0], self.lot_qty)

        # ====================== EAN-13 ======================
        elif re.match(ean13_pattern, barcode):
            match = re.match(ean13_pattern, barcode)
            ean13 = match.group(1)

            lots = self.env['stock.quant'].search([
                ('lot_id.ref', '=', ean13),
                ('quantity', '>', 0),
                ('location_id', '=', location)
            ])
            if not lots:
                raise ValidationError(f"No lots found for barcode {ean13}")

            product = lots[0].product_id

            # Serial
            if product.tracking == 'serial':
                lot = lots.lot_id.filtered(
                    lambda l: l.name not in self.move_ids_without_package.mapped('dummy_lot_ids.name')
                )
                if not lot:
                    raise ValidationError("All serials already used.")
                _merge_or_create_lot_line(product, lot[0], 1)

            # Lot (Progressive consumption with remaining qty check)
            else:
                if not self.lot_qty or self.lot_qty <= 0:
                    raise ValidationError("Please provide a valid quantity.")

                requested_qty = self.lot_qty
                remaining_qty = requested_qty

                consumed_map = {}
                for ml in self.move_ids_without_package:
                    for lot in ml.dummy_lot_ids:
                        consumed_map[lot.id] = consumed_map.get(lot.id, 0) + ml.product_uom_qty

                # Iterate largest lots first
                for quant in lots.sorted(lambda q: q.quantity, reverse=True):
                    if remaining_qty <= 0:
                        break

                    lot = quant.lot_id
                    # qty already consumed in this picking
                    used_in_this_picking = consumed_map.get(lot.id, 0)
                    # true available now
                    available = quant.quantity - used_in_this_picking
                    if available <= 0:
                        continue

                    take_qty = min(remaining_qty, available)
                    _merge_or_create_lot_line(product, lot, take_qty)
                    remaining_qty -= take_qty

                if remaining_qty > 0:
                    raise ValidationError(
                        f"Requested {requested_qty}, "
                        f"But only {requested_qty - remaining_qty} remaining."
                    )



        # ====================== CUSTOM R-SERIAL ======================
        elif re.match(custom_serial_pattern, barcode):
            Quant = self.env['stock.quant']

            # -------------------------------------------------------
            # EXISTING SEARCH — DO NOT MODIFY
            # -------------------------------------------------------
            lots_found = Quant.search([
                ('quantity', '>', 0),
                ('company_id', '=', self.company_id.id),
                ('location_id', '=', location),
                ('lot_id.name', '=', barcode),
            ])

            if not lots_found:
                lots_found = Quant.search([
                    ('quantity', '>', 0),
                    ('company_id', '=', self.company_id.id),
                    ('location_id', '=', location),
                    ('lot_id.ref', '=', barcode),
                ])

            if not lots_found:
                raise ValidationError(f"No lot found with ref '{barcode}'.")

            product = lots_found[0].product_id

            # -------------------------------------------------------
            # FIND ALREADY USED LOTS (ONLY FOR BRANDED)
            # -------------------------------------------------------
            used_lot_ids = self.move_ids_without_package.mapped(
                'dummy_lot_ids'
            ).ids

            # -------------------------------------------------------
            # SERIAL TRACKED → FETCH NEXT SERIAL (FIFO)
            # -------------------------------------------------------
            if product.tracking == 'serial':
                available_quants = lots_found.filtered(
                    lambda q: q.lot_id.id not in used_lot_ids
                ).sorted(key=lambda q: q.id)

                if not available_quants:
                    raise ValidationError(
                        "All serial numbers for this branded barcode are already scanned."
                    )

                _merge_or_create_lot_line(
                    product,
                    available_quants[0].lot_id,
                    1
                )

            # -------------------------------------------------------
            # LOT TRACKED → SAME LOT, SAME LOGIC
            # -------------------------------------------------------
            else:
                if not self.lot_qty or self.lot_qty <= 0:
                    raise ValidationError("Please enter valid quantity.")
                _merge_or_create_lot_line(
                    product,
                    lots_found[0].lot_id,
                    self.lot_qty
                )
        elif self.stock_barcode:
            raise ValidationError("Invalid barcode format")
        # Reset fields
        self.stock_barcode = False
        self.lot_qty = False

    def reset_product_lines(self):
        for rec in self:
            rec.move_ids_without_package.unlink()

    @api.constrains('driver_name')
    def _check_driver_name_characters(self):
        for rec in self:
            if rec.driver_name and not re.match(r'^[A-Za-z ]+$', rec.driver_name):
                raise ValidationError("Driver Name must contain alphabetic characters only.")

    def action_open_import_wizard(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Import Delivery',
            'res_model': 'delivery.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_picking_id': self.id
            }
        }

    # def get_pos_delivery_orders(self):
    #     ho_ids = self.env['nhcl.ho.store.master'].search([
    #         ('nhcl_store_type', '=', 'ho'),
    #         ('nhcl_active', '=', True)
    #     ])
    #
    #     for ho in ho_ids:
    #         ho_ip = ho.nhcl_terminal_ip
    #         ho_port = ho.nhcl_port_no
    #         store_api_key = ho.nhcl_api_key
    #         headers_source = {'api-key': f"{store_api_key}", 'Content-Type': 'application/json'}
    #
    #         picking_type_id = self.env['stock.picking.type'].search([('name', '=', "PoS Orders")])
    #         store_pos_delivery_orders = self.env['stock.picking'].search([
    #             ('picking_type_id', '=', picking_type_id.id),
    #             ('nhcl_replication_status', '=', False),
    #             ('state', '=', 'done')
    #         ])
    #         # store_pos_delivery_orders = self.env['stock.picking'].search([
    #         #     ('name', '=', 'CMR-H/POS/00567'),
    #         #     ('nhcl_replication_status', '=', False),
    #         #     ('state', '=', 'done')
    #         # ])
    #
    #         for order in self:
    #             try:
    #                 if order.location_dest_id.name != "Customers":
    #                     continue
    #
    #                 # -------------------- Company --------------------
    #                 company_url = f"http://{ho_ip}:{ho_port}/api/res.company/search"
    #                 company_domain = [('name', '=', order.company_id.name)]
    #                 company_data = requests.get(
    #                     f"{company_url}?domain={company_domain}", headers=headers_source
    #                 ).json()
    #                 company_id = company_data.get("data")
    #                 if not company_id:
    #                     msg = f"Company not found for order: {order.name}"
    #                     ho.create_cmr_transaction_server_replication_log("failure", msg)
    #                     continue
    #
    #                 # -------------------- Picking Type --------------------
    #                 picking_type_url = f"http://{ho_ip}:{ho_port}/api/stock.picking.type/search"
    #                 picking_type_domain = [('name', '=', "PoS Orders"), ('company_id', '=', company_id[0]['id'])]
    #                 picking_type_data = requests.get(
    #                     f"{picking_type_url}?domain={picking_type_domain}", headers=headers_source
    #                 ).json()
    #                 picking_type = picking_type_data.get("data")
    #                 if not picking_type:
    #                     msg = f"Picking Type not found for company {order.company_id.name}"
    #                     ho.create_cmr_transaction_server_replication_log("failure", msg)
    #                     continue
    #
    #                 # -------------------- Source Location --------------------
    #                 location_url = f"http://{ho_ip}:{ho_port}/api/stock.location/search"
    #                 location_domain = [
    #                     ('name', '=', order.location_id.name),
    #                     ('active', '!=', False),
    #                     ('usage', '=', 'internal'),
    #                     ('company_id', '=', company_id[0]['id'])
    #                 ]
    #                 location_data = requests.get(
    #                     f"{location_url}?domain={location_domain}&fields=['name','id']", headers=headers_source
    #                 ).json()
    #                 location_id = location_data.get("data")
    #                 if not location_id:
    #                     msg = f"Source Location not found for order: {order.name}"
    #                     ho.create_cmr_transaction_server_replication_log("failure", msg)
    #                     continue
    #
    #                 # -------------------- Destination Location --------------------
    #                 dest_location_url = f"http://{ho_ip}:{ho_port}/api/stock.location/search"
    #                 dest_domain = [
    #                     ('complete_name', '=', order.location_dest_id.complete_name),
    #                     ('active', '!=', False),
    #                     ('usage', '=', 'customer')
    #                 ]
    #                 dest_data = requests.get(
    #                     f"{dest_location_url}?domain={dest_domain}", headers=headers_source
    #                 ).json()
    #                 dest_location = dest_data.get("data")
    #                 if not dest_location:
    #                     msg = f"Destination Location not found for order: {order.name}"
    #                     ho.create_cmr_transaction_server_replication_log("failure", msg)
    #                     continue
    #
    #                 # -------------------- Create Picking --------------------
    #                 stock_picking_vals = {
    #                     'picking_type_id': picking_type[0]['id'],
    #                     'origin': order.name,
    #                     'location_id': location_id[0]['id'],
    #                     'location_dest_id': dest_location[0]['id'],
    #                     'company_id': company_id[0]['id'],
    #                     'move_type': 'direct',
    #                     'state': 'done',
    #                     'nhcl_store_delivery': True
    #                 }
    #
    #                 try:
    #                     picking_response = requests.post(
    #                         f"http://{ho_ip}:{ho_port}/api/stock.picking/create",
    #                         headers=headers_source, json=[stock_picking_vals]
    #                     )
    #                     picking_response.raise_for_status()
    #                     stock_picking = picking_response.json()
    #                 except Exception as req_err:
    #                     msg = f"Error creating picking for {order.name}: {req_err}"
    #                     ho.create_cmr_transaction_server_replication_log("failure", msg)
    #                     continue
    #
    #                 if not stock_picking.get("success"):
    #                     msg = f"Picking creation failed for order {order.name}: {stock_picking.get('message')}"
    #                     ho.create_cmr_transaction_server_replication_log("failure", msg)
    #                     continue
    #
    #                 picking_id = stock_picking.get("create_id")
    #
    #                 # -------------------- Create Move Lines --------------------
    #                 for line in order.move_line_ids_without_package:
    #                     try:
    #                         product_domain = [('nhcl_id', '=', line.product_id.nhcl_id)]
    #                         product_data = requests.get(
    #                             f"http://{ho_ip}:{ho_port}/api/product.product/search?domain={product_domain}",
    #                             headers=headers_source
    #                         ).json()
    #                         product_id = product_data.get("data")
    #                         if not product_id:
    #                             msg = f"Product not found for {line.product_id.display_name} in order {order.name}"
    #                             ho.create_cmr_transaction_server_replication_log("failure", msg)
    #                             continue
    #
    #                         lot_name = line.lot_id.name if line.lot_id else None
    #                         move_line_vals = {
    #                             "picking_id": picking_id,
    #                             "product_id": product_id[0]['id'],
    #                             "quantity": line.quantity,
    #                             "location_id": location_id[0]['id'],
    #                             "location_dest_id": dest_location[0]["id"],
    #                             "lot_name": lot_name,
    #                         }
    #
    #                         move_resp = requests.post(
    #                             f"http://{ho_ip}:{ho_port}/api/stock.move.line/create",
    #                             headers=headers_source, json=[move_line_vals]
    #                         )
    #                         move_resp.raise_for_status()
    #
    #                     except Exception as line_err:
    #                         msg = f"Move line creation failed for {order.name}: {line_err}"
    #                         ho.create_cmr_transaction_server_replication_log("failure", msg)
    #                         continue
    #
    #                 # -------------------- Success Marking --------------------
    #                 order.nhcl_replication_status = True
    #                 order.validate_orders(deliver_order='pos_order')
    #                 msg = f"Delivery Order successfully created for {order.name}"
    #                 ho.create_cmr_transaction_server_replication_log("success", msg)
    #
    #             except Exception as order_err:
    #                 # Any unexpected error per order
    #                 ho.create_cmr_transaction_server_replication_log("failure", str(order_err))
    #                 continue

    def get_pos_delivery_orders(self):

        import requests

        ho_ids = self.env['nhcl.ho.store.master'].search([
            ('nhcl_store_type', '=', 'ho'),
            ('nhcl_active', '=', True)
        ])

        for ho in ho_ids:

            session = requests.Session()

            base_url = f"http://{ho.nhcl_terminal_ip}:{ho.nhcl_port_no}/api"

            session.headers.update({
                'api-key': ho.nhcl_api_key,
                'Content-Type': 'application/json'
            })

            for order in self.filtered(
                    lambda x: x.location_dest_id.name == "Customers"
            ):

                try:

                    # -------------------------------------------------
                    # COMPANY
                    # -------------------------------------------------
                    company = session.get(
                        f"{base_url}/res.company/search",
                        params={
                            "domain": str([
                                ('name', '=', order.company_id.name)
                            ])
                        },
                        timeout=20
                    ).json().get("data", [])

                    if not company:
                        continue

                    company_id = company[0]['id']

                    # -------------------------------------------------
                    # PICKING TYPE
                    # -------------------------------------------------
                    picking_type = session.get(
                        f"{base_url}/stock.picking.type/search",
                        params={
                            "domain": str([
                                ('name', '=', order.picking_type_id.name),
                                ('company_id', '=', company_id)
                            ])
                        },
                        timeout=20
                    ).json().get("data", [])

                    if not picking_type:
                        continue

                    # -------------------------------------------------
                    # LOCATIONS
                    # -------------------------------------------------
                    source = session.get(
                        f"{base_url}/stock.location/search",
                        params={
                            "domain": str([
                                ('complete_name', '=', order.location_id.complete_name),
                                ('usage', '=', 'internal'),
                                ('company_id', '=', company_id)
                            ]),
                            "fields": "['id']"
                        },
                        timeout=20
                    ).json().get("data", [])

                    dest = session.get(
                        f"{base_url}/stock.location/search",
                        params={
                            "domain": str([
                                ('complete_name', '=', order.location_dest_id.complete_name),
                                ('usage', '=', 'customer')
                            ]),
                            "fields": "['id']"
                        },
                        timeout=20
                    ).json().get("data", [])

                    if not source or not dest:
                        continue

                    location_id = source[0]['id']
                    dest_id = dest[0]['id']

                    # -------------------------------------------------
                    # CHECK PICKING EXISTS
                    # -------------------------------------------------
                    existing = session.get(
                        f"{base_url}/stock.picking/search",
                        params={
                            "domain": str([
                                ('origin', '=', order.name)
                            ])
                        },
                        timeout=20
                    ).json().get("data", [])

                    if existing:
                        continue

                    # -------------------------------------------------
                    # PRODUCTS SINGLE FETCH
                    # -------------------------------------------------
                    nhcl_ids = list(set(
                        order.move_line_ids_without_package.mapped(
                            'product_id.nhcl_id'
                        )
                    ))

                    products = session.get(
                        f"{base_url}/product.product/search",
                        params={
                            "domain": str([
                                ('nhcl_id', 'in', nhcl_ids)
                            ]),
                            "fields": "['id','nhcl_id']"
                        },
                        timeout=60
                    ).json().get("data", [])

                    product_map = {
                        str(p['nhcl_id']): p['id']
                        for p in products
                    }

                    # -------------------------------------------------
                    # MOVE LINES
                    # -------------------------------------------------
                    move_lines = []

                    for line in order.move_line_ids_without_package:

                        product_id = product_map.get(
                            str(line.product_id.nhcl_id)
                        )

                        if not product_id:
                            continue

                        move_lines.append((0, 0, {
                            'product_id': product_id,
                            'quantity': line.quantity,
                            'location_id': location_id,
                            'location_dest_id': dest_id,
                            'lot_name': line.lot_id.name if line.lot_id else False,
                        }))

                    if not move_lines:
                        continue

                    # -------------------------------------------------
                    # SINGLE BULK CREATE
                    # -------------------------------------------------
                    payload = [{
                        'picking_type_id': picking_type[0]['id'],
                        'origin': order.name,
                        'location_id': location_id,
                        'location_dest_id': dest_id,
                        'company_id': company_id,
                        'move_type': 'direct',
                        'state': 'done',
                        'nhcl_store_delivery': True,
                        'move_line_ids_without_package': move_lines
                    }]

                    response = session.post(
                        f"{base_url}/stock.picking/create",
                        json=payload,
                        timeout=300
                    ).json()

                    if response.get("success"):

                        order.nhcl_replication_status = True

                        order.validate_orders(
                            deliver_order='pos_order'
                        )

                        ho.create_cmr_transaction_server_replication_log(
                            "success",
                            f"DO Created : {order.name}"
                        )

                    else:

                        ho.create_cmr_transaction_server_replication_log(
                            "failure",
                            response.get("message")
                        )

                except Exception as e:

                    ho.create_cmr_transaction_server_replication_log(
                        "failure",
                        str(e)
                    )

            session.close()

    @api.model
    def transfer_damage_main(self, barcode_list):
        context = {
            'restricted_picking_type_code': 'incoming',
            'search_default_damage_main': 1,
            'default_stock_picking_type': 'damage_main',
            'default_stock_type': 'ho_operation',
        }
        self = self.sudo().with_context(context).create({})
        if self:
            for rec in barcode_list:
                self.lot_qty = rec['qty']
                self.stock_barcode = rec['text']
                self._onchange_stock_barcode()
            self.action_confirm()
            self.button_validate()
            if self.state == 'done':
                self.get_damage_main_delivery_orders()
            return self.name
        return ''

    @api.model
    def transfer_return_main(self, barcode_list):
        context = {
            'restricted_picking_type_code': 'incoming',
            'search_default_return_main': 1,
            'default_stock_picking_type': 'return_main',
            'default_stock_type': 'ho_operation',
        }
        self = self.sudo().with_context(context).create({})
        if self:
            for rec in barcode_list:
                self.lot_qty = rec['qty']
                self.stock_barcode = rec['text']
                self._onchange_stock_barcode()
            self.action_confirm()
            self.button_validate()
            if self.state == 'done':
                self.get_return_main_delivery_orders()
            return self.name
        return ''


class PickingType(models.Model):
    _inherit = "stock.picking.type"

    stock_picking_type = fields.Selection([('exchange', 'Customer-Return'), ('receipt', 'Receipt')
                                              , ('delivery', 'Delivery'), ('pos_order', 'POS Order'),
                                           ('manufacturing', 'Manufacturing'),
                                           ('regular', 'Regular'), ('damage', 'Damage'), ('return', 'Return'),
                                           ('damage_main', 'Damage-Main'), ('main_damage', 'Main-Damage'),
                                           ('return_main', 'Return-Main'),
                                           ('hpo', 'Hired Product Outward'), ('hpi', 'Hired Product Inward')],
                                          string='Type')


class StockMove(models.Model):
    """Inherited stock.move class to add fields and functions"""
    _inherit = "stock.move"

    serial_no = fields.Char(string="Serial No")
    dummy_lot_ids = fields.Many2many('stock.lot', string="Ref S.No")
    type_product = fields.Selection([('brand', 'Brand'), ('un_brand', 'UnBrand'), ('others', 'Others')],
                                    string='Brand Type', copy=False)
    pos_order_lines = fields.Many2one('pos.order.line', string='pos order lines', copy=False)
    nhcl_tax_ids = fields.Many2many('account.tax', 'nhcl_tax',
                                    domain=[('type_tax_use', '=', 'sale'), ('active', '=', True)],
                                    string="Tax")
    nhcl_total = fields.Float(string="Total", copy=False, digits=(16, 2))
    nhcl_rsp = fields.Float(string="RSP", copy=False, digits=(16, 2))
    nhcl_exchange = fields.Boolean(string="Exchange", copy=False)
    nhcl_discount = fields.Float(string="Discount (%)", copy=False, digits=(16, 2))
    nhcl_gdiscount = fields.Float(string="Global Discount (%)", copy=False, digits=(16, 2))
    nhcl_disc_lines = fields.Char(string="Disc", copy=False)
    nhcl_price_total = fields.Monetary(compute='_compute_amount', string='Total', store=True)
    nhcl_price_subtotal = fields.Monetary(compute='_compute_amount', string='Subtotal', store=True)
    nhcl_price_tax = fields.Float(compute='_compute_amount', string='Tax', store=True, digits=(16, 2))
    currency_id = fields.Many2one("res.currency", string="Currency", required=True,
                                  related='picking_id.currency_id')
    ref_pos_order_line_id = fields.Integer('Pos Order Line Id', default="0", copy=False)
    move_picking_type = fields.Selection(related='picking_id.stock_picking_type', string='Picking Type')
    s_no = fields.Integer(string="S.No", compute="_compute_s_no")
    nhcl_cost_price = fields.Float(string="Cost Price", copy=False, digits=(16, 2))
    nhcl_on_cart_discount_value = fields.Float(string="On Cart Disc Value", copy=False, digits=(16, 2))
    nhcl_on_cart_without_discount = fields.Boolean(string="On Cart Without Disc Value", copy=False)
    nhcl_rs_price = fields.Float(string="RS Price", copy=False)
    nhcl_mr_price = fields.Float(string="MR Price", copy=False)
    tracking = fields.Selection(
        related='product_id.tracking',
        string="Tracking", store=True)
    nhcl_move_qty = fields.Float(
        'Enter Quantity', digits='Product Unit of Measure', copy=False)
    nhcl_old_qty = fields.Float(
        'Original Quantity', digits='Product Unit of Measure', copy=False)
    move_brand_barcode = fields.Char(string="CMR Barcode", copy=False)
    move_cp = fields.Float(string="CMR CP", copy=False)
    move_mrp = fields.Float(string="CMR MRP", copy=False)
    move_rsp = fields.Float(string="CMR RSP", copy=False)
    nhcl_is_fix_discount_line = fields.Boolean(related="pos_order_lines.is_fix_discount_line",
                                               string='Is Fix Discounted Line', store=True)


    # @api.constrains('nhcl_move_qty')
    def change_exchange_move_qty(self):
        for rec in self:
            if rec.nhcl_move_qty and (rec.nhcl_move_qty > rec.quantity) and rec.product_id.tracking == 'lot':
                raise ValidationError("Entered quantity cannot be greater than the purchase quantity.")
            if rec.nhcl_move_qty:
                rec.product_uom_qty = rec.nhcl_move_qty
                rec.quantity = rec.nhcl_move_qty

    @api.depends('picking_id')
    def _compute_s_no(self):
        for rec in self:
            if rec.picking_id and rec.id in rec.picking_id.move_ids_without_package.ids:
                rec.s_no = rec.picking_id.move_ids_without_package.ids.index(rec.id) + 1
            else:
                rec.s_no = 0

    def _get_new_picking_values(self):
        res = super(StockMove, self)._get_new_picking_values()
        if res.get('origin'):
            sale_id = self.env['sale.order'].search([('name', '=', res.get('origin'))], limit=1)
            stock_operation_type = self.env['stock.picking.type'].search(
                [('stock_picking_type', '=', sale_id.transfer_type)], limit=1)
            if sale_id:
                res.update({
                    'stock_type': sale_id.so_type,
                    'stock_picking_type': sale_id.transfer_type,
                })
                if stock_operation_type:
                    res.update({
                        'picking_type_id': stock_operation_type.id,
                        'location_id': stock_operation_type.default_location_src_id.id

                    })
            for move in self:
                if stock_operation_type:
                    move.write({
                        'location_id': stock_operation_type.default_location_src_id.id

                    })
        return res

    def picking_unlink(self):
        for move in self:
            move.unlink()

    @api.model
    def _prepare_merge_moves_distinct_fields(self):
        distinct_fields = super(StockMove, self)._prepare_merge_moves_distinct_fields()
        distinct_fields.append('pos_order_lines')
        distinct_fields.append('ref_pos_order_line_id')
        distinct_fields.append('dummy_lot_ids')
        return distinct_fields

    @api.depends('quantity', 'nhcl_rsp', 'nhcl_tax_ids')
    def _compute_amount(self):
        for line in self:
            tax_results = self.env['account.tax']._compute_taxes([line._convert_to_tax_base_line_dict()])
            totals = next(iter(tax_results['totals'].values()))
            amount_untaxed = round(totals['amount_untaxed'], 2)
            amount_tax = round(totals['amount_tax'], 2)
            line.update({
                'nhcl_price_subtotal': amount_untaxed,
                'nhcl_price_tax': amount_tax,
                'nhcl_price_total': amount_untaxed + amount_tax,
            })

    # updating the price unit,currency,req qty,product,partner
    def _convert_to_tax_base_line_dict(self):
        # Hook method to returns the different argument values for the
        # compute_all method, due to the fact that discounts mechanism
        # is not implemented yet on the purchase orders.
        # This method should disappear as soon as this feature is
        # also introduced like in the sales module.
        self.ensure_one()
        price_unit = round(self.nhcl_rsp * (1 - (self.nhcl_discount or 0.0) / 100.0) * (
                1 - (self.nhcl_gdiscount or 0.0) / 100.0), 2)
        # print(price_unit, "price_unit")
        return self.env['account.tax']._convert_to_tax_base_line_dict(
            self,
            price_unit=price_unit,
            currency=self.picking_id.currency_id,
            quantity=self.quantity,
            product=self.product_id,
            taxes=self.nhcl_tax_ids,
            partner=self.picking_id.partner_id,
            price_subtotal=round(self.nhcl_price_subtotal, 2),
        )

    @api.onchange('product_id')
    def _onchange_product_id_stock(self):
        if self.picking_id and not self.picking_id.stock_type:
            # Clear the product_id and raise an error if no stock_type is selected
            self.product_id = False
            raise ValidationError(
                "You must select a Stock Type before selecting a product."
            )

    def action_assign_serial(self):
        if self.picking_id.picking_type_id.code == 'incoming' and self.picking_id.stock_type in ['ho_operation',
                                                                                                 'data_import']:
            raise ValidationError(
                _("You are not allowed to Assign the Serial Number For Product %s") % (self.product_id.name))
        else:
            return super(StockMove, self).action_assign_serial()

    def _update_reserved_quantity(self, need, location_id, lot_id=None, quant_ids=None, package_id=None, owner_id=None,
                                  strict=True):
        if self._context.get("sol_lot_id"):
            # Use sale line's lots if available, otherwise fallback to dummy lots
            if self.sale_line_id and self.sale_line_id.lot_ids:
                lot_id = self.sale_line_id.lot_ids
            elif self.dummy_lot_ids:
                lot_id = self.dummy_lot_ids
        return super()._update_reserved_quantity(
            need, location_id,
            lot_id=lot_id,
            quant_ids=quant_ids,
            package_id=package_id,
            owner_id=owner_id,
            strict=strict
        )

    def auto_generate_serial_numbers(self):
        MasterSeq = self.env['nhcl.master.sequence']
        auto_seq = MasterSeq.search(
            [('nhcl_code', '=', 'Auto Serial Number'),
             ('nhcl_state', '=', 'activate')],
            limit=1
        )

        if not auto_seq:
            return

        start_num = auto_seq.nhcl_next_number
        for move in self:
            pending_lines = move.move_line_ids.filtered(
                lambda l: not l.lot_name
            )
            for idx, line in enumerate(pending_lines):
                lot_name = f"{auto_seq.nhcl_prefix}{start_num}"
                line.lot_name = lot_name
                start_num += 1
        auto_seq.nhcl_next_number = start_num

    def _action_done(self, cancel_backorder=False):
        for move in self:
            if move.product_id.tracking in ['serial', 'lot'] and move.picking_id.stock_type == 'data_import':
                if float(move.quantity) == sum(
                        move.move_line_ids.filtered(lambda x: x.lot_name == False).mapped('quantity')):
                    auto_generate = self.env['nhcl.master.sequence'].search(
                        [('nhcl_code', '=', 'Auto Serial Number'), ('nhcl_state', '=', 'activate')])
                    if auto_generate:
                        move.auto_generate_serial_numbers()
                    else:
                        raise ValidationError("Not cofigured Auto Serial Number.")
            elif move.picking_id.stock_picking_type == 'exchange':
                move.change_exchange_move_qty()
        return super()._action_done(cancel_backorder=cancel_backorder)

    def _prepare_move_line_vals(self, quantity=None, reserved_quant=None):
        vals = super()._prepare_move_line_vals(quantity=quantity, reserved_quant=reserved_quant)
        sale_line = self.sale_line_id
        if sale_line:
            if self.product_id.tracking == 'lot' and sale_line.lot_ids:
                vals['lot_id'] = sale_line.lot_ids.id
            elif self.product_id.tracking == 'serial' and sale_line.lot_ids:
                vals['lot_id'] = sale_line.lot_ids[0].id
        if self.serial_no and self.picking_id.stock_picking_type == 'exchange' and self.picking_id.company_type == 'same':
            vals['lot_name'] = self.serial_no
            if self.dummy_lot_ids:
                lot = self.dummy_lot_ids[0]
                vals['cost_price'] = lot.cost_price
        if self.picking_id.move_type == 'direct' and self.dummy_lot_ids:
            vals['lot_id'] = self.dummy_lot_ids.id
        cat1 = self.product_id.categ_id
        # walk the parent chain safely
        p11 = cat1.parent_id
        p22 = p11.parent_id if p11 else False
        p33 = p22.parent_id if p22 else False
        vals.update({
            'brick': cat1.id,
            'class_level_id': p11.id if p11 else False,
            'category': p22.id if p22 else False,
            'family': p33.id if p33 else False,
        })
        return vals


class StockMoveLine(models.Model):
    _inherit = "stock.move.line"

    internal_ref_lot = fields.Char(string="Barcode", copy=False)

    type_product = fields.Selection([('brand', 'Brand'), ('un_brand', 'UnBrand'), ('others', 'Others')],
                                    string='Brand Type', copy=False)
    categ_1 = fields.Many2one('product.attribute.value', string='Color', copy=False,
                              domain=[('attribute_id.name', '=', 'Color')])
    categ_2 = fields.Many2one('product.attribute.value', string='Fit', copy=False,
                              domain=[('attribute_id.name', '=', 'Fit')])
    categ_3 = fields.Many2one('product.attribute.value', string='Brand', copy=False,
                              domain=[('attribute_id.name', '=', 'Brand')])
    categ_4 = fields.Many2one('product.attribute.value', string='Pattern', copy=False,
                              domain=[('attribute_id.name', '=', 'Pattern')])
    categ_5 = fields.Many2one('product.attribute.value', string='Border Type', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Type')])
    categ_6 = fields.Many2one('product.attribute.value', string='Border Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Size')])
    categ_7 = fields.Many2one('product.attribute.value', string='Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Size')])
    categ_8 = fields.Many2one('product.attribute.value', string='Design', copy=False,
                              domain=[('attribute_id.name', '=', 'Design')])
    descrip_1 = fields.Many2one('product.aging.line', string="Product Aging", copy=False)
    descrip_2 = fields.Many2one('product.attribute.value', string='Range', copy=False,
                                domain=[('attribute_id.name', '=', 'Range')])
    descrip_3 = fields.Many2one('product.attribute.value', string='Collection', copy=False,
                                domain=[('attribute_id.name', '=', 'Collection')])
    descrip_4 = fields.Many2one('product.attribute.value', string='Fabric', copy=False,
                                domain=[('attribute_id.name', '=', 'Fabric')])
    descrip_5 = fields.Many2one('product.attribute.value', string='Exclusive', copy=False,
                                domain=[('attribute_id.name', '=', 'Exclusive')])
    descrip_6 = fields.Many2one('product.attribute.value', string='Print', copy=False,
                                domain=[('attribute_id.name', '=', 'Print')])
    descrip_7 = fields.Many2one('product.attribute.value', string='Days Ageing', copy=False,
                                domain=[('attribute_id.name', '=', 'Days Ageing')])
    descrip_8 = fields.Many2one('product.attribute.value', string='Description 8', copy=False)
    descrip_9 = fields.Many2one('product.attribute.value', string='Discount', copy=False,
                                domain=[('attribute_id.name', '=', 'Discount')])
    cost_price = fields.Float(string='CP', copy=False)
    mr_price = fields.Float(string='MRP', copy=False)
    rs_price = fields.Float(string='RSP', copy=False)
    segment = fields.Selection([('apparel', 'Apparel'), ('non_apparel', 'Non Apparel'), ('others', 'Others')],
                               string="Segment", copy=False, store=True)
    move_line_picking_type = fields.Selection(related='picking_id.stock_picking_type', string='Picking Type')

    s_no = fields.Integer(string="S.No", compute="_compute_s_no")
    family = fields.Many2one('product.category', string="Family", domain="[('parent_id','=',False)]")
    category = fields.Many2one('product.category', string="Category",
                               domain="[('parent_id','=',family)]")
    class_level_id = fields.Many2one('product.category', string="Class",
                                     domain="[('parent_id','=',category)]")
    brick = fields.Many2one('product.category', string="Brick",
                            domain="[('parent_id','=',class_level_id)]")

    nhcl_purchase_indent_reference = fields.Char(string="NHCL Purchase Indent Reference", copy=False)
    nhcl_lot_hsn_code = fields.Char(string="HSN Code", copy=False)
    sale_tax_ids = fields.Many2many('account.tax', string="Sale Taxes", copy=False)
    purchase_tax_ids = fields.Many2many('account.tax', 'purch_move_line_tax', string="Purchase Taxes", copy=False)

    @api.depends('picking_id')
    def get_from_info(self):
        for move_line in self:
            value = False
            picking = move_line.picking_id
            if picking and picking.stock_picking_type == 'pos_order':
                value = "Main"
            elif picking and picking.stock_picking_type == 'receipt':
                value = "Vendor"
            elif picking and picking.stock_picking_type == 'exchange':
                value = "Customer"
            elif picking and picking.stock_picking_type == 'damage_main':
                value = "Damage"
            elif picking and picking.stock_picking_type == 'main_damage':
                value = "Main"
            elif picking and picking.stock_picking_type == 'return_main':
                value = "Return"
            elif picking and picking.stock_picking_type == 'damage':
                value = "Damage"
            elif picking and picking.stock_picking_type == 'return':
                value = "Main"
            move_line.nhcl_from = value

    @api.depends('picking_id')
    def get_to_info(self):
        for move_line in self:
            value = False
            picking = move_line.picking_id
            if picking and picking.stock_picking_type == 'pos_order':
                value = "Customer"
            elif picking and picking.stock_picking_type == 'receipt':
                value = "Main"
            elif picking and picking.stock_picking_type == 'exchange':
                value = "Return"
            elif picking and picking.stock_picking_type == 'damage_main':
                value = "Main"
            elif picking and picking.stock_picking_type == 'main_damage':
                value = "Damage"
            elif picking and picking.stock_picking_type == 'return_main':
                value = "Main"
            elif picking and picking.stock_picking_type == 'damage':
                value = "Vendor"
            elif picking and picking.stock_picking_type == 'return':
                value = "Vendor"
            move_line.nhcl_to = value

    nhcl_from = fields.Char(string="From", compute='get_from_info')
    nhcl_to = fields.Char(string="To", compute='get_to_info')

    @api.depends('batch_id')
    def _compute_s_no(self):
        for rec in self:
            if rec.batch_id and rec.id in rec.batch_id.move_line_ids.ids:
                rec.s_no = rec.batch_id.move_line_ids.ids.index(rec.id) + 1
            else:
                rec.s_no = 0

    def compute_get_unit_price(self):
        for rec in self:
            if rec.picking_id.picking_type_id.stock_picking_type != 'receipt':
                if rec.lot_id:
                    rec.cost_price = rec.lot_id.cost_price
                    rec.internal_ref_lot = rec.lot_id.ref
                    rec.type_product = rec.lot_id.type_product
                    rec.mr_price = rec.lot_id.mr_price
                    rec.rs_price = rec.lot_id.rs_price
                    rec.segment = rec.lot_id.segment
                    rec.categ_1 = rec.lot_id.category_1
                    rec.categ_2 = rec.lot_id.category_2
                    rec.categ_3 = rec.lot_id.category_3
                    rec.categ_4 = rec.lot_id.category_4
                    rec.categ_5 = rec.lot_id.category_5
                    rec.categ_6 = rec.lot_id.category_6
                    rec.categ_7 = rec.lot_id.category_7
                    rec.categ_8 = rec.lot_id.category_8
                    rec.descrip_1 = rec.lot_id.description_1
                    rec.descrip_2 = rec.lot_id.description_2
                    rec.descrip_3 = rec.lot_id.description_3
                    rec.descrip_4 = rec.lot_id.description_4
                    rec.descrip_5 = rec.lot_id.description_5
                    rec.descrip_6 = rec.lot_id.description_6
                    rec.descrip_7 = rec.lot_id.description_7
                    rec.descrip_8 = rec.lot_id.description_8
                else:
                    rec.cost_price = 0.0

    def get_product_attributes(self):
        for rec in self:
            val = rec.product_id.product_template_attribute_value_ids
            for i in val:
                attribute = self.env['product.attribute.value'].search([('name', '=', i.name)])
                for j in attribute:
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Color'):
                        rec.categ_1 = j.id

    @api.model_create_multi
    def create(self, vals_list):
        records = super(StockMoveLine, self).create(vals_list)

        for record in records:
            record.get_product_attributes()
            record.compute_get_unit_price()

            if (
                    record.move_id and
                    record.move_id.product_id.id == record.product_id.id and
                    record.picking_id.stock_picking_type == 'receipt'
            ):
                record.move_id.type_product = record.type_product

        return records

    def hsn_values_update(self):
        for rec in self:
            rec.lot_id.nhcl_lot_hsn_code = rec.nhcl_lot_hsn_code
            rec.lot_id.sale_tax_ids = rec.sale_tax_ids.ids
            rec.lot_id.purchase_tax_ids = rec.purchase_tax_ids.ids



    def write(self, vals):
        res = super(StockMoveLine, self).write(vals)
        for move_line in self:
            if move_line.lot_id:
                lot_values = {}
                for categ_field in ['categ_1', 'categ_2', 'categ_3', 'categ_4', 'categ_5', 'categ_6', 'categ_7',
                                    'categ_8']:
                    if categ_field in vals:
                        lot_values[categ_field.replace('categ', 'category')] = vals[categ_field]
                for desc_field in ['descrip_1', 'descrip_2', 'descrip_3', 'descrip_4', 'descrip_5', 'descrip_6',
                                   'descrip_7', 'descrip_8']:
                    if desc_field in vals:
                        lot_values[desc_field.replace('descrip', 'description')] = vals[desc_field]
                for price_field in ['cost_price', 'mr_price', 'rs_price']:
                    if price_field in vals:
                        lot_values[price_field] = vals[price_field]
                if 'internal_ref_lot' in vals:
                    lot_values['ref'] = vals['internal_ref_lot']
                if 'nhcl_purchase_indent_reference' in vals:
                    lot_values['nhcl_purchase_indent_referenace'] = vals['nhcl_purchase_indent_reference']
                if lot_values:
                    move_line.lot_id.write(lot_values)
        return res

    def _create_and_assign_production_lot(self):
        res = super(StockMoveLine, self)._create_and_assign_production_lot()
        if self.picking_id.picking_type_id.code == 'incoming':
            for rec in self:
                rec.lot_id.write({
                    'type_product': rec.type_product,
                    'picking_id': rec.picking_id.id,
                    'ref': rec.internal_ref_lot,
                    'nhcl_lot_hsn_code': rec.nhcl_lot_hsn_code,
                    'sale_tax_ids': rec.sale_tax_ids.ids,
                    'purchase_tax_ids': rec.purchase_tax_ids.ids,
                    'category_1': rec.categ_1,
                    'category_2': rec.categ_2,
                    'category_3': rec.categ_3,
                    'category_4': rec.categ_4,
                    'category_5': rec.categ_5,
                    'category_6': rec.categ_6,
                    'category_7': rec.categ_7,
                    'category_8': rec.categ_8,
                    'description_1': rec.descrip_1,
                    'description_2': rec.descrip_2,
                    'description_3': rec.descrip_3,
                    'description_4': rec.descrip_4,
                    'description_5': rec.descrip_5,
                    'description_6': rec.descrip_6,
                    'description_7': rec.descrip_7,
                    'description_8': rec.descrip_8,
                    'mr_price': rec.mr_price,
                    'cost_price': rec.cost_price,
                    'rs_price': rec.rs_price,
                    # 'product_aging': rec.product_aging,
                    'segment': rec.segment,
                    'nhcl_purchase_indent_reference': rec.nhcl_purchase_indent_reference if rec.nhcl_purchase_indent_reference else None,
                    # 'transfer_price': rec.transfer_price,
                    # 'transfer_percent': rec.transfer_percent,
                })
                if rec.type_product == 'brand':
                    # Search for existing barcode
                    existing_barcode = self.env['product.barcode'].sudo().search(
                        [('barcode', '=', rec.internal_ref_lot)],
                        limit=1)

                    if rec.picking_id.picking_type_id.code == 'incoming':  # If the picking code is for incoming transfers (receipts)
                        if existing_barcode:
                            # Increment nhcl_inward_qty for existing barcode
                            existing_barcode.sudo().write({'nhcl_inward_qty': existing_barcode.nhcl_inward_qty + 1})
                        else:
                            # Create a new barcode and set nhcl_inward_qty to 1
                            p = self.env['product.barcode'].search([], order='id desc', limit=1).id
                            print(p,'iudshfilsknfmlkfmd;zl')
                            rec.lot_id.product_id.sudo().product_barcode = [(0, 0, {
                                'barcode': rec.internal_ref_lot,
                                'nhcl_inward_qty': 1,
                            })]

                    elif rec.picking_id.picking_type_id.code == 'outgoing':  # If the picking code is for outgoing transfers (deliveries)
                        if existing_barcode:
                            # Increment nhcl_outward_qty for existing barcode
                            existing_barcode.sudo().write({'nhcl_outward_qty': existing_barcode.nhcl_outward_qty + 1})
                        else:
                            # Create a new barcode and set nhcl_outward_qty to 1
                            rec.lot_id.product_id.sudo().product_barcode = [(0, 0, {
                                'barcode': rec.internal_ref_lot,
                                'nhcl_outward_qty': 1,
                            })]
        return res

    @api.onchange('internal_ref_lot')
    def sending_no_to_lot(self):
        for rec in self:
            if rec.lot_id and rec.internal_ref_lot:
                # Set the reference on the lot
                # rec.lot_id.ref = rec.internal_ref_lot
                # Search for existing barcode
                existing_barcode = self.env['product.barcode'].sudo().search([('barcode', '=', rec.internal_ref_lot)],
                                                                             limit=1)

                if rec.picking_id.picking_type_id.code == 'incoming':  # If the picking code is for incoming transfers (receipts)
                    if existing_barcode:
                        # Increment nhcl_inward_qty for existing barcode
                        existing_barcode.sudo().write({'nhcl_inward_qty': existing_barcode.nhcl_inward_qty + 1})
                    else:
                        # Create a new barcode and set nhcl_inward_qty to 1
                        rec.lot_id.product_id.sudo().product_barcode = [(0, 0, {
                            'barcode': rec.internal_ref_lot,
                            'nhcl_inward_qty': 1,
                        })]

                elif rec.picking_id.picking_type_id.code == 'outgoing':  # If the picking code is for outgoing transfers (deliveries)
                    if existing_barcode:
                        # Increment nhcl_outward_qty for existing barcode
                        existing_barcode.sudo().write({'nhcl_outward_qty': existing_barcode.nhcl_outward_qty + 1})
                    else:
                        # Create a new barcode and set nhcl_outward_qty to 1
                        rec.lot_id.product_id.sudo().product_barcode = [(0, 0, {
                            'barcode': rec.internal_ref_lot,
                            'nhcl_outward_qty': 1,
                        })]


class StockLot(models.Model):
    """Inherited stock.lot class to add fields and functions"""
    _inherit = 'stock.lot'

    category_1 = fields.Many2one('product.attribute.value', string='Color', copy=False,
                                 domain=[('attribute_id.name', '=', 'Color')])
    category_2 = fields.Many2one('product.attribute.value', string='Fit', copy=False,
                                 domain=[('attribute_id.name', '=', 'Fit')])
    category_3 = fields.Many2one('product.attribute.value', string='Brand', copy=False,
                                 domain=[('attribute_id.name', '=', 'Brand')])
    category_4 = fields.Many2one('product.attribute.value', string='Pattern', copy=False,
                                 domain=[('attribute_id.name', '=', 'Pattern')])
    category_5 = fields.Many2one('product.attribute.value', string='Border Type', copy=False,
                                 domain=[('attribute_id.name', '=', 'Border Type')])
    category_6 = fields.Many2one('product.attribute.value', string='Border Size', copy=False,
                                 domain=[('attribute_id.name', '=', 'Border Size')])
    category_7 = fields.Many2one('product.attribute.value', string='Size', copy=False,
                                 domain=[('attribute_id.name', '=', 'Size')])
    category_8 = fields.Many2one('product.attribute.value', string='Design', copy=False,
                                 domain=[('attribute_id.name', '=', 'Design')])
    description_1 = fields.Many2one('product.aging.line', string="Product Aging", copy=False)
    description_2 = fields.Many2one('product.attribute.value', string='Range', copy=False,
                                    domain=[('attribute_id.name', '=', 'Range')])
    description_3 = fields.Many2one('product.attribute.value', string='Collection', copy=False,
                                    domain=[('attribute_id.name', '=', 'Collection')])
    description_4 = fields.Many2one('product.attribute.value', string='Fabric', copy=False,
                                    domain=[('attribute_id.name', '=', 'Fabric')])
    description_5 = fields.Many2one('product.attribute.value', string='Exclusive', copy=False,
                                    domain=[('attribute_id.name', '=', 'Exclusive')])
    description_6 = fields.Many2one('product.attribute.value', string='Print', copy=False,
                                    domain=[('attribute_id.name', '=', 'Print')])
    description_7 = fields.Many2one('product.attribute.value', string='Days Ageing', copy=False,
                                    domain=[('attribute_id.name', '=', 'Days Ageing')])
    description_8 = fields.Many2one('product.attribute.value', string='Description 8', copy=False)
    description_9 = fields.Many2one('product.attribute.value', string='Discount', copy=False,
                                    domain=[('attribute_id.name', '=', 'Discount')])
    product_description = fields.Html(string="Product Description", copy=False)
    web_product = fields.Char(string="Website Product Name", copy=False)
    cost_price = fields.Float(string='CP', copy=False)
    mr_price = fields.Float(string='MRP', copy=False)
    rs_price = fields.Float(string='RSP', copy=False)
    # transfer_price = fields.Float(string='TRP', copy=False)
    type_product = fields.Selection([('brand', 'Brand'), ('un_brand', 'UnBrand'), ('others', 'Others')],
                                    string='Brand Type', copy=False)
    picking_id = fields.Many2one('stock.picking', string="GRC No", copy=False)
    # transfer_percent = fields.Float(string="TP %", copy=False)
    is_used = fields.Boolean(string='POS Posted')
    # product_aging = fields.Many2one('product.aging.line', string="Product Aging", copy=False)
    segment = fields.Selection([('apparel', 'Apparel'), ('non_apparel', 'Non Apparel'), ('others', 'Others')],
                               string="Segment", copy=False, tracking=True, store=True)
    ref = fields.Char('Barcode',
                      help="Internal reference number in case it differs from the manufacturer's lot/serial number")
    ho_grc_no = fields.Char(string='HO GRC NO.')
    serial_type = fields.Selection([('regular', 'Regular'), ('return', 'Returned')],
                                   string='Serial Type', copy=False, tracking=True, default='regular')
    is_uploaded = fields.Boolean('Is Uploaded', copy=False)
    hired_product = fields.Boolean(string="Hired Product")
    nhcl_purchase_indent_reference = fields.Char(string="NHCL Purchase Indent Reference", copy=False)
    nhcl_lot_hsn_code = fields.Char(string="HSN Code", copy=False)
    sale_tax_ids = fields.Many2many('account.tax', string="Sale Taxes", copy=False)
    purchase_tax_ids = fields.Many2many('account.tax','purch_tax', string="Purchase Taxes", copy=False)

    def _compute_traceability_line_ids(self):
        TraceReport = self.env["stock.traceability.report"]
        for lot in self:
            context = {
                "active_id": lot.id,
                "model": "stock.lot",
            }

            # Get full traceability structure (inward + outward)
            roots = TraceReport.with_context(context).get_lines() or []

            visited = set()
            all_lines = self.env["stock.move.line"]

            # Prepare BFS queue with all initial lines
            queue = []
            for r in roots:
                line = self.env[r["model"]].browse(r["model_id"])
                if line and line.id not in visited:
                    visited.add(line.id)
                    all_lines |= line
                    queue.append(line)

            # BFS traversal through entire chain
            while queue:
                move_line = queue.pop(0)

                # 1. Get parents/children linked via stock.traceability.report
                linked_lines, is_used = TraceReport._get_linked_move_lines(move_line)
                for l in linked_lines:
                    if l.id not in visited:
                        visited.add(l.id)
                        all_lines |= l
                        queue.append(l)

                    # 2. For each linked line, also fetch its move lines
                    move_lines = TraceReport._get_move_lines(l)
                    for ml in move_lines:
                        if ml.id not in visited:
                            visited.add(ml.id)
                            all_lines |= ml
                            queue.append(ml)

            lot.traceability_line_ids = all_lines

    traceability_line_ids = fields.Many2many(
        "stock.move.line", "Traceability Lines", compute=_compute_traceability_line_ids
    )

    def action_traceability_list(self):
        tree_view_id = self.env.ref("stock.view_move_line_tree").id
        form_view_id = self.env.ref("stock.view_move_line_form").id
        domain = [("id", "in", self.traceability_line_ids.ids)]
        action = {
            "type": "ir.actions.act_window",
            "views": [(tree_view_id, "tree"), (form_view_id, "form")],
            "view_mode": "tree,form",
            "name": _("Traceability List"),
            "res_model": "stock.move.line",
            "domain": domain,
            "context": {"create": False, "edit": False, "delete": False},
        }
        return action

    def action_traceability_dynamic(self):
        """Open the Traceability action (ID 638) for this specific lot"""
        self.ensure_one()
        try:
            action = self.env.ref('stock.action_stock_report').read()[0]
        except ValueError:
            action = self.env['ir.actions.act_window'].browse(638).read()[0]

        action['context'] = dict(self.env.context or {}, search_default_lot_id=self.id)
        return action

    @api.onchange('category_1')
    def updating_line_to_lot_category_1(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_1 = self.category_1

    @api.onchange('category_2')
    def updating_line_to_lot_category_2(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_2 = self.category_2

    @api.onchange('category_3')
    def updating_line_to_lot_category_3(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_3 = self.category_3

    @api.onchange('category_4')
    def updating_line_to_lot_category_4(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_4 = self.category_4

    @api.onchange('category_5')
    def updating_line_to_lot_category_5(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_5 = self.category_5

    @api.onchange('category_6')
    def updating_line_to_lot_category_6(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_6 = self.category_6

    @api.onchange('category_7')
    def updating_line_to_lot_category_7(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_7 = self.category_7

    @api.onchange('category_8')
    def updating_line_to_lot_category_8(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_8 = self.category_8

    @api.onchange('description_1')
    def updating_line_to_lot_description_1(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_1 = self.description_1

    @api.onchange('description_2')
    def updating_line_to_lot_description_2(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_2 = self.description_2

    @api.onchange('description_3')
    def updating_line_to_lot_description_3(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_3 = self.description_3

    @api.onchange('description_4')
    def updating_line_to_lot_description_4(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_4 = self.description_4

    @api.onchange('description_5')
    def updating_line_to_lot_description_5(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_5 = self.description_5

    @api.onchange('description_6')
    def updating_line_to_lot_description_6(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_6 = self.description_6

    @api.onchange('description_7')
    def updating_line_to_lot_description_7(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_7 = self.description_7

    @api.onchange('description_8')
    def updating_line_to_lot_description_8(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_8 = self.description_8

    @api.onchange('mr_price')
    def updating_line_to_lot_mr_price(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.mr_price = self.mr_price

    @api.onchange('rs_price')
    def updating_line_to_lot_rs_price(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.rs_price = self.rs_price

    # @api.onchange('transfer_price')
    # def updating_line_to_lot_transfer_price(self):
    # detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
    # detail.transfer_price = self.transfer_price

    @api.onchange('transfer_percent')
    def updating_line_to_lot_transfer_percent(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.transfer_percent = self.transfer_percent

    @api.onchange('ref')
    def updating_line_to_lot_ref(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.internal_ref_lot = self.ref

    def get_attributes(self):
        for rec in self:
            val = rec.product_id.product_template_attribute_value_ids
            for i in val:
                attribute = self.env['product.attribute.value'].search([('name', '=', i.name)])
                for j in attribute:
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith(
                            'Color'):
                        rec.category_1 = j.id

    @api.model_create_multi
    def create(self, vals_list):
        records = super(StockLot, self).create(vals_list)

        for record in records:
            record.get_attributes()

        return records

    def search_by_loyalty_rule(self, loyalty_rule):
        domain = []
        today = datetime.today()

        # Get the user's timezone
        user_tz = self.env.user.tz or pytz.utc
        local = pytz.timezone(user_tz)

        # Get the selected ageing slab range
        if loyalty_rule.day_ageing_slab:
            # Mapping for ageing slabs
            slab_mapping = {
                '1': (0, 30),
                '2': (30, 60),
                '3': (60, 90),
                '4': (90, 120),
                '5': (120, 150),
                '6': (150, 180),
                '7': (180, 210),
                '8': (210, 240),
                '9': (240, 270),
                '10': (270, 300),
                '11': (300, 330),
                '12': (330, 360)
            }
            # Get the start and end days for the slab
            slab_start, slab_end = slab_mapping.get(str(loyalty_rule.day_ageing_slab), (0, 360))

            # Calculate the lower and upper bounds for the matching date range
            ageing_date_start = today - timedelta(days=slab_end)
            ageing_date_end = today - timedelta(days=slab_start)

            # Ensure the start date is earlier than the end date
            if ageing_date_start > ageing_date_end:
                ageing_date_start, ageing_date_end = ageing_date_end, ageing_date_start

            # Localize the dates to the user's timezone
            from_date_local = ageing_date_start.replace(hour=0, minute=0, second=0, microsecond=0)
            to_date_local = ageing_date_end.replace(hour=23, minute=59, second=59, microsecond=999999)

            from_date_local = local.localize(from_date_local)
            to_date_local = local.localize(to_date_local)

            # Convert the localized dates to UTC
            from_date_utc = from_date_local.astimezone(pytz.utc)
            to_date_utc = to_date_local.astimezone(pytz.utc)

            # Format the dates into ISO 8601 format
            from_date_str = from_date_utc.strftime("%Y-%m-%dT%H:%M:%S")
            to_date_str = to_date_utc.strftime("%Y-%m-%dT%H:%M:%S")

            # Add create_date range condition to the domain
            domain.append(('create_date', '>=', from_date_str))
            domain.append(('create_date', '<=', to_date_str))
        # Loop through all category and description fields
        for i in range(1, 8):
            # Dynamically construct field names
            category_field = f'category_{i}'
            description_field = f'description_{i}'
            # Get the corresponding many2many fields in loyalty.rule
            category_rule_field = f'category_{i}_ids'
            description_rule_field = f'description_{i}_ids' if i != 7 else None

            # Add to domain if the loyalty rule fields have values
            if getattr(loyalty_rule, category_rule_field):
                domain.append((category_field, 'in', getattr(loyalty_rule, category_rule_field).ids))
            if description_rule_field and getattr(loyalty_rule, description_rule_field):
                domain.append((description_field, 'in', getattr(loyalty_rule, description_rule_field).ids))

        # Add product filtering if ref_product_ids is defined in the loyalty rule
        if loyalty_rule.ref_product_ids:
            domain.append(('product_id', 'in', loyalty_rule.ref_product_ids.ids))
        # Add category filtering if product_category_id is defined in the loyalty rule
        if loyalty_rule.product_category_ids:
            selected_categories = loyalty_rule.product_category_id.ids
            for category in loyalty_rule.product_category_ids:
                selected_categories += category.search([('id', 'child_of', category.id)]).ids
            selected_categories = list(set(selected_categories))
            domain.append(('product_id.categ_id', 'in', selected_categories))
        # Add product tag filtering if product_tag_id is defined in the loyalty rule
        if loyalty_rule.product_tag_id:
            domain.append(('product_id.product_tag_ids', '=', loyalty_rule.product_tag_id.id))
        # Additional checks for stock.lot records
        domain.append(('product_qty', '>=', 1))
        domain.append(('product_id.item_type', '=', 'inventory'))
        # Add company filtering if company_id is defined
        if self.company_id:
            domain.append(('company_id', '=', self.company_id.id))
        lots = self.env['stock.lot'].search(domain)
        return lots

    @api.model
    def _get_next_serial(self, company, product):
        """Return the next serial number to be attributed to the product."""
        if product.tracking == "serial":
            auto_generate_sequence = self.env['nhcl.master.sequence'].search(
                [('nhcl_code', '=', 'Auto Serial Number'), ('nhcl_state', '=', 'activate')])
            if auto_generate_sequence:
                if auto_generate_sequence.nhcl_next_number == 1:
                    last_serial = auto_generate_sequence.nhcl_prefix + '0'
                else:
                    last_serial = auto_generate_sequence.nhcl_prefix + str(auto_generate_sequence.nhcl_next_number)
                if last_serial:
                    return self.env['stock.lot'].generate_lot_names(last_serial, 2)[1]['lot_name']
            else:
                raise ValidationError(
                    "Serial sequence is not configured in the Sequence Master. Please configure it.")
        else:
            return super(StockLot, self)._get_next_serial(company, product)


class StockBackorderConfirmation(models.TransientModel):
    """Inherited stock.backorder.confirmation class to override existing functions"""
    _inherit = 'stock.backorder.confirmation'

    def process(self):
        res = super(StockBackorderConfirmation, self).process()
        for pick_id in self.pick_ids:
            if pick_id.picking_type_id.code == 'incoming' and pick_id.stock_picking_type == 'receipt' and pick_id.is_confirm == False:
                backorder = self.env['stock.picking'].search([('backorder_id', '=', pick_id.id)])
                un_matched_ids = pick_id.stock_verification_ids.filtered(lambda x: x.stock_status == 'un_matched')
                pending_ids = pick_id.stock_picking_delivery_ids.filtered(lambda y:y.status == 'pending')
                barcode_vals_list = []
                for pending in pending_ids:
                    if pending.sequence and pending.sequence != "New" and pending.sequence.startswith("LR"):
                        barcode_vals_list.append({
                            'lr_number': pending.lr_number,
                            'sequence': pending.sequence,
                            'barcode': pending.barcode,
                            'status': pending.status,
                            'stock_picking_delivery_id': backorder.id,
                        })
                if barcode_vals_list:
                    self.env['stock.picking.barcode'].create(barcode_vals_list)
                if backorder:
                    backorder.move_ids.move_line_ids.filtered(lambda x: x.quantity == 0).unlink()
                    for un_matched_id in un_matched_ids:
                        move_lines = backorder.move_ids.move_line_ids.filtered(lambda
                                                                                   x: x.product_id == un_matched_id.stock_product_id and x.type_product != un_matched_id.type_product)
                        for move_line in move_lines:
                            move_line.type_product = un_matched_id.type_product
                        vals = {
                            'stock_product_id': un_matched_id.stock_product_id.id,
                            'stock_serial': un_matched_id.stock_serial,
                            'stock_qty': un_matched_id.stock_qty - un_matched_id.stock_actual_qty,
                            # 'stock_actual_qty': un_matched_id.stock_actual_qty,
                            'stock_status': un_matched_id.stock_status,
                            'type_product': un_matched_id.type_product,
                            'nhcl_lot_hsn_code': un_matched_id.nhcl_lot_hsn_code,
                            'sale_tax_ids': un_matched_id.sale_tax_ids.ids,
                            'purchase_tax_ids': un_matched_id.purchase_tax_ids.ids,
                            'stock_product_barcode': un_matched_id.stock_product_barcode,
                            'stock_picking_id': backorder.id,
                            'mr_price': un_matched_id.mr_price,
                            'rs_price': un_matched_id.rs_price if un_matched_id.rs_price else 0,
                            'cost_price': un_matched_id.cost_price if un_matched_id.cost_price else 0,
                            'segment': un_matched_id.segment,
                            'categ_1': un_matched_id.categ_1.id if un_matched_id.categ_1 else False,
                            'categ_2': un_matched_id.categ_2.id if un_matched_id.categ_2 else False,
                            'categ_3': un_matched_id.categ_3.id if un_matched_id.categ_3 else False,
                            'categ_4': un_matched_id.categ_4.id if un_matched_id.categ_4 else False,
                            'categ_5': un_matched_id.categ_5.id if un_matched_id.categ_5 else False,
                            'categ_6': un_matched_id.categ_6.id if un_matched_id.categ_6 else False,
                            'categ_7': un_matched_id.categ_7.id if un_matched_id.categ_7 else False,
                            'categ_8': un_matched_id.categ_8.id if un_matched_id.categ_8 else False,
                            'descrip_1': un_matched_id.descrip_1.id if un_matched_id.descrip_1 else False,
                            'descrip_2': un_matched_id.descrip_2.id if un_matched_id.descrip_2 else False,
                            'descrip_3': un_matched_id.descrip_3.id if un_matched_id.descrip_3 else False,
                            'descrip_4': un_matched_id.descrip_4.id if un_matched_id.descrip_4 else False,
                            'descrip_5': un_matched_id.descrip_5.id if un_matched_id.descrip_5 else False,
                            'descrip_6': un_matched_id.descrip_6.id if un_matched_id.descrip_6 else False,
                            'nhcl_categ_1': un_matched_id.nhcl_categ_1 if un_matched_id.nhcl_categ_1 else False,
                            'nhcl_categ_2': un_matched_id.nhcl_categ_2 if un_matched_id.nhcl_categ_2 else False,
                            'nhcl_categ_3': un_matched_id.nhcl_categ_3 if un_matched_id.nhcl_categ_3 else False,
                            'nhcl_categ_4': un_matched_id.nhcl_categ_4 if un_matched_id.nhcl_categ_4 else False,
                            'nhcl_categ_5': un_matched_id.nhcl_categ_5 if un_matched_id.nhcl_categ_5 else False,
                            'nhcl_categ_6': un_matched_id.nhcl_categ_6 if un_matched_id.nhcl_categ_6 else False,
                            'nhcl_categ_7': un_matched_id.nhcl_categ_7 if un_matched_id.nhcl_categ_7 else False,
                            'nhcl_categ_8': un_matched_id.nhcl_categ_8 if un_matched_id.nhcl_categ_8 else False,
                            'nhcl_descrip_1': un_matched_id.nhcl_descrip_1 if un_matched_id.nhcl_descrip_1 else False,
                            'nhcl_descrip_2': un_matched_id.nhcl_descrip_2 if un_matched_id.nhcl_descrip_2 else False,
                            'nhcl_descrip_3': un_matched_id.nhcl_descrip_3 if un_matched_id.nhcl_descrip_3 else False,
                            'nhcl_descrip_4': un_matched_id.nhcl_descrip_4 if un_matched_id.nhcl_descrip_4 else False,
                            'nhcl_descrip_5': un_matched_id.nhcl_descrip_5 if un_matched_id.nhcl_descrip_5 else False,
                            'nhcl_descrip_6': un_matched_id.nhcl_descrip_6 if un_matched_id.nhcl_descrip_6 else False,
                            'nhcl_descrip_7': un_matched_id.nhcl_descrip_7 if un_matched_id.nhcl_descrip_7 else False,
                            'nhcl_descrip_8': un_matched_id.nhcl_descrip_8 if un_matched_id.nhcl_descrip_8 else False,
                        }
                        self.env['stock.verification'].create(vals)
            elif pick_id.picking_type_id.code == 'incoming' and pick_id.stock_type in [
                'data_import'] and pick_id.is_confirm == False:
                # pick_id.move_ids.auto_generate_serial_numbers()
                pick_id.is_confirm = True
        return res

    def process_cancel_backorder(self):
        for pick_id in self.pick_ids:
            if pick_id.picking_type_id.code == 'incoming' and pick_id.stock_type in [
                'data_import'] and pick_id.is_confirm == False:
                pick_id.move_ids.auto_generate_serial_numbers()
                pick_id.is_confirm = True
        return super(StockBackorderConfirmation, self).process_cancel_backorder()


class StockVerification(models.Model):
    _name = 'stock.verification'
    _description = "stock verification"

    stock_picking_id = fields.Many2one('stock.picking', copy=False)
    stock_product_id = fields.Many2one('product.product', string='Product', copy=False)
    stock_serial = fields.Char(string="Serial's", copy=False)
    stock_product_barcode = fields.Char(string="Barcode", copy=False)
    stock_qty = fields.Float(string='Qty', copy=False)
    stock_actual_qty = fields.Float(string='Act Qty', copy=False)
    type_product = fields.Selection([('brand', 'Brand'), ('un_brand', 'Un Brand'),
                                     ('others', 'Others')], string='Type Product')
    stock_status = fields.Selection([('matched', 'Matched'), ('un_matched', 'Un Matched')], string='Status',
                                    default='un_matched')
    categ_1 = fields.Many2one('product.attribute.value', string='Color', copy=False,
                              domain=[('attribute_id.name', '=', 'Color')])
    categ_2 = fields.Many2one('product.attribute.value', string='Fit', copy=False,
                              domain=[('attribute_id.name', '=', 'Fit')])
    categ_3 = fields.Many2one('product.attribute.value', string='Brand', copy=False,
                              domain=[('attribute_id.name', '=', 'Brand')])
    categ_4 = fields.Many2one('product.attribute.value', string='Pattern', copy=False,
                              domain=[('attribute_id.name', '=', 'Pattern')])
    categ_5 = fields.Many2one('product.attribute.value', string='Border Type', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Type')])
    categ_6 = fields.Many2one('product.attribute.value', string='Border Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Size')])
    categ_7 = fields.Many2one('product.attribute.value', string='Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Size')])
    categ_8 = fields.Many2one('product.attribute.value', string='Design', copy=False,
                              domain=[('attribute_id.name', '=', 'Design')])
    descrip_1 = fields.Many2one('product.aging.line', string="Product Aging", copy=False)
    descrip_2 = fields.Many2one('product.attribute.value', string='Range', copy=False,
                                domain=[('attribute_id.name', '=', 'Range')])
    descrip_3 = fields.Many2one('product.attribute.value', string='Collection', copy=False,
                                domain=[('attribute_id.name', '=', 'Collection')])
    descrip_4 = fields.Many2one('product.attribute.value', string='Fabric', copy=False,
                                domain=[('attribute_id.name', '=', 'Fabric')])
    descrip_5 = fields.Many2one('product.attribute.value', string='Exclusive', copy=False,
                                domain=[('attribute_id.name', '=', 'Exclusive')])
    descrip_6 = fields.Many2one('product.attribute.value', string='Print', copy=False,
                                domain=[('attribute_id.name', '=', 'Print')])
    descrip_7 = fields.Many2one('product.attribute.value', string='Days Ageing', copy=False,
                                domain=[('attribute_id.name', '=', 'Days Ageing')])
    descrip_8 = fields.Many2one('product.attribute.value', string='Description 8', copy=False)
    cost_price = fields.Float(string='CP', copy=False)
    mr_price = fields.Float(string='MRP', copy=False)
    rs_price = fields.Float(string='RSP', copy=False)
    segment = fields.Selection([('apparel', 'Apparel'), ('non_apparel', 'Non Apparel'), ('others', 'Others')],
                               string="Segment", copy=False, store=True)

    nhcl_categ_1 = fields.Char(string='Color', copy=False,
                               )
    nhcl_categ_2 = fields.Char(string='Fit', copy=False,
                               )
    nhcl_categ_3 = fields.Char(string='Brand', copy=False,
                               )
    nhcl_categ_4 = fields.Char(string='Pattern', copy=False,
                               )
    nhcl_categ_5 = fields.Char(string='Border Type', copy=False,
                               )
    nhcl_categ_6 = fields.Char(string='Border Size', copy=False,
                               )
    nhcl_categ_7 = fields.Char(string='Size', copy=False,
                               )
    nhcl_categ_8 = fields.Char(string='Design', copy=False,
                               )
    nhcl_descrip_1 = fields.Char(string="Product Aging", copy=False)
    nhcl_descrip_2 = fields.Char(string='Range', copy=False,
                                 )
    nhcl_descrip_3 = fields.Char(string='Collection', copy=False,
                                 )
    nhcl_descrip_4 = fields.Char(string='Fabric', copy=False,
                                 )
    nhcl_descrip_5 = fields.Char(string='Exclusive', copy=False,
                                 )
    nhcl_descrip_6 = fields.Char(string='Print', copy=False,
                                 )
    nhcl_descrip_7 = fields.Char(string='Days Ageing', copy=False)
    nhcl_descrip_8 = fields.Char(string='Description 8', copy=False)
    nhcl_lot_hsn_code = fields.Char(string="HSN Code", copy=False)
    sale_tax_ids = fields.Many2many('account.tax', string="Sale Taxes", copy=False)
    purchase_tax_ids = fields.Many2many('account.tax', 'purch_verification_tax', string="Purchase Taxes", copy=False)

    # def write(self, vals):
    #     res = super(StockVerification,self).write(vals)
    #     if res.product_id.tracking == 'serial' and res.stock_qty > 1:
    #         raise ValidationError("You cannot add more than 1.")

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            product = self.env['product.product'].browse(vals.get('stock_product_id'))
            serial_or_barcode = vals.get('stock_serial')
            qty = vals.get('stock_qty', 0)

            # Validation
            if product and product.tracking == 'serial' and qty > 1:
                if serial_or_barcode and not re.match(r'^\d{13}$', serial_or_barcode):
                    raise ValidationError(
                        "You cannot add more than 1 quantity for a serial-tracked product (by serial number)."
                    )

        return super().create(vals_list)

    def write(self, vals):
        res = super().write(vals)
        for record in self:
            product = record.stock_product_id
            serial_or_barcode = record.stock_serial
            qty = record.stock_qty

            if product.tracking == 'serial' and qty > 1:
                if not re.match(r'^\d{13}$', serial_or_barcode):
                    raise ValidationError(
                        "You cannot set quantity > 1 for a serial-tracked product (by serial number).")
        return res

    @api.constrains('stock_actual_qty', 'stock_qty')
    def _check_actual_qty_not_greater(self):
        for rec in self:
            if rec.stock_actual_qty > rec.stock_qty:
                raise ValidationError(
                    _("Actual Quantity cannot be greater than Available Quantity.")
                )


class StockPickingBatch(models.Model):
    _inherit = 'stock.picking.batch'

    nhcl_replication_status = fields.Boolean(string='Replication Status')
    stock_picking_type = fields.Selection(string='Type',
                                          tracking=True, related='picking_type_id.stock_picking_type')
    total_delivered_qty = fields.Float(
        string="Quantity",
        compute="_compute_batch_totals"
    )

    total_untaxed_amount = fields.Float(
        string="Untaxed Amount",
        compute="_compute_batch_totals"
    )

    total_amount = fields.Float(
        string="Net Amount",
        compute="_compute_batch_totals"
    )
    transpoter_id = fields.Many2one('dev.transport.details', string='Transport by')
    transpoter_route_id = fields.Many2one('dev.routes.details', string='Transporter Route')
    no_of_parcel = fields.Integer(string='No Of Parcels')
    tracking_number = fields.Char(string='Tracking Number')
    nhcl_tracking_number = fields.Char(string='Source Tracking Number')
    lr_number = fields.Char(string='LR Number')
    vehicle_number = fields.Char(string='Vehicle Number')
    driver_name = fields.Char(string='Driver Name')
    invoice_count = fields.Integer(
        string="Invoices",
        compute="_compute_invoice_count"
    )

    invoice_ids = fields.Many2many(
        'account.move',
        compute="_compute_invoice_count",
        string="Invoices"
    )

    invoice_dates_display = fields.Char(
        string="Invoice Date",
        compute="_compute_invoice_data"
    )
    transfer_count = fields.Integer(
        string="No. of Packets",
        compute="_compute_transfer_count",
        store=True
    )
    transfer_type = fields.Selection(
        [
            ('in', 'Transfer In'),
            ('out', 'Transfer Out')
        ],
        string="Transfer Type",
        compute="_compute_transfer_type",
        store=True
    )
    family_id = fields.Many2one('product.category', string="Family", compute="_compute_family", store=True)
    hsn_code = fields.Char(string="HSN Code", compute="_compute_hsn", store=True)
    division_onhand_qty = fields.Float(string="SOH", compute="_compute_division_onhand", store=False)
    gst_types = fields.Char(string="GST Types", compute="_compute_gst_types", store=True)

    @api.depends(
        'picking_ids.move_ids_without_package.purchase_line_id.taxes_id',
        'picking_ids.move_ids_without_package.sale_line_id.tax_id'
    )
    def _compute_gst_types(self):
        for batch in self:
            moves = batch.picking_ids.mapped('move_ids_without_package')
            purchase_taxes = moves.mapped('purchase_line_id.taxes_id')
            sale_taxes = moves.mapped('sale_line_id.tax_id')
            taxes = purchase_taxes | sale_taxes
            gst_set = set()
            for tax in taxes:
                children = tax.children_tax_ids if tax.amount_type == 'group' else tax
                for child in children:
                    name = (child.name or '').upper()
                    if "CGST" in name:
                        gst_set.add("CGST")
                    elif "SGST" in name:
                        gst_set.add("SGST")
                    elif "IGST" in name:
                        gst_set.add("IGST")
            batch.gst_types = (
                "GST: " + ", ".join(sorted(gst_set))
                if gst_set else ""
            )

    def _compute_division_onhand(self):
        Quant = self.env['stock.quant']
        products = self.env['product.product'].search([
            ('family_categ_id', 'in', self.mapped('family_id').ids)
        ])
        product_family_map = {
            p.id: p.family_categ_id.id for p in products
        }
        grouped = Quant.read_group(
            [
                ('product_id', 'in', products.ids),
                ('location_id.usage', '=', 'internal')
            ],
            ['quantity:sum'],
            ['product_id']
        )
        family_map = {}
        for g in grouped:
            product_id = g['product_id'][0]
            family_id = product_family_map.get(product_id)
            if family_id:
                family_map.setdefault(family_id, 0.0)
                family_map[family_id] += g['quantity']
        # Step 4: assign
        for batch in self:
            batch.division_onhand_qty = family_map.get(batch.family_id.id, 0.0)

    @api.depends('picking_ids.move_ids_without_package.product_id.categ_id')
    def _compute_family(self):
        for batch in self:
            family = False
            picking = batch.picking_ids[:1]
            if picking:
                move = picking.move_ids_without_package[:1]
                if move and move.product_id.categ_id:
                    categ = move.product_id.categ_id
                    # safely go up hierarchy
                    family = (
                                     categ.parent_id
                                     and categ.parent_id.parent_id
                                     and categ.parent_id.parent_id.parent_id
                             ) or categ.parent_id or categ
            batch.family_id = family

    @api.depends('picking_ids.move_ids_without_package.product_id.l10n_in_hsn_code')
    def _compute_hsn(self):
        for batch in self:
            hsn_code = False
            picking = batch.picking_ids[:1]
            if picking:
                move = picking.move_ids_without_package[:1]
                if move and move.product_id.l10n_in_hsn_code:
                    hsn_code = move.product_id.l10n_in_hsn_code

            batch.hsn_code = hsn_code

    @api.depends('stock_picking_type')
    def _compute_transfer_type(self):
        for batch in self:
            if batch.stock_picking_type == 'receipt':
                batch.transfer_type = 'in'
            elif batch.stock_picking_type in ['return', 'damage']:
                batch.transfer_type = 'out'
            else:
                batch.transfer_type = False

    @api.depends('picking_ids')
    def _compute_transfer_count(self):
        for rec in self:
            rec.transfer_count = len(rec.picking_ids)

    @api.depends(
        'picking_ids.sale_id',
        'picking_ids.sale_id.invoice_ids',
        'picking_ids.sale_id.invoice_ids.invoice_date'
    )
    def _compute_invoice_data(self):
        for rec in self:
            sale_orders = rec.mapped('picking_ids.sale_id')

            invoices = sale_orders.mapped('invoice_ids').filtered(
                lambda m: m.move_type == 'out_invoice'
            )

            rec.invoice_ids = invoices
            rec.invoice_count = len(invoices)

            # Format invoice dates
            dates = invoices.mapped('invoice_date')
            formatted_dates = [
                d.strftime('%d-%m-%Y') for d in dates if d
            ]

            rec.invoice_dates_display = ", ".join(formatted_dates)

    @api.depends('picking_ids.sale_id.invoice_ids')
    def _compute_invoice_count(self):
        for rec in self:
            sale_orders = rec.mapped('picking_ids.sale_id')
            invoices = sale_orders.mapped('invoice_ids').filtered(
                lambda m: m.move_type == 'out_invoice'
            )
            rec.invoice_ids = invoices
            rec.invoice_count = len(invoices)

    def action_view_nhcl_invoices(self):
        self.ensure_one()

        return {
            'type': 'ir.actions.act_window',
            'name': 'Invoices',
            'res_model': 'account.move',
            'view_mode': 'tree,form',
            'domain': [('id', 'in', self.invoice_ids.ids)],
            'context': {'default_move_type': 'out_invoice', 'create': False, 'delete': False, 'duplicate': False,
                        'edit': False},
        }

    def _compute_batch_totals(self):
        for batch in self:
            qty_total = 0.0
            untaxed_total = 0.0
            total_with_tax = 0.0

            for picking in batch.picking_ids:
                for move in picking.move_ids_without_package:

                    so_line = move.sale_line_id
                    if not so_line:
                        continue

                    delivered_qty = move.quantity
                    price = so_line.price_unit
                    taxes = so_line.tax_id

                    qty_total += delivered_qty
                    untaxed_total += delivered_qty * price

                    tax_result = taxes.compute_all(
                        price,
                        quantity=delivered_qty,
                        product=move.product_id,
                        partner=picking.partner_id,
                    )

                    total_with_tax += tax_result['total_included']

            batch.total_delivered_qty = qty_total
            batch.total_untaxed_amount = untaxed_total
            batch.total_amount = total_with_tax

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', '/') == '/':
                company_id = vals.get('company_id', self.env.company.id)
                picking_type = self.env['stock.picking.type'].search([('stock_picking_type', '=', 'regular')])
                damage_picking_type = self.env['stock.picking.type'].search([('stock_picking_type', '=', 'damage')])
                return_picking_type = self.env['stock.picking.type'].search([('stock_picking_type', '=', 'return')])
                if vals.get('is_wave'):
                    vals['name'] = self.env['ir.sequence'].with_company(company_id).next_by_code('picking.wave') or '/'
                elif vals.get('picking_type_id') == picking_type.id:
                    sequence = self.env['ir.sequence'].with_company(company_id).next_by_code(
                        'picking.batch.return') or '/'
                    company = self.env.company.name
                    operation = picking_type.stock_picking_type
                    vals['name'] = f'Batch/{company}/{operation}/{sequence}'

                elif vals.get('picking_type_id') == damage_picking_type.id:
                    sequence = self.env['ir.sequence'].with_company(company_id).next_by_code(
                        'picking.batch.return') or '/'
                    company = self.env.company.name
                    operation = damage_picking_type.stock_picking_type
                    vals['name'] = f'Batch/{company}/{operation}/{sequence}'

                elif vals.get('picking_type_id') == return_picking_type.id:
                    sequence = self.env['ir.sequence'].with_company(company_id).next_by_code(
                        'picking.batch.return') or '/'
                    company = self.env.company.name
                    operation = return_picking_type.stock_picking_type
                    vals['name'] = f'Batch/{company}/{operation}/{sequence}'

                else:
                    vals['name'] = self.env['ir.sequence'].with_company(company_id).next_by_code('picking.batch') or '/'
        res = super().create(vals_list)
        print("batch", res)
        if res.stock_picking_type == 'receipt':
            for pick in res.picking_ids:
                pick.nhcl_batch_number = res.name
        return res

    # def action_confirm(self):
    #     for batch in self:
    #         for picking in batch.picking_ids:
    #             if batch.stock_picking_type in ['return', 'damage']:
    #                 if batch.no_of_parcel <= 0:
    #                     raise ValidationError("Number of Parcels should be more than 0.")
    #                 picking.write({
    #                     # 'transpoter_id': batch.transpoter_id.id,
    #                     # 'transpoter_route_id': batch.transpoter_route_id.id,
    #                     'no_of_parcel': batch.no_of_parcel,
    #                     # 'lr_number': batch.lr_number,
    #                     # 'driver_name': batch.driver_name,
    #                     # 'vehicle_number': batch.vehicle_number,
    #                 })
    #     return super().action_confirm()

    def action_create_invoices(self):
        sale_orders = self.mapped('picking_ids.sale_id')
        sale_orders = sale_orders.filtered(lambda so: so.state in ['sale'])
        if not sale_orders:
            raise ValidationError("No confirmed Sale Orders found for invoicing.")
        return {
            'type': 'ir.actions.act_window',
            'name': 'Create Invoices',
            'res_model': 'sale.advance.payment.inv',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_model': 'sale.order',
                'active_ids': sale_orders.ids,
            }
        }


class StockVerificationImport(models.TransientModel):
    _name = 'stock.verification.import'
    _description = 'Import Stock Verification'

    stock_picking_id = fields.Many2one('stock.picking', string="Stock Picking", required=True)
    file_type = fields.Selection([
        ('excel', 'Excel'),
        ('csv', 'CSV')
    ], string="File Type", required=True, default='excel')
    file_data = fields.Binary(string="Upload File", required=True)
    file_name = fields.Char(string="File Name", required=True)

    def is_valid_file_extension(self, file_name):
        valid_extensions = ['.xls', '.xlsx', '.ods', '.csv', '.txt']
        return any(file_name.lower().endswith(ext) for ext in valid_extensions)

    def clean_string(self, value):
        if value is None:
            return ''
        if isinstance(value, float):
            return str(int(value))
        return str(value).strip().replace('.0', '')

    def action_import(self):
        if not self.file_data:
            raise UserError(_("Please upload an Excel file."))

        if not any(line.status == 'matched' for line in self.stock_picking_id.stock_picking_delivery_ids):
            raise ValidationError(_("You cannot proceed. Please scan at least one bale barcode first."))


        try:
            data = base64.b64decode(self.file_data)
            wb = load_workbook(filename=BytesIO(data), read_only=True)
            sheet = wb.active
        except Exception as e:
            raise UserError(_("Error reading Excel file: %s") % e)

        gs1_pattern = r'01(\d{14})21([A-Za-z0-9]+)'  # GS1 pattern for unbranded
        ean13_pattern = r'^\d{13}$'  # EAN13 for branded

        # Accumulators
        unbrand_serials_seen = {}
        unbrand_lot_qty = {}
        branded_serial_qty = {}  # barcode → total excel qty
        branded_lot_qty = {}
        # NEW (R barcode accumulators)
        rbrand_serial_qty = {}
        rbrand_lot_qty = {}

        # ===== FIRST PASS: Read Excel & validate =====
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            raw_code = str(row[0]).strip() if row[0] else ''
            qty = float(row[1]) if row[1] else 0

            if not raw_code:
                raise UserError(_("Missing Serial/Barcode in row %s.") % row_index)
            if qty <= 0:
                raise UserError(_("Quantity must be greater than 0 in row %s.") % row_index)

            # Try GS1 extraction for unbranded
            match = re.match(gs1_pattern, raw_code)
            unbrand_serial = match.group(2) if match else raw_code

            # Check if unbranded line exists
            unbrand_line = self.env['stock.verification'].search([
                ('stock_serial', 'ilike', unbrand_serial),
                ('stock_picking_id', '=', self.stock_picking_id.id),
                ('type_product', '=', 'un_brand')
            ], limit=1)

            if unbrand_line:
                tracking_type = unbrand_line.stock_product_id.tracking
                if tracking_type == 'serial':
                    # Count total qty per serial in Excel
                    unbrand_serials_seen[unbrand_serial] = unbrand_serials_seen.get(unbrand_serial, 0) + qty
                    if unbrand_serials_seen[unbrand_serial] > 1:
                        raise UserError(
                            _("Duplicate or excess quantity for unbranded serial '%s'. Total in Excel = %s. Only 1 allowed.") %
                            (unbrand_serial, unbrand_serials_seen[unbrand_serial])
                        )
                else:  # lot
                    unbrand_lot_qty[unbrand_serial] = unbrand_lot_qty.get(unbrand_serial, 0) + qty
            else:
                # ===== NEW CONDITION FOR R BARCODE =====
                # if raw_code.startswith('R'):
                if re.search(r'[A-Za-z]', raw_code) and re.search(r'\d', raw_code):
                    rbrand_line = self.env['stock.verification'].search([
                        ('stock_product_barcode', '=', raw_code),
                        ('stock_picking_id', '=', self.stock_picking_id.id),
                        ('type_product', '=', 'brand')
                    ])

                    if not rbrand_line:
                        raise UserError(
                            _("No verification line found for R barcode '%s' in row %s.") % (raw_code, row_index)
                        )

                    tracking_type = rbrand_line[0].stock_product_id.tracking

                    if tracking_type == 'serial':
                        rbrand_serial_qty[raw_code] = rbrand_serial_qty.get(raw_code, 0) + qty
                    else:
                        rbrand_lot_qty[raw_code] = rbrand_lot_qty.get(raw_code, 0) + qty

                    continue
                # Branded product
                if not re.match(ean13_pattern, raw_code):
                    raise UserError(_("Invalid barcode for branded product in row %s.") % row_index)

                brand_line = self.env['stock.verification'].search([
                    ('stock_product_barcode', '=', raw_code),
                    ('stock_picking_id', '=', self.stock_picking_id.id),
                    ('type_product', '=', 'brand')
                ])

                if not brand_line:
                    raise UserError(
                        _("No verification line found for branded barcode '%s' in row %s.") % (raw_code, row_index))

                tracking_type = brand_line[0].stock_product_id.tracking
                if tracking_type == 'serial':
                    branded_serial_qty[raw_code] = branded_serial_qty.get(raw_code, 0) + qty
                else:  # lot
                    branded_lot_qty[raw_code] = branded_lot_qty.get(raw_code, 0) + qty

        # ===== SECOND PASS: Update Odoo lines =====

        # Unbranded Serial → qty always 1
        for serial in unbrand_serials_seen:
            line = self.env['stock.verification'].search([
                ('stock_serial', 'ilike', serial),
                ('stock_picking_id', '=', self.stock_picking_id.id),
                ('type_product', '=', 'un_brand')
            ], limit=1)
            if line:
                line.stock_actual_qty = 1
                line.stock_status = 'matched' if line.stock_actual_qty == line.stock_qty else 'un_matched'

        # Unbranded Lot → set summed qty (with validation)
        for serial, total_qty in unbrand_lot_qty.items():
            lines = self.env['stock.verification'].search([
                ('stock_serial', 'ilike', serial),
                ('stock_picking_id', '=', self.stock_picking_id.id),
                ('type_product', '=', 'un_brand')
            ])
            if lines:
                total_stock_qty = sum(lines.mapped('stock_qty'))
                if total_qty > total_stock_qty:
                    raise UserError(_(
                        "Excel quantity (%s) for unbranded lot '%s' exceeds available stock quantity (%s)."
                    ) % (total_qty, serial, total_stock_qty))

                for line in lines:
                    line.stock_actual_qty = total_qty
                    line.stock_status = 'matched' if line.stock_actual_qty == line.stock_qty else 'un_matched'

        # Branded Serial → validate & fill first N lines
        for barcode, excel_qty in branded_serial_qty.items():
            lines = self.env['stock.verification'].search([
                ('stock_product_barcode', '=', barcode),
                ('stock_picking_id', '=', self.stock_picking_id.id),
                ('type_product', '=', 'brand')
            ])
            if not lines:
                continue

            total_stock_qty = sum(lines.mapped('stock_qty'))
            if excel_qty > total_stock_qty:
                raise UserError(_(
                    "Excel quantity (%s) for branded serial '%s' exceeds available stock quantity (%s)."
                ) % (excel_qty, barcode, total_stock_qty))

            count_assigned = 0
            for line in lines:
                if count_assigned < excel_qty:
                    line.stock_actual_qty = 1
                    line.stock_status = 'matched' if line.stock_actual_qty == line.stock_qty else 'un_matched'
                    count_assigned += 1
                else:
                    pass  # untouched lines

        # Branded Lot → allocate progressively
        for barcode, excel_qty in branded_lot_qty.items():
            lines = self.env['stock.verification'].search([
                ('stock_product_barcode', '=', barcode),
                ('stock_picking_id', '=', self.stock_picking_id.id),
                ('type_product', '=', 'brand')
            ], order="id")

            if not lines:
                continue

            total_stock_qty = sum(lines.mapped('stock_qty'))
            if excel_qty > total_stock_qty:
                raise UserError(_(
                    "Excel quantity (%s) for branded lot '%s' exceeds available stock quantity (%s)."
                ) % (excel_qty, barcode, total_stock_qty))

            remaining_qty = excel_qty

            for line in lines:
                if remaining_qty <= 0:
                    break

                allowed_qty = min(line.stock_qty, remaining_qty)

                line.stock_actual_qty = allowed_qty
                line.stock_status = 'matched' if allowed_qty == line.stock_qty else 'un_matched'

                remaining_qty -= allowed_qty

            # ===== NEW R BRAND SERIAL =====
        for barcode, excel_qty in rbrand_serial_qty.items():

            lines = self.env['stock.verification'].search([
                ('stock_product_barcode', '=', barcode),
                ('stock_picking_id', '=', self.stock_picking_id.id),
                ('type_product', '=', 'brand')
            ])

            total_stock_qty = sum(lines.mapped('stock_qty'))

            if excel_qty > total_stock_qty:
                raise UserError(_(
                    "Excel quantity (%s) for R barcode '%s' exceeds available stock quantity (%s)."
                ) % (excel_qty, barcode, total_stock_qty))

            count = 0
            for line in lines:
                if count < excel_qty:
                    line.stock_actual_qty = 1
                    line.stock_status = 'matched' if line.stock_actual_qty == line.stock_qty else 'un_matched'
                    count += 1

            # ===== NEW R BRAND LOT =====
        for barcode, excel_qty in rbrand_lot_qty.items():

            lines = self.env['stock.verification'].search([
                ('stock_product_barcode', '=', barcode),
                ('stock_picking_id', '=', self.stock_picking_id.id),
                ('type_product', '=', 'brand')
            ], order="id")

            total_stock_qty = sum(lines.mapped('stock_qty'))

            if excel_qty > total_stock_qty:
                raise UserError(_(
                    "Excel quantity (%s) for R barcode '%s' exceeds available stock quantity (%s)."
                ) % (excel_qty, barcode, total_stock_qty))

            remaining_qty = excel_qty

            for line in lines:

                if remaining_qty <= 0:
                    break

                allowed_qty = min(line.stock_qty, remaining_qty)

                line.stock_actual_qty = allowed_qty
                line.stock_status = 'matched' if allowed_qty == line.stock_qty else 'un_matched'

                remaining_qty -= allowed_qty

        self.file_data = False

        return {
            'effect': {
                'fadeout': 'slow',
                'message': _("Excel processed successfully for Brand and UnBrand products."),
                'type': 'rainbow_man',
            }
        }

    # def action_import(self):
    #     _logger.info("Starting barcode verification import process.")
    #
    #     if self.stock_picking_id.verification_success == 'matched':
    #         raise UserError("This picking is already verified and matched. Further imports are not allowed.")
    #
    #     if not self.file_name or isinstance(self.file_name, bool):
    #         raise UserError("File name is missing or invalid.")
    #
    #     if not self.is_valid_file_extension(self.file_name):
    #         raise UserError("Invalid file format! Allowed: .xls, .xlsx, .csv, .txt")
    #
    #     file_content = base64.b64decode(self.file_data)
    #     if len(file_content) == 0:
    #         raise UserError("File is empty!")
    #
    #     _logger.info("File successfully read and decoded.")
    #
    #     verification_lines = []
    #     found_serials = set()
    #     barcode_count = defaultdict(int)
    #     move_lines = self.stock_picking_id.move_line_ids_without_package
    #
    #     try:
    #         if self.file_type == 'excel':
    #             workbook = openpyxl.load_workbook(io.BytesIO(file_content))
    #             sheet = workbook.active
    #             _logger.info("Processing Excel file...")
    #
    #             # First pass: count barcodes and collect unique serials/lots
    #             for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #                 serial_no = self.clean_string(row[0])
    #                 barcode = self.clean_string(row[1])
    #                 qty = row[2] or 0
    #
    #                 if not serial_no and not barcode:
    #                     raise UserError(f"Row {idx}: Either Serial Number or Barcode must be provided.")
    #
    #                 if not isinstance(qty, (int, float)) or qty <= 0:
    #                     raise UserError(f"Row {idx}: Quantity must be a positive number.")
    #
    #                 if serial_no:
    #                     if serial_no in found_serials:
    #                         raise UserError(f"Row {idx}: Duplicate serial/lot number in file: {serial_no}")
    #
    #                     move_line = move_lines.filtered(lambda ml: ml.lot_name == serial_no)
    #                     if not move_line:
    #                         raise UserError(f"Row {idx}: Serial/Lot number '{serial_no}' not found in move lines.")
    #
    #                     tracking = move_line.product_id.tracking
    #                     if tracking == 'serial':
    #                         if qty != 1:
    #                             raise UserError(
    #                                 f"Row {idx}: Quantity must be 1 for serial-tracked product '{serial_no}'.")
    #                     elif tracking == 'lot':
    #                         pass  # Allow quantity > 1 for lot-tracked
    #                     else:
    #                         raise UserError(
    #                             f"Row {idx}: Product tracking not defined for product '{move_line.product_id.display_name}'.")
    #
    #                     found_serials.add(serial_no)
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': move_line.product_id.id,
    #                         'stock_serial': serial_no,
    #                         'stock_qty': qty,
    #                     }))
    #                 else:
    #                     # Barcode line, to be processed in second pass
    #                     barcode_count[barcode] += 1
    #
    #             # Second pass: process grouped barcodes
    #             for barcode, excel_qty in barcode_count.items():
    #                 matching_lines = move_lines.filtered(
    #                     lambda ml: ml.internal_ref_lot == barcode and ml.lot_name not in found_serials)
    #                 available_serials = [ml.lot_name for ml in matching_lines]
    #
    #                 if len(available_serials) < excel_qty:
    #                     raise UserError(
    #                         f"Barcode '{barcode}' mismatch: Excel shows {excel_qty}, "
    #                         f"but only {len(available_serials)} serials found in stock move lines."
    #                     )
    #
    #                 used_lines = matching_lines[:excel_qty]
    #
    #                 for ml in used_lines:
    #                     found_serials.add(ml.lot_name)
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': ml.product_id.id,
    #                         'stock_serial': ml.lot_name,
    #                         'stock_qty': 1,
    #                     }))
    #
    #         # Clear existing lines and write new ones
    #         self.stock_picking_id.stock_verification_ids = [(5, 0, 0)]
    #         self.stock_picking_id.stock_verification_ids = verification_lines
    #         _logger.info(f"Total verification lines created: {len(verification_lines)}")
    #
    #         operation_qty = sum(self.stock_picking_id.move_line_ids_without_package.mapped('quantity'))
    #         verification_qty = sum(line[2]['stock_qty'] for line in verification_lines)
    #
    #         if operation_qty == verification_qty:
    #             self.stock_picking_id.verification_success = 'matched'
    #         else:
    #             self.stock_picking_id.verification_success = ''
    #
    #     except Exception as e:
    #         _logger.error(f"Error: {str(e)}")
    #         raise UserError(f"Error importing file: {str(e)}")
    #
    #     return {'type': 'ir.actions.act_window_close'}

    # def action_import(self):
    #     _logger.info("Starting barcode verification import process.")
    #
    #     if self.stock_picking_id.verification_success == 'matched':
    #         raise UserError("This picking is already verified and matched. Further imports are not allowed.")
    #
    #     if not self.file_name or isinstance(self.file_name, bool):
    #         raise UserError("File name is missing or invalid.")
    #
    #     if not self.is_valid_file_extension(self.file_name):
    #         raise UserError("Invalid file format! Allowed: .xls, .xlsx, .csv, .txt")
    #
    #     file_content = base64.b64decode(self.file_data)
    #     if len(file_content) == 0:
    #         raise UserError("File is empty!")
    #
    #     _logger.info("File successfully read and decoded.")
    #
    #     verification_lines = []
    #     found_serials = set()
    #     barcode_count = defaultdict(int)
    #     move_lines = self.stock_picking_id.move_line_ids_without_package
    #
    #     try:
    #         if self.file_type == 'excel':
    #             workbook = openpyxl.load_workbook(io.BytesIO(file_content))
    #             sheet = workbook.active
    #             _logger.info("Processing Excel file...")
    #
    #             # First pass: count barcodes and collect unique serials/lots
    #             for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #                 serial_no = self.clean_string(row[0])
    #                 barcode = self.clean_string(row[1])
    #                 qty = row[2] or 0
    #
    #                 if not serial_no and not barcode:
    #                     raise UserError(f"Row {idx}: Either Serial Number or Barcode must be provided.")
    #
    #                 if not isinstance(qty, (int, float)) or qty <= 0:
    #                     raise UserError(f"Row {idx}: Quantity must be a positive number.")
    #
    #                 if serial_no:
    #                     if serial_no in found_serials:
    #                         raise UserError(f"Row {idx}: Duplicate serial/lot number in file: {serial_no}")
    #
    #                     move_line = move_lines.filtered(lambda ml: ml.lot_name == serial_no)
    #                     if not move_line:
    #                         raise UserError(f"Row {idx}: Serial/Lot number '{serial_no}' not found in move lines.")
    #
    #                     tracking = move_line.product_id.tracking
    #                     if tracking == 'serial':
    #                         if qty != 1:
    #                             raise UserError(
    #                                 f"Row {idx}: Quantity must be 1 for serial-tracked product '{serial_no}'.")
    #                     elif tracking == 'lot':
    #                         pass  # Allow quantity > 1 for lot-tracked
    #                     else:
    #                         raise UserError(
    #                             f"Row {idx}: Product tracking not defined for product '{move_line.product_id.display_name}'.")
    #
    #                     found_serials.add(serial_no)
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': move_line.product_id.id,
    #                         'stock_serial': serial_no,
    #                         'stock_qty': qty,
    #                     }))
    #                 else:
    #                     # Barcode line, to be processed in second pass
    #                     barcode_count[barcode] += qty
    #
    #             # Second pass: process grouped barcodes
    #             for barcode, excel_qty in barcode_count.items():
    #                 matching_lines = move_lines.filtered(
    #                     lambda ml: ml.internal_ref_lot == barcode and ml.lot_name not in found_serials)
    #
    #                 if not matching_lines:
    #                     raise UserError(
    #                         f"Barcode '{barcode}' not found in stock move lines."
    #                     )
    #
    #                 ml = matching_lines[0]  # ✅ get first match
    #                 tracking = ml.product_id.tracking  # ✅ Get product tracking
    #
    #                 if tracking == 'serial':
    #                     if len(matching_lines) < excel_qty:
    #                         raise UserError(
    #                             f"Barcode '{barcode}' mismatch: Excel shows {excel_qty}, "
    #                             f"but only {len(matching_lines)} serials found in stock move lines."
    #                         )
    #                     used_lines = matching_lines[:excel_qty]
    #
    #                     for each_ml in used_lines:
    #                         found_serials.add(each_ml.lot_name)
    #
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': ml.product_id.id,
    #                         'stock_serial': barcode,
    #                         'stock_qty': excel_qty,  # ✅ total quantity from Excel
    #                     }))
    #
    #                 elif tracking == 'lot':
    #                     # ✅ Allow single lot line with qty > 1
    #                     found_serials.add(ml.lot_name)
    #
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': ml.product_id.id,
    #                         'stock_serial': barcode,
    #                         'stock_qty': excel_qty,  # ✅ total qty from Excel for lot
    #                     }))
    #                 else:
    #                     raise UserError(f"Product '{ml.product_id.display_name}' has undefined tracking.")
    #
    #         # Clear existing lines and write new ones
    #         self.stock_picking_id.stock_verification_ids = [(5, 0, 0)]  # Reset old lines
    #         self.stock_picking_id.stock_verification_ids = verification_lines
    #         _logger.info(f"Total verification lines created: {len(verification_lines)}")
    #
    #         operation_qty = sum(self.stock_picking_id.move_line_ids_without_package.mapped('quantity'))
    #         verification_qty = sum(line[2]['stock_qty'] for line in verification_lines)
    #
    #         if operation_qty == verification_qty:
    #             self.stock_picking_id.verification_success = 'matched'
    #         else:
    #             self.stock_picking_id.verification_success = ''
    #
    #     except Exception as e:
    #         _logger.error(f"Error: {str(e)}")
    #         raise UserError(f"Error importing file: {str(e)}")
    #
    #     return {'type': 'ir.actions.act_window_close'}

    # def action_import(self):
    #     _logger.info("Starting barcode verification import process.")
    #
    #     if self.stock_picking_id.verification_success == 'matched':
    #         raise UserError("This picking is already verified and matched. Further imports are not allowed.")
    #
    #     if not self.file_name or isinstance(self.file_name, bool):
    #         raise UserError("File name is missing or invalid.")
    #
    #     if not self.is_valid_file_extension(self.file_name):
    #         raise UserError("Invalid file format! Allowed: .xls, .xlsx, .csv, .txt")
    #
    #     file_content = base64.b64decode(self.file_data)
    #     if len(file_content) == 0:
    #         raise UserError("File is empty!")
    #
    #     verification_lines = []
    #     found_serials = set()
    #     barcode_qty_map = defaultdict(int)
    #     barcode_product_map = {}
    #     move_lines = self.stock_picking_id.move_line_ids_without_package
    #
    #     try:
    #         if self.file_type == 'excel':
    #             workbook = openpyxl.load_workbook(io.BytesIO(file_content))
    #             sheet = workbook.active
    #             _logger.info("Processing Excel file...")
    #
    #             for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #                 serial_no = self.clean_string(row[0])
    #                 barcode = self.clean_string(row[1])
    #                 qty = row[2] or 0
    #
    #                 if not serial_no and not barcode:
    #                     raise UserError(f"Row {idx}: Either Serial or Barcode must be provided.")
    #
    #                 if serial_no and barcode:
    #                     raise UserError(f"Row {idx}: Provide either Serial or Barcode, not both.")
    #
    #                 if not isinstance(qty, (int, float)) or qty <= 0:
    #                     raise UserError(f"Row {idx}: Quantity must be a positive number.")
    #
    #                 # --- Handle Unbranded via Serial ---
    #                 if serial_no:
    #                     matched_lines = move_lines.filtered(lambda ml: ml.lot_name == serial_no)
    #                     if not matched_lines:
    #                         raise UserError(f"Row {idx}: Serial '{serial_no}' not found in move lines.")
    #
    #                     move_line = next((ml for ml in matched_lines if ml.product_id), None)
    #                     if not move_line:
    #                         raise UserError(f"Row {idx}: No product found for Serial '{serial_no}'.")
    #
    #                     if move_line.type_product == 'brand':
    #                         raise UserError(f"Row {idx}: '{serial_no}' is a branded product. Use Barcode.")
    #
    #                     tracking = move_line.product_id.tracking
    #                     if tracking == 'serial' and qty != 1:
    #                         raise UserError(f"Row {idx}: Quantity must be 1 for serial-tracked product '{serial_no}'.")
    #
    #                     if serial_no in found_serials:
    #                         raise UserError(f"Row {idx}: Duplicate serial '{serial_no}' in file.")
    #
    #                     found_serials.add(serial_no)
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': move_line.product_id.id,
    #                         'stock_serial': serial_no,
    #                         'stock_qty': qty,
    #                     }))
    #
    #                 # --- Handle Branded via Barcode ---
    #                 elif barcode:
    #                     if not re.match(r'^\d{13}$', barcode):
    #                         raise UserError(f"Row {idx}: Barcode '{barcode}' is not valid EAN-13.")
    #
    #                     matched_lines = move_lines.filtered(lambda ml: ml.internal_ref_lot == barcode and ml.lot_name)
    #                     if not matched_lines:
    #                         raise UserError(f"Row {idx}: Barcode '{barcode}' not found in move lines.")
    #
    #                     branded_lines = matched_lines.filtered(lambda ml: ml.type_product == 'brand')
    #                     if not branded_lines:
    #                         raise UserError(f"Row {idx}: Barcode '{barcode}' refers to unbranded product. Use Serial.")
    #
    #                     available_lines = [ml for ml in branded_lines if ml.lot_name not in found_serials]
    #                     if len(available_lines) < qty:
    #                         raise UserError(
    #                             f"Row {idx}: Qty = {qty}, but only {len(available_lines)} available for barcode '{barcode}'.")
    #
    #                     used_lines = available_lines[:int(qty)]
    #                     for ml in used_lines:
    #                         found_serials.add(ml.lot_name)
    #
    #                     if barcode not in barcode_product_map:
    #                         product_line = next((ml for ml in branded_lines if ml.product_id), None)
    #                         if not product_line:
    #                             raise UserError(f"Row {idx}: No product found for barcode '{barcode}'.")
    #                         barcode_product_map[barcode] = product_line.product_id.id
    #
    #                     barcode_qty_map[barcode] += qty
    #
    #         # --- Finalize barcode-based lines (grouped) ---
    #         for barcode, total_qty in barcode_qty_map.items():
    #             verification_lines.append((0, 0, {
    #                 'stock_product_id': barcode_product_map[barcode],
    #                 'stock_serial': barcode,
    #                 'stock_qty': total_qty,
    #             }))
    #
    #         # --- Save verification lines ---
    #         self.stock_picking_id.stock_verification_ids = [(5, 0, 0)]
    #         self.stock_picking_id.stock_verification_ids = verification_lines
    #         _logger.info(f"Total verification lines created: {len(verification_lines)}")
    #
    #         # --- Check if fully matched ---
    #         operation_qty = sum(self.stock_picking_id.move_line_ids_without_package.mapped('quantity'))
    #         verification_qty = sum(line[2]['stock_qty'] for line in verification_lines)
    #
    #         self.stock_picking_id.verification_success = 'matched' if operation_qty == verification_qty else ''
    #
    #     except Exception as e:
    #         _logger.error(f"Error: {str(e)}")
    #         raise UserError(f"Error importing file: {str(e)}")
    #
    #     return {'type': 'ir.actions.act_window_close'}

    # main
    # def action_import(self):
    #     _logger.info("Starting barcode verification import process.")
    #
    #
    #     if not self.file_name or isinstance(self.file_name, bool):
    #         raise UserError("File name is missing or invalid.")
    #
    #     if not self.is_valid_file_extension(self.file_name):
    #         raise UserError("Invalid file format! Allowed: .xls, .xlsx, .csv, .txt")
    #
    #     file_content = base64.b64decode(self.file_data)
    #     if len(file_content) == 0:
    #         raise UserError("File is empty!")
    #
    #     # ✅ Reset verification lines before starting fresh
    #     self.stock_picking_id.stock_verification_ids = [(5, 0, 0)]
    #
    #     gs1_pattern = r'01(\d{14})21([A-Za-z0-9]+)'  # GS1 format
    #     ean13_pattern = r'^\d{13}$'
    #
    #     verification_lines = []
    #     found_serials = set()
    #     barcode_qty_map = defaultdict(float)
    #     barcode_row_map = {}
    #     move_lines = self.stock_picking_id.move_line_ids_without_package
    #
    #     try:
    #         if self.file_type == 'excel':
    #             workbook = openpyxl.load_workbook(io.BytesIO(file_content))
    #             sheet = workbook.active
    #
    #             for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #                 raw_serial_input = self.clean_string(row[0])
    #                 barcode = self.clean_string(row[1])
    #                 qty = row[2] or 0
    #
    #                 if not raw_serial_input and not barcode:
    #                     raise UserError(f"Row {idx}: Either Serial or Barcode must be provided.")
    #                 if raw_serial_input and barcode:
    #                     raise UserError(f"Row {idx}: Provide either Serial or Barcode, not both.")
    #                 if not isinstance(qty, (int, float)) or qty <= 0:
    #                     raise UserError(f"Row {idx}: Quantity must be a positive number.")
    #
    #                 # GS1 Serial Parse
    #                 gs1_match = re.match(gs1_pattern, raw_serial_input)
    #                 serial_no = gs1_match.group(2) if gs1_match else raw_serial_input
    #
    #                 # Unbranded: Use Serial/GS1
    #                 if serial_no:
    #                     matching_lines = move_lines.filtered(lambda ml: ml.lot_name == serial_no)
    #                     if not matching_lines:
    #                         raise UserError(f"Row {idx}: Serial/Lot '{serial_no}' not found in move lines.")
    #
    #                     line = next((ml for ml in matching_lines if ml.product_id), False)
    #                     if not line:
    #                         raise UserError(f"Row {idx}: No product found for serial '{serial_no}'.")
    #
    #                     if line.type_product == 'brand':
    #                         raise UserError(f"Row {idx}: Branded product '{serial_no}' must use barcode, not serial.")
    #
    #                     tracking = line.product_id.tracking
    #                     if tracking == 'serial' and qty != 1:
    #                         raise UserError(f"Row {idx}: Serial-tracked product '{serial_no}' must have quantity 1.")
    #
    #                     if serial_no in found_serials:
    #                         raise UserError(f"Row {idx}: Duplicate serial '{serial_no}' in file.")
    #
    #                     found_serials.add(serial_no)
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': line.product_id.id,
    #                         'stock_serial': serial_no,
    #                         'stock_qty': qty,
    #                     }))
    #
    #                 # Branded: Use Barcode
    #                 elif barcode:
    #                     if not re.match(ean13_pattern, barcode):
    #                         raise UserError(f"Row {idx}: Invalid barcode format '{barcode}'. Must be 13 digits.")
    #                     barcode_qty_map[barcode] += qty
    #                     barcode_row_map[barcode] = idx
    #
    #         # ✅ Second pass: process barcode lines
    #         for barcode, total_qty in barcode_qty_map.items():
    #             idx = barcode_row_map.get(barcode, 0)
    #             matching_lines = move_lines.filtered(lambda ml: ml.internal_ref_lot == barcode and ml.lot_name)
    #             if not matching_lines:
    #                 raise UserError(f"Row {idx}: Barcode '{barcode}' not found in move lines.")
    #
    #             branded_lines = matching_lines.filtered(lambda ml: ml.type_product == 'brand')
    #             if not branded_lines:
    #                 raise UserError(f"Row {idx}: Barcode '{barcode}' refers to unbranded product. Use serial.")
    #
    #             product_line = next((ml for ml in branded_lines if ml.product_id), False)
    #             if not product_line:
    #                 raise UserError(f"Row {idx}: Product not found for barcode '{barcode}'.")
    #
    #             tracking = product_line.product_id.tracking
    #
    #             if tracking == 'serial':
    #                 available_lines = [ml for ml in branded_lines if ml.lot_name not in found_serials]
    #                 if len(available_lines) < total_qty:
    #                     raise UserError(
    #                         f"Row {idx}: Qty = {total_qty}, but only {len(available_lines)} available for barcode '{barcode}'.")
    #
    #                 for ml in available_lines[:int(total_qty)]:
    #                     found_serials.add(ml.lot_name)
    #
    #                 verification_lines.append((0, 0, {
    #                     'stock_product_id': product_line.product_id.id,
    #                     'stock_serial': barcode,
    #                     'stock_qty': total_qty,
    #                 }))
    #
    #             elif tracking == 'lot':
    #                 available_qty = sum(ml.quantity for ml in branded_lines if ml.lot_name not in found_serials)
    #                 if available_qty < total_qty:
    #                     raise UserError(
    #                         f"Row {idx}: Qty = {total_qty}, but only {available_qty} available for barcode '{barcode}'.")
    #
    #                 for ml in branded_lines:
    #                     found_serials.add(ml.lot_name)
    #
    #                 verification_lines.append((0, 0, {
    #                     'stock_product_id': product_line.product_id.id,
    #                     'stock_serial': barcode,
    #                     'stock_qty': total_qty,
    #                 }))
    #
    #         # ✅ Push new verification lines
    #         self.stock_picking_id.stock_verification_ids = verification_lines
    #
    #         # ✅ Check final quantity match
    #         operation_qty = sum(self.stock_picking_id.move_line_ids_without_package.mapped('quantity'))
    #         verification_qty = sum(line[2]['stock_qty'] for line in verification_lines)
    #
    #     except Exception as e:
    #         _logger.error(f"Import error: {str(e)}")
    #         raise UserError(f"Error importing file: {str(e)}")
    #
    #     return {'type': 'ir.actions.act_window_close'}


class StockPickingBarcode(models.Model):
    _name = 'stock.picking.barcode'
    _description = "stock picking barcode"

    @api.model_create_multi
    def create(self, vals_list):
        filtered_vals_list = []
        for vals in vals_list:
            picking_id = vals.get('stock_picking_delivery_id')
            picking = self.env['stock.picking'].browse(picking_id) if picking_id else False
            serial_no = vals.get('serial_no')

            if picking and picking.picking_type_code == 'outgoing':
                final_barcode = f"{picking.lr_number}-{serial_no}" if serial_no else False
                barcode = f"{picking.name}-{serial_no}" if serial_no else False

                if barcode:
                    vals['barcode'] = barcode

                if vals.get('sequence', 'New') == 'New' and final_barcode:
                    vals['sequence'] = final_barcode
                if not vals.get('barcode'):
                    continue

            filtered_vals_list.append(vals)
        if not filtered_vals_list:
            return self.browse()
        return super(StockPickingBarcode, self).create(filtered_vals_list)

    barcode = fields.Char(string='Barcode')
    stock_picking_delivery_id = fields.Many2one('stock.picking', string="Delivery Number")

    lr_number = fields.Char(
        string="LR Number",
        related='stock_picking_delivery_id.lr_number',
        store=False,
        readonly=True
    )
    sequence = fields.Char(string="Sequence", copy=False, default=lambda self: _("New"))
    serial_no = fields.Integer(string='S.NO')
    status = fields.Selection([
        ('pending', 'Pending'),
        ('matched', 'Matched')
    ], string="Status", default='pending')


class NhclStockMoveLine(models.Model):
    _name = "nhcl.stock.move.line"
    _description = "nhcl stock move line"

    product_id = fields.Many2one('product.product', 'Product', ondelete="cascade", check_company=True,
                                 domain="[('type', '!=', 'service')]", index=True)
    lot_name = fields.Char('Lot/Serial Number Name')
    quantity = fields.Float(
        'Quantity', digits='Product Unit of Measure', copy=False, store=True,
        readonly=False)
    internal_ref_lot = fields.Char(string="Barcode", copy=False)

    type_product = fields.Selection([('brand', 'Brand'), ('un_brand', 'UnBrand'), ('others', 'Others')],
                                    string='Brand Type', copy=False)

    categ_1 = fields.Many2one('product.attribute.value', string='Color', copy=False,
                              domain=[('attribute_id.name', '=', 'Color')])
    categ_2 = fields.Many2one('product.attribute.value', string='Fit', copy=False,
                              domain=[('attribute_id.name', '=', 'Fit')])
    categ_3 = fields.Many2one('product.attribute.value', string='Brand', copy=False,
                              domain=[('attribute_id.name', '=', 'Brand')])
    categ_4 = fields.Many2one('product.attribute.value', string='Pattern', copy=False,
                              domain=[('attribute_id.name', '=', 'Pattern')])
    categ_5 = fields.Many2one('product.attribute.value', string='Border Type', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Type')])
    categ_6 = fields.Many2one('product.attribute.value', string='Border Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Size')])
    categ_7 = fields.Many2one('product.attribute.value', string='Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Size')])
    categ_8 = fields.Many2one('product.attribute.value', string='Design', copy=False,
                              domain=[('attribute_id.name', '=', 'Design')])
    descrip_1 = fields.Many2one('product.aging.line', string="Product Aging", copy=False)
    descrip_2 = fields.Many2one('product.attribute.value', string='Range', copy=False,
                                domain=[('attribute_id.name', '=', 'Range')])
    descrip_3 = fields.Many2one('product.attribute.value', string='Collection', copy=False,
                                domain=[('attribute_id.name', '=', 'Collection')])
    descrip_4 = fields.Many2one('product.attribute.value', string='Fabric', copy=False,
                                domain=[('attribute_id.name', '=', 'Fabric')])
    descrip_5 = fields.Many2one('product.attribute.value', string='Exclusive', copy=False,
                                domain=[('attribute_id.name', '=', 'Exclusive')])
    descrip_6 = fields.Many2one('product.attribute.value', string='Print', copy=False,
                                domain=[('attribute_id.name', '=', 'Print')])
    descrip_7 = fields.Many2one('product.attribute.value', string='Days Ageing', copy=False,
                                domain=[('attribute_id.name', '=', 'Days Ageing')])
    descrip_8 = fields.Many2one('product.attribute.value', string='Description 8', copy=False)
    cost_price = fields.Float(string='CP', copy=False)
    mr_price = fields.Float(string='MRP', copy=False)
    rs_price = fields.Float(string='RSP', copy=False)
    segment = fields.Selection([('apparel', 'Apparel'), ('non_apparel', 'Non Apparel'), ('others', 'Others')],
                               string="Segment", copy=False, store=True)
    # move_line_picking_type = fields.Selection(related='picking_id.stock_picking_type', string='Picking Type')


class StockReturnPicking(models.TransientModel):
    _inherit = 'stock.return.picking'

    def _create_returns(self):
        new_picking_id, picking_type_id = super()._create_returns()

        new_picking = self.env['stock.picking'].browse(new_picking_id)
        original_picking = self.picking_id

        # Copy transfer type
        new_picking.transfer_type = original_picking.transfer_type

        # Copy hired products
        for line in original_picking.hired_product_ids:
            self.env['sale.order.hired.product'].create({
                'picking_id': new_picking.id,
                'product_id': line.product_id.id,
                'lot_number': line.lot_number.id,
                'barcode': line.barcode,
                'quantity': line.quantity,
            })

        return new_picking_id, picking_type_id
